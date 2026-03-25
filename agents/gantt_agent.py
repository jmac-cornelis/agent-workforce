##########################################################################################
#
# Module: agents/gantt_agent.py
#
# Description: Gantt Project Planner Agent.
#              Produces evidence-backed planning snapshots, milestone proposals,
#              dependency views, and planning-risk summaries from Jira data.
#
# Author: Cornelis Networks
#
##########################################################################################

from __future__ import annotations

import logging
import os
import sys
import time
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Mapping, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from agents.base import BaseAgent, AgentConfig, AgentResponse
from agents.gantt_components import (
    BacklogInterpreter,
    DependencyMapper,
    MilestonePlanner,
    PlanningSummarizer,
    RiskProjector,
)
from agents.gantt_models import (
    BugSummary,
    DependencyEdge,
    DependencyGraph,
    MilestoneProposal,
    PlanningRequest,
    PlanningRiskRecord,
    PlanningSnapshot,
    ReleaseMonitorReport,
    ReleaseMonitorRequest,
    RoadmapGap,
    RoadmapItem,
    RoadmapRequest,
    RoadmapSection,
    RoadmapSnapshot,
)
from agents.pm_runtime import normalize_csv_list, notify_shannon
from collections import Counter
from core.evidence import EvidenceBundle, load_evidence_bundle
from core.release_tracking import (
    build_snapshot as build_release_snapshot,
    compute_delta,
    compute_velocity,
    compute_cycle_time_stats,
    assess_readiness,
    format_summary as format_release_summary,
    TrackerConfig,
)
from state.gantt_dependency_review_store import GanttDependencyReviewStore
from state.gantt_release_monitor_store import GanttReleaseMonitorStore
from state.gantt_snapshot_store import GanttSnapshotStore
from excel_utils import (
    STATUS_FILL_COLORS,
    PRIORITY_FILL_COLORS,
    _apply_header_style,
    _apply_status_conditional_formatting,
    _apply_priority_conditional_formatting,
    _auto_fit_columns,
)
from tools.jira_tools import (
    JiraTools, get_jira, get_project_info, get_releases,
    search_tickets, get_children_hierarchy,
)
from tools.knowledge_tools import search_knowledge, list_knowledge_files, read_knowledge_file

# Logging config - follows jira_utils.py pattern
log = logging.getLogger(os.path.basename(sys.argv[0]))

# ---------------------------------------------------------------------------
# Roadmap constants (from roadmap_agent.py)
# ---------------------------------------------------------------------------

# Jira base URL for hyperlinks
JIRA_BASE_URL = 'https://cornelisnetworks.atlassian.net'

# Issue types to exclude from the roadmap (Bugs are not roadmap items)
_EXCLUDED_TYPES = {'Bug', 'bug'}

# Depth mapping for issue types
_TYPE_DEPTH = {
    'Initiative': 0,
    'Epic': 1,
    'Story': 2,
    'Task': 2,
    'Sub-task': 2,
}

# Done statuses — used when include_closed is False
_DONE_STATUSES = {'Closed', 'Done', 'Resolved'}


class GanttProjectPlannerAgent(BaseAgent):
    '''
    Agent for producing project-planning snapshots from Jira data.

    The implementation is deterministic-first. Specialized planner components
    handle backlog interpretation, dependency mapping, milestone planning,
    risk projection, and summary formatting.
    '''

    STALE_DAYS = 30

    def __init__(self, project_key: Optional[str] = None, **kwargs):
        '''
        Initialize the Gantt Project Planner agent.
        '''
        instruction = self._load_prompt_file()
        if not instruction:
            raise FileNotFoundError(
                'config/prompts/gantt_agent.md is required but not found. '
                'The Gantt Project Planner Agent has no hardcoded fallback prompt.'
            )

        config = AgentConfig(
            name='gantt_project_planner',
            description='Builds planning snapshots, milestones, dependencies, and risks',
            instruction=instruction,
            max_iterations=10,
        )

        super().__init__(config=config, tools=[JiraTools()], **kwargs)

        self.register_tool(search_knowledge)
        self.register_tool(list_knowledge_files)
        self.register_tool(read_knowledge_file)

        self.project_key = project_key
        self.backlog_interpreter = BacklogInterpreter(
            jira_provider=get_jira,
            now_provider=self._utc_now,
            stale_days=self.STALE_DAYS,
        )
        self.dependency_mapper = DependencyMapper(
            review_store=GanttDependencyReviewStore()
        )
        self.milestone_planner = MilestonePlanner()
        self.risk_projector = RiskProjector()
        self.planning_summarizer = PlanningSummarizer()

    @staticmethod
    def _load_prompt_file() -> Optional[str]:
        '''Load the Gantt agent prompt from config/prompts/.'''
        prompt_path = os.path.join('config', 'prompts', 'gantt_agent.md')
        if os.path.exists(prompt_path):
            try:
                with open(prompt_path, 'r', encoding='utf-8') as f:
                    return f.read()
            except Exception as e:
                log.warning(f'Failed to load Gantt agent prompt: {e}')
        return None

    @staticmethod
    def _utc_now() -> datetime:
        '''Return the current UTC time as a timezone-aware datetime.'''
        return datetime.now(timezone.utc)

    def run(self, input_data: Any) -> AgentResponse:
        '''
        Build a planning snapshot from Jira data.
        '''
        log.debug(f'GanttProjectPlannerAgent.run(input_data={input_data})')

        if isinstance(input_data, str):
            request = PlanningRequest(project_key=input_data)
        elif isinstance(input_data, dict):
            request = PlanningRequest(
                project_key=input_data.get('project_key', self.project_key or ''),
                planning_horizon_days=int(input_data.get('planning_horizon_days', 90)),
                limit=int(input_data.get('limit', 200)),
                include_done=bool(input_data.get('include_done', False)),
                backlog_jql=input_data.get('backlog_jql'),
                policy_profile=input_data.get('policy_profile', 'default'),
                evidence_paths=[
                    str(path) for path in (input_data.get('evidence_paths') or [])
                ],
            )
        else:
            return AgentResponse.error_response(
                'Invalid input: expected project key string or request dict'
            )

        if not request.project_key:
            return AgentResponse.error_response('No project_key provided')

        try:
            snapshot = self.create_snapshot(request)
        except Exception as e:
            log.error(f'Gantt planning snapshot failed: {e}')
            return AgentResponse.error_response(str(e))

        return AgentResponse.success_response(
            content=snapshot.summary_markdown,
            metadata={'planning_snapshot': snapshot.to_dict()},
        )

    def run_once(
        self,
        task_type: str,
        request: PlanningRequest | ReleaseMonitorRequest,
        *,
        persist: bool = True,
    ) -> Dict[str, Any]:
        '''
        Execute one deterministic Gantt task and optionally persist the result.
        '''
        if task_type == 'planning_snapshot':
            if not isinstance(request, PlanningRequest):
                raise TypeError('planning_snapshot requires a PlanningRequest')

            self.project_key = request.project_key or self.project_key
            snapshot = self.create_snapshot(request)
            result: Dict[str, Any] = {
                'ok': True,
                'task_type': task_type,
                'project_key': snapshot.project_key,
                'snapshot': snapshot.to_dict(),
                'summary_markdown': snapshot.summary_markdown,
            }
            if persist:
                result['stored'] = GanttSnapshotStore().save_snapshot(
                    snapshot,
                    summary_markdown=snapshot.summary_markdown,
                )
            return result

        if task_type == 'release_monitor':
            if not isinstance(request, ReleaseMonitorRequest):
                raise TypeError('release_monitor requires a ReleaseMonitorRequest')

            self.project_key = request.project_key or self.project_key
            report = self.create_release_monitor(request)
            result = {
                'ok': True,
                'task_type': task_type,
                'project_key': report.project_key,
                'report': report.to_dict(),
                'summary_markdown': report.summary_markdown,
            }
            if persist:
                result['stored'] = GanttReleaseMonitorStore().save_report(
                    report,
                    summary_markdown=report.summary_markdown,
                )
            return result

        raise ValueError(f'Unsupported Gantt task_type: {task_type}')

    @staticmethod
    def _build_snapshot_notification_payload(result: Dict[str, Any]) -> Dict[str, Any]:
        snapshot = result.get('snapshot', {})
        overview = snapshot.get('backlog_overview', {})
        stored = result.get('stored', {})
        text = (
            f'{snapshot.get("project_key", "")}: '
            f'{overview.get("total_issues", 0)} issues, '
            f'{overview.get("blocked_issues", 0)} blocked, '
            f'{len(snapshot.get("milestones", []))} milestones'
        )
        body_lines = [
            f'Snapshot ID: {stored.get("snapshot_id") or snapshot.get("snapshot_id", "")}',
            f'Planning horizon: {snapshot.get("planning_horizon_days", 0)} days',
            f'Risks: {len(snapshot.get("risks", []))}',
            f'Dependencies: {snapshot.get("dependency_graph", {}).get("edge_count", 0)}',
        ]
        return {
            'title': f'Gantt Planning Snapshot — {snapshot.get("project_key", "")}',
            'text': text,
            'body_lines': body_lines,
        }

    @staticmethod
    def _build_release_monitor_notification_payload(result: Dict[str, Any]) -> Dict[str, Any]:
        report = result.get('report', {})
        stored = result.get('stored', {})
        text = (
            f'{report.get("project_key", "")}: '
            f'{report.get("total_bugs", 0)} bugs, '
            f'P0={report.get("total_p0", 0)}, '
            f'P1={report.get("total_p1", 0)}'
        )
        body_lines = [
            f'Report ID: {stored.get("report_id") or report.get("report_id", "")}',
            f'Releases: {", ".join(report.get("releases_monitored", [])) or "none"}',
        ]
        readiness = report.get('readiness') or {}
        if readiness:
            body_lines.append(
                f'Readiness: {readiness.get("total_open", 0)} open, '
                f'{readiness.get("p0_open", 0)} P0, '
                f'{readiness.get("p1_open", 0)} P1'
            )
        return {
            'title': f'Gantt Release Monitor — {report.get("project_key", "")}',
            'text': text,
            'body_lines': body_lines,
        }

    def tick(self, poller_spec: Dict[str, Any]) -> Dict[str, Any]:
        '''
        Execute one scheduled Gantt cycle.
        '''
        project_key = str(
            poller_spec.get('project_key') or self.project_key or ''
        ).strip()
        if not project_key:
            raise ValueError('Gantt tick requires project_key')

        self.project_key = project_key
        persist = bool(poller_spec.get('persist', True))
        notify_enabled = bool(poller_spec.get('notify_shannon', False))
        shannon_base_url = poller_spec.get('shannon_base_url')
        tasks: List[Dict[str, Any]] = []
        notifications: List[Dict[str, Any]] = []
        errors: List[str] = []

        if bool(poller_spec.get('run_planning', True)):
            try:
                planning_request = PlanningRequest(
                    project_key=project_key,
                    planning_horizon_days=int(
                        poller_spec.get('planning_horizon_days', 90)
                    ),
                    limit=int(poller_spec.get('limit', 200)),
                    include_done=bool(poller_spec.get('include_done', False)),
                    backlog_jql=poller_spec.get('backlog_jql'),
                    policy_profile=str(
                        poller_spec.get('policy_profile', 'default') or 'default'
                    ),
                    evidence_paths=normalize_csv_list(
                        poller_spec.get('evidence_paths')
                    ),
                )
                planning_result = self.run_once(
                    'planning_snapshot',
                    planning_request,
                    persist=persist,
                )
                tasks.append(planning_result)
                if notify_enabled:
                    notifications.append(
                        notify_shannon(
                            agent_id='gantt',
                            shannon_base_url=shannon_base_url,
                            **self._build_snapshot_notification_payload(planning_result),
                        )
                    )
            except Exception as e:
                errors.append(f'planning_snapshot: {e}')

        release_names = normalize_csv_list(poller_spec.get('releases'))
        run_release_monitor = bool(poller_spec.get('run_release_monitor', False))
        if release_names:
            run_release_monitor = True

        if run_release_monitor:
            try:
                release_request = ReleaseMonitorRequest(
                    project_key=project_key,
                    releases=release_names or None,
                    scope_label=poller_spec.get('scope_label'),
                    include_gap_analysis=bool(
                        poller_spec.get('include_gap_analysis', True)
                    ),
                    include_bug_report=bool(
                        poller_spec.get('include_bug_report', True)
                    ),
                    include_velocity=bool(
                        poller_spec.get('include_velocity', True)
                    ),
                    include_readiness=bool(
                        poller_spec.get('include_readiness', True)
                    ),
                    compare_to_previous=bool(
                        poller_spec.get('compare_to_previous', True)
                    ),
                    output_file=poller_spec.get('output_file'),
                )
                release_result = self.run_once(
                    'release_monitor',
                    release_request,
                    persist=persist,
                )
                tasks.append(release_result)
                if notify_enabled:
                    notifications.append(
                        notify_shannon(
                            agent_id='gantt',
                            shannon_base_url=shannon_base_url,
                            **self._build_release_monitor_notification_payload(
                                release_result
                            ),
                        )
                    )
            except Exception as e:
                errors.append(f'release_monitor: {e}')

        return {
            'ok': not errors,
            'agent_id': 'gantt',
            'project_key': project_key,
            'tasks': tasks,
            'notifications': notifications,
            'errors': errors,
            'completed_at': datetime.now(timezone.utc).isoformat(),
        }

    def run_poller(
        self,
        poller_spec: Dict[str, Any],
        *,
        sleep_fn: Any = time.sleep,
    ) -> Dict[str, Any]:
        '''
        Run the Gantt polling loop for a bounded number of cycles.
        '''
        interval_seconds = max(int(poller_spec.get('interval_seconds', 300) or 0), 0)
        max_cycles = int(poller_spec.get('max_cycles', 1) or 0)
        if max_cycles < 0:
            raise ValueError('max_cycles must be >= 0')
        if max_cycles == 0 and interval_seconds <= 0:
            raise ValueError(
                'Continuous polling requires interval_seconds > 0'
            )

        cycle_summaries: List[Dict[str, Any]] = []
        cycles_run = 0
        last_tick: Optional[Dict[str, Any]] = None

        while True:
            cycles_run += 1
            last_tick = self.tick(poller_spec)
            cycle_summaries.append({
                'cycle_number': cycles_run,
                'ok': last_tick.get('ok', False),
                'task_count': len(last_tick.get('tasks', [])),
                'notification_count': len(last_tick.get('notifications', [])),
                'errors': list(last_tick.get('errors', [])),
                'completed_at': last_tick.get('completed_at', ''),
            })

            if max_cycles > 0 and cycles_run >= max_cycles:
                break
            if interval_seconds > 0:
                sleep_fn(interval_seconds)

        return {
            'ok': all(item.get('ok', False) for item in cycle_summaries),
            'agent_id': 'gantt',
            'project_key': str(
                poller_spec.get('project_key') or self.project_key or ''
            ).strip(),
            'cycles_run': cycles_run,
            'cycle_summaries': cycle_summaries,
            'last_tick': last_tick or {},
        }

    def create_snapshot(self, request: PlanningRequest) -> PlanningSnapshot:
        '''
        Create a deterministic planning snapshot from Jira backlog data.
        '''
        log.info(f'Creating Gantt planning snapshot for {request.project_key}')

        project_info = self._load_project_info(request.project_key)
        releases = self._load_releases(request.project_key)
        issues = self._load_backlog_issues(request)
        evidence_bundle = load_evidence_bundle(request.evidence_paths)
        dependency_graph = self._build_dependency_graph(issues)
        milestones = self._build_milestones(issues, releases, dependency_graph)
        evidence_gaps = self._build_evidence_gaps(issues, evidence_bundle)
        risks = self._build_risks(issues, milestones, dependency_graph)
        backlog_overview = self._build_backlog_overview(
            issues,
            milestones,
            dependency_graph,
            risks,
        )

        snapshot = PlanningSnapshot(
            project_key=request.project_key,
            planning_horizon_days=request.planning_horizon_days,
            project_info=project_info,
            backlog_overview=backlog_overview,
            milestones=milestones,
            dependency_graph=dependency_graph,
            risks=risks,
            issues=issues,
            evidence_summary=evidence_bundle.to_summary(),
            evidence_gaps=evidence_gaps,
        )
        snapshot.summary_markdown = self._format_snapshot(snapshot)
        return snapshot

    def _load_project_info(self, project_key: str) -> Dict[str, Any]:
        result = get_project_info(project_key)
        if result.is_success:
            return result.data
        raise RuntimeError(result.error or f'Failed to load project info for {project_key}')

    def _load_releases(self, project_key: str) -> List[Dict[str, Any]]:
        result = get_releases(
            project_key,
            include_released=True,
            include_unreleased=True,
        )
        if result.is_success:
            return result.data
        raise RuntimeError(result.error or f'Failed to load releases for {project_key}')

    def _load_backlog_issues(self, request: PlanningRequest) -> List[Dict[str, Any]]:
        '''
        Query Jira directly so dependency-related fields remain available.
        '''
        issues = self.backlog_interpreter.load_backlog_issues(request)
        return self.dependency_mapper.attach_dependency_edges(
            issues,
            project_key=request.project_key,
        )

    @staticmethod
    def _build_backlog_jql(request: PlanningRequest) -> str:
        return BacklogInterpreter.build_backlog_jql(request)

    def _normalize_issue(self, issue: Any) -> Dict[str, Any]:
        normalized = self.backlog_interpreter.normalize_issue(issue)
        return self.dependency_mapper.attach_dependency_edges(
            [normalized],
            project_key=self.project_key,
        )[0]

    def _extract_edges(
        self,
        issue_key: str,
        parent_key: str,
        issue_links: Iterable[Any],
    ) -> List[DependencyEdge]:
        return self.dependency_mapper.extract_edges(issue_key, parent_key, issue_links)

    @staticmethod
    def _slugify_relationship(value: str) -> str:
        return DependencyMapper.slugify_relationship(value)

    @staticmethod
    def _parse_jira_datetime(value: str) -> Optional[datetime]:
        return BacklogInterpreter.parse_jira_datetime(value)

    @staticmethod
    def _is_done_status(status: str) -> bool:
        return BacklogInterpreter.is_done_status(status)

    @staticmethod
    def _is_high_priority(priority: str) -> bool:
        return BacklogInterpreter.is_high_priority(priority)

    @staticmethod
    def _is_blocked_status(status: str) -> bool:
        return DependencyMapper.is_blocked_status(status)

    def _build_dependency_graph(self, issues: List[Dict[str, Any]]) -> DependencyGraph:
        return self.dependency_mapper.build_graph(issues)

    def _build_milestones(
        self,
        issues: List[Dict[str, Any]],
        releases: List[Dict[str, Any]],
        dependency_graph: DependencyGraph,
    ) -> List[MilestoneProposal]:
        return self.milestone_planner.build_milestones(
            issues,
            releases,
            dependency_graph,
        )

    def _milestone_risk_level(
        self,
        open_issues: int,
        blocked_issues: int,
        unassigned_issues: int,
        release_date: Optional[str],
    ) -> str:
        return self.milestone_planner.milestone_risk_level(
            open_issues,
            blocked_issues,
            unassigned_issues,
            release_date,
        )

    @staticmethod
    def _milestone_sort_key(milestone: MilestoneProposal) -> tuple[int, str, str]:
        return MilestonePlanner.milestone_sort_key(milestone)

    def _build_evidence_gaps(
        self,
        issues: List[Dict[str, Any]],
        evidence_bundle: Optional[EvidenceBundle] = None,
    ) -> List[str]:
        return self.risk_projector.build_evidence_gaps(issues, evidence_bundle)

    def _build_risks(
        self,
        issues: List[Dict[str, Any]],
        milestones: List[MilestoneProposal],
        dependency_graph: DependencyGraph,
    ) -> List[PlanningRiskRecord]:
        return self.risk_projector.build_risks(issues, milestones, dependency_graph)

    def _build_backlog_overview(
        self,
        issues: List[Dict[str, Any]],
        milestones: List[MilestoneProposal],
        dependency_graph: DependencyGraph,
        risks: List[PlanningRiskRecord],
    ) -> Dict[str, Any]:
        return self.planning_summarizer.build_backlog_overview(
            issues,
            milestones,
            dependency_graph,
            risks,
        )

    def _format_snapshot(self, snapshot: PlanningSnapshot) -> str:
        return self.planning_summarizer.format_snapshot(snapshot)

    # ===========================================================================
    # Roadmap Analysis — Initiative/Epic/Story hierarchy with gap analysis
    # ===========================================================================

    def create_roadmap_snapshot(self, request: RoadmapRequest) -> RoadmapSnapshot:
        '''
        Create a roadmap snapshot from Jira data with optional gap analysis.

        Algorithm:
        1. Discover initiatives (explicit keys or JQL search).
        2. Build hierarchy — children of each initiative (epics, stories).
        3. Organize into RoadmapSection objects.
        4. Optionally run LLM gap analysis.
        5. Generate xlsx output.
        6. Return RoadmapSnapshot.
        '''
        log.info(
            f'Creating roadmap snapshot for {request.project_key} '
            f'(scope={request.scope_label})'
        )

        # Step 1: Discover initiatives
        initiative_keys = self._discover_initiatives(request)
        log.info(f'Discovered {len(initiative_keys)} initiatives')

        # Step 2+3: Build hierarchy and organize into sections
        sections = self._build_sections(request, initiative_keys)
        log.info(
            f'Built {len(sections)} sections with '
            f'{sum(len(s.items) for s in sections)} total items'
        )

        # Step 4: Gap analysis (LLM)
        if request.include_gap_analysis:
            self._run_gap_analysis(request, sections)
            total_gaps = sum(len(s.gaps) for s in sections)
            log.info(f'Gap analysis identified {total_gaps} proposed gaps')

        # Build snapshot
        snapshot = RoadmapSnapshot(
            project_key=request.project_key,
            scope_label=request.scope_label,
            sections=sections,
        )
        snapshot.summary_markdown = self._format_summary(snapshot)

        # Step 5: Generate xlsx
        if request.output_file:
            output_path = self._write_xlsx(snapshot, request.output_file, request)
            log.info(f'Wrote roadmap xlsx to {output_path}')

        return snapshot

    def _discover_initiatives(self, request: RoadmapRequest) -> List[str]:
        '''
        Discover initiative ticket keys — either from explicit list or JQL.

        If request.initiative_keys is provided, use those directly.
        Otherwise, search Jira for initiatives matching the scope_label.
        '''
        if request.initiative_keys:
            return list(request.initiative_keys)

        # Build JQL to find initiatives by scope label
        jql_parts = [
            f'project = {request.project_key}',
            'issuetype = Initiative',
        ]
        if request.scope_label:
            jql_parts.append(f'summary ~ "{request.scope_label}"')
        jql = ' AND '.join(jql_parts) + ' ORDER BY priority ASC'

        log.debug(f'Initiative discovery JQL: {jql}')

        try:
            result = search_tickets(jql=jql, limit=50)
            if not result.is_success:
                log.warning(f'Initiative search failed: {result.error}')
                return []
            return [ticket['key'] for ticket in result.data]
        except Exception as e:
            log.error(f'Initiative discovery failed: {e}')
            return []

    def _build_sections(
        self,
        request: RoadmapRequest,
        initiative_keys: List[str],
    ) -> List[RoadmapSection]:
        '''
        Build RoadmapSection objects from initiative keys.

        For each initiative, fetches the child hierarchy and organizes
        items into a section. Filters out Bugs and optionally closed items.
        '''
        sections: List[RoadmapSection] = []

        for init_key in initiative_keys:
            try:
                section = self._build_section_for_initiative(request, init_key)
                if section:
                    sections.append(section)
            except Exception as e:
                log.error(f'Failed to build section for {init_key}: {e}')

        return sections

    def _build_section_for_initiative(
        self,
        request: RoadmapRequest,
        initiative_key: str,
    ) -> Optional[RoadmapSection]:
        '''
        Build a single RoadmapSection for an initiative and its children.

        Uses get_children_hierarchy to fetch the full tree, then converts
        each Jira issue into a RoadmapItem.
        '''
        # Fetch the full hierarchy rooted at this initiative
        try:
            result = get_children_hierarchy(root_key=initiative_key, limit=500)
            if not result.is_success:
                log.warning(
                    f'get_children_hierarchy failed for {initiative_key}: '
                    f'{result.error}'
                )
                return None
        except Exception as e:
            log.error(f'Hierarchy fetch failed for {initiative_key}: {e}')
            return None

        hierarchy_items = result.data
        if not hierarchy_items:
            return None

        # The first item is the initiative itself — use its summary as title
        root = hierarchy_items[0]
        section_title = root.get('summary', initiative_key)

        items: List[RoadmapItem] = []
        for ticket in hierarchy_items:
            item = self._ticket_to_roadmap_item(ticket, section_title)
            if item is None:
                continue

            # Filter: exclude Bugs
            if item.issue_type in _EXCLUDED_TYPES:
                continue

            # Filter: exclude closed items unless requested
            if not request.include_closed and item.status in _DONE_STATUSES:
                continue

            # Filter: fix_versions if specified
            if request.fix_versions and item.fix_version:
                if not any(
                    fv in item.fix_version for fv in request.fix_versions
                ):
                    continue

            # Filter: respect hierarchy_depth
            if item.depth > request.hierarchy_depth:
                continue

            items.append(item)

        if not items:
            return None

        return RoadmapSection(title=section_title, items=items)

    def _ticket_to_roadmap_item(
        self,
        ticket: Dict[str, Any],
        section_title: str,
    ) -> Optional[RoadmapItem]:
        '''
        Convert a Jira ticket dict (from get_children_hierarchy or
        search_tickets) into a RoadmapItem.

        get_children_hierarchy returns: key, summary, type, status,
        priority, assignee, fix_versions (list), components (list),
        labels (list), depth.

        search_tickets returns: key, summary, issue_type/type, status,
        priority, assignee, fix_version (csv), component (csv),
        labels (list), etc.
        '''
        key = ticket.get('key', '')
        if not key:
            return None

        # Determine issue type — get_children_hierarchy uses 'type',
        # search_tickets uses 'issue_type' or 'type'
        issue_type = (
            ticket.get('issue_type')
            or ticket.get('type')
            or ''
        )

        # Determine depth — get_children_hierarchy provides it directly,
        # otherwise infer from issue type
        depth = ticket.get('depth')
        if depth is None:
            depth = _TYPE_DEPTH.get(issue_type, 2)

        # Handle fix_versions — may be a list or a csv string
        fix_versions_raw = ticket.get('fix_versions') or ticket.get('fix_version', '')
        if isinstance(fix_versions_raw, list):
            fix_version = ', '.join(fix_versions_raw)
        else:
            fix_version = str(fix_versions_raw)

        # Handle components — may be a list or a csv string
        components_raw = ticket.get('components') or ticket.get('component', '')
        if isinstance(components_raw, list):
            component = ', '.join(components_raw)
        else:
            component = str(components_raw)

        # Handle labels — may be a list or a csv string
        labels_raw = ticket.get('labels') or ticket.get('labels_csv', '')
        if isinstance(labels_raw, list):
            labels = ', '.join(labels_raw)
        else:
            labels = str(labels_raw)

        # Parent key — from search_tickets result or hierarchy context
        parent_key = ticket.get('parent_key', '')

        return RoadmapItem(
            key=key,
            summary=ticket.get('summary', ''),
            issue_type=issue_type,
            status=ticket.get('status', ''),
            priority=ticket.get('priority', ''),
            assignee=ticket.get('assignee', ''),
            fix_version=fix_version,
            component=component,
            labels=labels,
            parent_key=parent_key,
            depth=depth,
            source='Jira',
            section=section_title,
        )

    def _run_gap_analysis(
        self,
        request: RoadmapRequest,
        sections: List[RoadmapSection],
    ) -> None:
        '''
        Run LLM gap analysis on the current roadmap sections.

        Builds a text summary of all sections and items, sends to the LLM
        via _run_with_tools(), parses the JSON response, and assigns
        RoadmapGap objects to the appropriate sections.
        '''
        # Build the roadmap summary for the LLM
        roadmap_text = self._build_gap_analysis_prompt(request, sections)

        log.info('Running LLM gap analysis...')
        try:
            response = self._run_with_tools(roadmap_text)
            if not response.success:
                log.warning(f'Gap analysis LLM call failed: {response.error}')
                return

            # Parse the JSON response
            gaps_json = self._extract_json_block(response.content or '')
            if not gaps_json:
                log.warning('No JSON block found in gap analysis response')
                return

            # Convert to RoadmapGap objects and assign to sections
            self._assign_gaps_to_sections(gaps_json, sections)

        except Exception as e:
            log.error(f'Gap analysis failed: {e}')

    def _build_gap_analysis_prompt(
        self,
        request: RoadmapRequest,
        sections: List[RoadmapSection],
    ) -> str:
        '''Build the user prompt for LLM gap analysis.'''
        lines = [
            f'## Roadmap Gap Analysis Request',
            f'',
            f'**Project**: {request.project_key}',
            f'**Scope**: {request.scope_label or "All"}',
            f'',
            f'Analyze the following Jira roadmap and identify gaps — '
            f'missing Epics or Stories that should exist but do not yet '
            f'have Jira tickets. Return your analysis as a JSON block.',
            f'',
            f'### Current Roadmap',
            f'',
        ]

        for section in sections:
            lines.append(f'#### {section.title}')
            for item in section.items:
                indent = '  ' * item.depth
                status_tag = f' [{item.status}]' if item.status else ''
                lines.append(
                    f'{indent}- {item.key} ({item.issue_type}): '
                    f'{item.summary}{status_tag}'
                )
            lines.append('')

        lines.extend([
            '### Output Format',
            '',
            'Return a JSON block with this structure:',
            '```json',
            '{',
            '  "gaps": [',
            '    {',
            '      "section": "Section Title",',
            '      "summary": "Proposed ticket summary",',
            '      "issue_type": "Epic or Story",',
            '      "priority": "P0-Stopper | P1-Critical | P2-Major | P3-Minor",',
            '      "suggested_component": "component name",',
            '      "acceptance_criteria": "AC text",',
            '      "dependencies": "STL-123; STL-456",',
            '      "suggested_fix_version": "version string",',
            '      "labels": "label1, label2",',
            '      "depth": 1,',
            '      "parent_summary": "Parent epic summary (for stories)"',
            '    }',
            '  ]',
            '}',
            '```',
        ])

        return '\n'.join(lines)

    def _assign_gaps_to_sections(
        self,
        gaps_json: Dict[str, Any],
        sections: List[RoadmapSection],
    ) -> None:
        '''
        Convert parsed gap JSON into RoadmapGap objects and assign
        them to the matching section.
        '''
        # Build a section lookup by title
        section_map: Dict[str, RoadmapSection] = {
            s.title: s for s in sections
        }

        gaps_list = gaps_json.get('gaps', [])
        if isinstance(gaps_json, list):
            # LLM may have returned a bare list instead of {"gaps": [...]}
            gaps_list = gaps_json

        for gap_data in gaps_list:
            if not isinstance(gap_data, dict):
                continue

            gap = RoadmapGap(
                summary=gap_data.get('summary', ''),
                issue_type=gap_data.get('issue_type', 'Epic'),
                priority=gap_data.get('priority', 'P2-Major'),
                suggested_component=gap_data.get('suggested_component', ''),
                acceptance_criteria=gap_data.get('acceptance_criteria', ''),
                dependencies=gap_data.get('dependencies', ''),
                suggested_fix_version=gap_data.get('suggested_fix_version', ''),
                labels=gap_data.get('labels', ''),
                depth=int(gap_data.get('depth', 1)),
                section=gap_data.get('section', ''),
                parent_summary=gap_data.get('parent_summary', ''),
            )

            # Find the matching section — fall back to first section
            target_section = section_map.get(gap.section)
            if not target_section and sections:
                target_section = sections[0]
            if target_section:
                target_section.gaps.append(gap)

    # -- XLSX generation — 3-sheet workbook ------------------------------------

    @staticmethod
    def _col(headers: List[str], name: str) -> int:
        '''Return 1-based column index for *name* in *headers*.'''
        return headers.index(name) + 1

    def _write_xlsx(
        self,
        snapshot: RoadmapSnapshot,
        output_file: str,
        request: Optional[RoadmapRequest] = None,
    ) -> str:
        '''
        Generate a 3-sheet xlsx workbook from the roadmap snapshot.

        Sheet 1: "Current Jira" — all Jira-sourced items
        Sheet 2: "Proposed (Gaps)" — LLM-identified gaps
        Sheet 3: "Merged Roadmap" — both interleaved

        Uses excel_utils formatting helpers for consistent styling.
        '''
        if request is None:
            request = RoadmapRequest()

        wb = Workbook()

        # -- Sheet 1: Current Jira --
        ws_jira = wb.active or wb.create_sheet()
        ws_jira.title = 'Current Jira'
        self._write_jira_sheet(ws_jira, snapshot, request)

        # -- Sheet 2: Proposed (Gaps) --
        ws_gaps = wb.create_sheet('Proposed (Gaps)')
        self._write_gaps_sheet(ws_gaps, snapshot, request)

        # -- Sheet 3: Merged Roadmap --
        ws_merged = wb.create_sheet('Merged Roadmap')
        self._write_merged_sheet(ws_merged, snapshot, request)

        # Save workbook
        os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
        wb.save(output_file)
        return output_file

    def _write_jira_sheet(
        self,
        ws: Any,
        snapshot: RoadmapSnapshot,
        request: RoadmapRequest,
    ) -> None:
        '''Write the "Current Jira" sheet with all Jira-sourced items.'''
        headers = [
            'Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)',
            'project', 'issue_type', 'status', 'priority', 'key',
            'summary', 'assignee', 'fix_version', 'component', 'labels',
        ]
        if not request.show_priority:
            headers.remove('priority')

        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        row = 2
        for section in snapshot.sections:
            row = self._write_section_divider(ws, row, section.title, len(headers))

            for item in section.items:
                if item.source != 'Jira':
                    continue
                self._write_jira_item_row(
                    ws, row, item, snapshot.project_key, headers, request,
                )
                row += 1

        _apply_header_style(ws, len(headers))
        _apply_status_conditional_formatting(ws, headers)
        if request.show_priority:
            _apply_priority_conditional_formatting(ws, headers)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(row - 1, 1)}'
        self._set_column_widths(ws, headers)

    def _write_jira_item_row(
        self,
        ws: Any,
        row: int,
        item: RoadmapItem,
        project_key: str,
        headers: List[str],
        request: RoadmapRequest,
    ) -> None:
        '''Write a single Jira item row to the Current Jira sheet.'''
        c = self._col
        for d in range(3):
            depth_header = ['Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)'][d]
            if item.depth == d:
                ws.cell(row=row, column=c(headers, depth_header), value=item.key)
            else:
                ws.cell(row=row, column=c(headers, depth_header), value='')

        ws.cell(row=row, column=c(headers, 'project'), value=project_key)

        issue_type_cell = ws.cell(
            row=row, column=c(headers, 'issue_type'), value=item.issue_type,
        )
        if request.bold_stories and item.issue_type == 'Story':
            issue_type_cell.font = Font(bold=True)

        ws.cell(row=row, column=c(headers, 'status'), value=item.status)

        if 'priority' in headers:
            ws.cell(row=row, column=c(headers, 'priority'), value=item.priority)

        key_cell = ws.cell(row=row, column=c(headers, 'key'), value=item.key)
        if item.key:
            key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{item.key}'
            key_cell.font = Font(color='0563C1', underline='single')

        ws.cell(row=row, column=c(headers, 'summary'), value=item.summary)

        assignee_val = item.assignee
        if request.blank_unassigned and assignee_val in ('Unassigned', '', None):
            assignee_val = ''
        ws.cell(row=row, column=c(headers, 'assignee'), value=assignee_val)

        ws.cell(row=row, column=c(headers, 'fix_version'), value=item.fix_version)
        ws.cell(row=row, column=c(headers, 'component'), value=item.component)
        ws.cell(row=row, column=c(headers, 'labels'), value=item.labels)

        if item.depth == 0:
            bold_font = Font(bold=True)
            key_col = c(headers, 'key')
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col_idx)
                if col_idx == key_col and item.key:
                    cell.font = Font(
                        bold=True, color='0563C1', underline='single'
                    )
                else:
                    cell.font = bold_font

    def _write_gaps_sheet(
        self,
        ws: Any,
        snapshot: RoadmapSnapshot,
        request: RoadmapRequest,
    ) -> None:
        '''Write the "Proposed (Gaps)" sheet with LLM-identified gaps.'''
        headers = [
            'Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)',
            'project', 'issue_type', 'priority', 'summary',
            'suggested_component', 'acceptance_criteria', 'dependencies',
            'suggested_fix_version', 'labels',
        ]
        if not request.show_priority:
            headers.remove('priority')

        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        row = 2
        for section in snapshot.sections:
            if not section.gaps:
                continue

            row = self._write_section_divider(ws, row, section.title, len(headers))

            for gap in section.gaps:
                self._write_gap_row(
                    ws, row, gap, snapshot.project_key, headers, request,
                )
                row += 1

        _apply_header_style(ws, len(headers))
        if request.show_priority:
            _apply_priority_conditional_formatting(ws, headers)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(row - 1, 1)}'
        self._set_column_widths(ws, headers, is_gaps_sheet=True)

    def _write_gap_row(
        self,
        ws: Any,
        row: int,
        gap: RoadmapGap,
        project_key: str,
        headers: List[str],
        request: RoadmapRequest,
    ) -> None:
        '''Write a single gap row to the Proposed (Gaps) sheet.'''
        c = self._col
        for d in range(3):
            depth_header = ['Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)'][d]
            if gap.depth == d:
                ws.cell(row=row, column=c(headers, depth_header), value=gap.summary[:40])
            else:
                ws.cell(row=row, column=c(headers, depth_header), value='')

        ws.cell(row=row, column=c(headers, 'project'), value=project_key)

        issue_type_cell = ws.cell(
            row=row, column=c(headers, 'issue_type'), value=gap.issue_type,
        )
        if request.bold_stories and gap.issue_type == 'Story':
            issue_type_cell.font = Font(bold=True)

        if 'priority' in headers:
            ws.cell(row=row, column=c(headers, 'priority'), value=gap.priority)

        ws.cell(row=row, column=c(headers, 'summary'), value=gap.summary)
        ws.cell(row=row, column=c(headers, 'suggested_component'), value=gap.suggested_component)
        ws.cell(row=row, column=c(headers, 'acceptance_criteria'), value=gap.acceptance_criteria)
        ws.cell(row=row, column=c(headers, 'dependencies'), value=gap.dependencies)
        ws.cell(row=row, column=c(headers, 'suggested_fix_version'), value=gap.suggested_fix_version)
        ws.cell(row=row, column=c(headers, 'labels'), value=gap.labels)

    def _write_merged_sheet(
        self,
        ws: Any,
        snapshot: RoadmapSnapshot,
        request: RoadmapRequest,
    ) -> None:
        '''Write the "Merged Roadmap" sheet with both Jira and proposed items.'''
        headers = [
            'Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)',
            'project', 'issue_type', 'status', 'priority', 'key',
            'summary', 'assignee', 'fix_version', 'component', 'labels',
            'source', 'acceptance_criteria', 'dependencies',
        ]
        if not request.show_priority:
            headers.remove('priority')

        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        proposed_fill = PatternFill(
            start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'
        )

        row = 2
        for section in snapshot.sections:
            row = self._write_section_divider(ws, row, section.title, len(headers))

            for item in section.items:
                if item.source != 'Jira':
                    continue
                self._write_merged_jira_row(
                    ws, row, item, snapshot.project_key, headers, request,
                )
                row += 1

            for gap in section.gaps:
                self._write_merged_gap_row(
                    ws, row, gap, snapshot.project_key, proposed_fill,
                    headers, request,
                )
                row += 1

        _apply_header_style(ws, len(headers))
        _apply_status_conditional_formatting(ws, headers)
        if request.show_priority:
            _apply_priority_conditional_formatting(ws, headers)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(row - 1, 1)}'
        self._set_column_widths(ws, headers, is_merged_sheet=True)

    def _write_merged_jira_row(
        self,
        ws: Any,
        row: int,
        item: RoadmapItem,
        project_key: str,
        headers: List[str],
        request: RoadmapRequest,
    ) -> None:
        '''Write a Jira item row to the Merged Roadmap sheet.'''
        c = self._col
        for d in range(3):
            depth_header = ['Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)'][d]
            if item.depth == d:
                ws.cell(row=row, column=c(headers, depth_header), value=item.key)
            else:
                ws.cell(row=row, column=c(headers, depth_header), value='')

        ws.cell(row=row, column=c(headers, 'project'), value=project_key)

        issue_type_cell = ws.cell(
            row=row, column=c(headers, 'issue_type'), value=item.issue_type,
        )
        if request.bold_stories and item.issue_type == 'Story':
            issue_type_cell.font = Font(bold=True)

        ws.cell(row=row, column=c(headers, 'status'), value=item.status)

        if 'priority' in headers:
            ws.cell(row=row, column=c(headers, 'priority'), value=item.priority)

        key_cell = ws.cell(row=row, column=c(headers, 'key'), value=item.key)
        if item.key:
            key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{item.key}'
            key_cell.font = Font(color='0563C1', underline='single')

        ws.cell(row=row, column=c(headers, 'summary'), value=item.summary)

        assignee_val = item.assignee
        if request.blank_unassigned and assignee_val in ('Unassigned', '', None):
            assignee_val = ''
        ws.cell(row=row, column=c(headers, 'assignee'), value=assignee_val)

        ws.cell(row=row, column=c(headers, 'fix_version'), value=item.fix_version)
        ws.cell(row=row, column=c(headers, 'component'), value=item.component)
        ws.cell(row=row, column=c(headers, 'labels'), value=item.labels)
        ws.cell(row=row, column=c(headers, 'source'), value='Jira')
        ws.cell(row=row, column=c(headers, 'acceptance_criteria'), value='')
        ws.cell(row=row, column=c(headers, 'dependencies'), value='')

        if item.depth == 0:
            bold_font = Font(bold=True)
            key_col = c(headers, 'key')
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col_idx)
                if col_idx == key_col and item.key:
                    cell.font = Font(
                        bold=True, color='0563C1', underline='single'
                    )
                else:
                    cell.font = bold_font

    def _write_merged_gap_row(
        self,
        ws: Any,
        row: int,
        gap: RoadmapGap,
        project_key: str,
        proposed_fill: PatternFill,
        headers: List[str],
        request: RoadmapRequest,
    ) -> None:
        '''Write a proposed gap row to the Merged Roadmap sheet (highlighted).'''
        c = self._col
        for d in range(3):
            depth_header = ['Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)'][d]
            if gap.depth == d:
                ws.cell(row=row, column=c(headers, depth_header), value=gap.summary[:40])
            else:
                ws.cell(row=row, column=c(headers, depth_header), value='')

        ws.cell(row=row, column=c(headers, 'project'), value=project_key)

        issue_type_cell = ws.cell(
            row=row, column=c(headers, 'issue_type'), value=gap.issue_type,
        )
        if request.bold_stories and gap.issue_type == 'Story':
            issue_type_cell.font = Font(bold=True)

        ws.cell(row=row, column=c(headers, 'status'), value='')

        if 'priority' in headers:
            ws.cell(row=row, column=c(headers, 'priority'), value=gap.priority)

        ws.cell(row=row, column=c(headers, 'key'), value='')
        ws.cell(row=row, column=c(headers, 'summary'), value=gap.summary)
        ws.cell(row=row, column=c(headers, 'assignee'), value='')
        ws.cell(row=row, column=c(headers, 'fix_version'), value=gap.suggested_fix_version)
        ws.cell(row=row, column=c(headers, 'component'), value=gap.suggested_component)
        ws.cell(row=row, column=c(headers, 'labels'), value=gap.labels)
        ws.cell(row=row, column=c(headers, 'source'), value='Proposed')
        ws.cell(row=row, column=c(headers, 'acceptance_criteria'), value=gap.acceptance_criteria)
        ws.cell(row=row, column=c(headers, 'dependencies'), value=gap.dependencies)

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row, column=col_idx).fill = proposed_fill

    @staticmethod
    def _write_section_divider(
        ws: Any,
        row: int,
        title: str,
        num_cols: int,
    ) -> int:
        '''
        Write a section divider row — bold white text on dark blue fill.

        Returns the next row number (row + 1).
        '''
        divider_font = Font(bold=True, color='FFFFFF')
        divider_fill = PatternFill(
            start_color='2F5496', end_color='2F5496', fill_type='solid'
        )

        ws.cell(row=row, column=1, value=title)
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = divider_font
            cell.fill = divider_fill

        return row + 1

    @staticmethod
    def _set_column_widths(
        ws: Any,
        headers: List[str],
        is_gaps_sheet: bool = False,
        is_merged_sheet: bool = False,
    ) -> None:
        '''
        Set explicit column widths for the worksheet.

        Column widths are tuned for readability in each sheet type.
        Uses header-name lookup so widths stay correct when columns are
        added or removed (e.g. priority omitted).
        '''
        _BASE: Dict[str, int] = {
            'Depth 0 (Initiative)': 42,
            'Depth 1 (Epic)': 42,
            'Depth 2 (Story)': 48,
            'project': 7,
            'issue_type': 12,
            'status': 12,
            'priority': 6,
            'key': 12,
            'summary': 65,
            'assignee': 20,
            'fix_version': 18,
            'component': 30,
            'labels': 14,
            'source': 50,
            'acceptance_criteria': 25,
            'dependencies': 25,
        }

        _GAPS: Dict[str, int] = {
            'summary': 65,
            'suggested_component': 25,
            'acceptance_criteria': 50,
            'dependencies': 25,
            'suggested_fix_version': 18,
            'labels': 25,
        }

        width_map = {**_BASE, **(_GAPS if is_gaps_sheet else {})}

        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width_map.get(header, 14)

    @staticmethod
    def _format_summary(snapshot: RoadmapSnapshot) -> str:
        '''Build a human-readable Markdown summary of the roadmap snapshot.'''
        lines = [
            f'# Roadmap Analysis: {snapshot.project_key}',
            f'',
            f'**Scope**: {snapshot.scope_label or "All"}',
            f'**Snapshot ID**: {snapshot.snapshot_id}',
            f'**Created**: {snapshot.created_at}',
            f'**Jira Items**: {snapshot.total_jira_items}',
            f'**Proposed Gaps**: {snapshot.total_proposed_gaps}',
            f'**Total Items**: {snapshot.total_items}',
            f'',
            '---',
            '',
        ]

        for section in snapshot.sections:
            lines.append(f'## {section.title}')
            lines.append(f'  Items: {len(section.items)}, Gaps: {len(section.gaps)}')
            lines.append('')

            for item in section.items:
                indent = '  ' * item.depth
                lines.append(
                    f'{indent}- [{item.key}] {item.summary} '
                    f'({item.status}, {item.priority})'
                )

            if section.gaps:
                lines.append('')
                lines.append('  **Proposed Gaps:**')
                for gap in section.gaps:
                    lines.append(
                        f'  - [{gap.issue_type}] {gap.summary} '
                        f'({gap.priority})'
                    )

            lines.append('')

        return '\n'.join(lines)

    # ===========================================================================
    # Release Monitor — Bug tracking, velocity, readiness, and gap analysis
    # ===========================================================================

    def create_release_monitor(
        self,
        request: ReleaseMonitorRequest,
    ) -> ReleaseMonitorReport:
        '''
        Create a release health monitoring report.

        Algorithm:
        1. Discover releases (explicit or unreleased versions from Jira).
        2. Query all tickets per release; build bug summaries.
        3. Build ReleaseSnapshots via core.release_tracking.
        4. Optionally compute deltas, velocity, readiness, and gap analysis.
        5. Generate xlsx output.
        6. Return ReleaseMonitorReport.
        '''
        project_key = request.project_key or self.project_key or 'STL'
        log.info(
            f'Creating release monitor for {project_key} '
            f'(releases={request.releases}, scope={request.scope_label})'
        )

        # Step 1: Discover releases to monitor
        if request.releases:
            release_names = list(request.releases)
        else:
            # Query Jira for unreleased versions
            result = get_releases(
                project_key,
                include_released=False,
                include_unreleased=True,
            )
            if result.is_success:
                release_names = [
                    v.get('name', '') for v in result.data
                    if not v.get('released', False) and v.get('name')
                ]
            else:
                log.warning(f'Failed to fetch releases: {result.error}')
                release_names = []

        if not release_names:
            log.warning('No releases found to monitor')

        # Step 2: Query tickets and build bug summaries per release
        all_tickets_by_release: Dict[str, List[Dict[str, Any]]] = {}
        bug_summaries: List[BugSummary] = []
        snapshots: List[Any] = []

        for release in release_names:
            tickets = self._query_release_tickets(release, request.scope_label)
            all_tickets_by_release[release] = tickets

            mapping_tickets: list[Mapping[str, Any]] = list(tickets)
            snapshot = build_release_snapshot(mapping_tickets, release)
            snapshots.append(snapshot)

            # Build bug summary (bugs only)
            bug_summary = self._build_bug_summary(release, tickets)
            bug_summaries.append(bug_summary)

        # Step 3: Delta comparison with previous snapshot
        delta_data: Optional[Dict[str, Any]] = None
        if request.compare_to_previous and len(snapshots) >= 2:
            # Reasoning: when monitoring multiple releases, compare the last
            # two snapshots chronologically. For single-release use, the caller
            # should supply a stored previous snapshot externally.
            try:
                delta = compute_delta(snapshots[-1], snapshots[0])
                delta_data = {
                    'release': delta.release,
                    'new_tickets': delta.new_tickets,
                    'closed_tickets': delta.closed_tickets,
                    'status_changes': delta.status_changes,
                    'priority_changes': delta.priority_changes,
                    'new_p0_p1': delta.new_p0_p1,
                    'velocity': delta.velocity,
                }
                # Enrich bug summaries with delta info
                for bs in bug_summaries:
                    if bs.release == delta.release:
                        bs.new_since_last = delta.new_tickets
                        bs.closed_since_last = delta.closed_tickets
                        bs.priority_changes = delta.priority_changes
            except Exception as e:
                log.warning(f'Delta computation failed: {e}')

        # Step 4: Velocity
        velocity_data: Optional[Dict[str, Any]] = None
        if request.include_velocity and snapshots:
            try:
                velocity_data = compute_velocity(snapshots, window_days=14)
            except Exception as e:
                log.warning(f'Velocity computation failed: {e}')

        # Step 5: Readiness assessment
        readiness_data: Optional[Dict[str, Any]] = None
        if request.include_readiness and snapshots:
            try:
                config = TrackerConfig(
                    project=project_key,
                    releases=release_names,
                )
                # Use the last snapshot for readiness
                vel = velocity_data or {}
                readiness_report = assess_readiness(
                    snapshots[-1], vel, [], config,
                )
                readiness_data = {
                    'release': readiness_report.release,
                    'total_open': readiness_report.total_open,
                    'p0_open': readiness_report.p0_open,
                    'p1_open': readiness_report.p1_open,
                    'daily_close_rate': readiness_report.daily_close_rate,
                    'estimated_days_remaining': readiness_report.estimated_days_remaining,
                    'stale_tickets': readiness_report.stale_tickets,
                    'component_risks': readiness_report.component_risks,
                }
            except Exception as e:
                log.warning(f'Readiness assessment failed: {e}')

        # Step 6: Gap analysis (reuse roadmap snapshot machinery)
        roadmap_data: Optional[Dict[str, Any]] = None
        if request.include_gap_analysis:
            try:
                roadmap_request = RoadmapRequest(
                    project_key=project_key,
                    scope_label=request.scope_label or '',
                    fix_versions=release_names,
                    include_gap_analysis=True,
                )
                roadmap_snap = self.create_roadmap_snapshot(roadmap_request)
                roadmap_data = roadmap_snap.to_dict()
            except Exception as e:
                log.warning(f'Gap analysis failed: {e}')

        # Build the report
        report = ReleaseMonitorReport(
            project_key=project_key,
            created_at=datetime.now(timezone.utc).isoformat(),
            releases_monitored=release_names,
            bug_summaries=bug_summaries,
            velocity=velocity_data,
            readiness=readiness_data,
            roadmap_snapshot=roadmap_data,
            delta=delta_data,
        )

        # Generate markdown summary
        report.summary_markdown = self._format_release_monitor_summary(report)

        # Step 7: Generate xlsx
        if request.output_file:
            output_path = self._write_release_monitor_xlsx(
                report, request.output_file, all_tickets_by_release,
            )
            report.output_file = output_path
            log.info(f'Wrote release monitor xlsx to {output_path}')

        return report

    def _query_release_tickets(
        self,
        release: str,
        scope_label: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        '''
        Query Jira for all tickets in a given release.

        Builds JQL with fixVersion filter and optional scope_label keyword.
        Returns a list of normalized ticket dicts.
        '''
        jql_parts = [
            f'project = {self.project_key or "STL"}',
            f'fixVersion = "{release}"',
        ]
        if scope_label:
            jql_parts.append(f'labels = "{scope_label}"')
        jql = ' AND '.join(jql_parts) + ' ORDER BY priority ASC, updated DESC'

        log.debug(f'Release ticket query JQL: {jql}')

        try:
            result = search_tickets(jql=jql, limit=500)
            if not result.is_success:
                log.warning(f'Release ticket search failed: {result.error}')
                return []
            tickets = result.data or []
        except Exception as e:
            log.error(f'Release ticket query failed for {release}: {e}')
            return []

        # Normalize ticket dicts to ensure consistent field names
        normalized: List[Dict[str, Any]] = []
        for t in tickets:
            normalized.append({
                'key': t.get('key', ''),
                'summary': t.get('summary', ''),
                'status': t.get('status', ''),
                'priority': t.get('priority', ''),
                'issuetype': t.get('issue_type') or t.get('type') or '',
                'assignee': t.get('assignee', ''),
                'fix_versions': t.get('fix_versions') or t.get('fix_version', ''),
                'components': t.get('components') or t.get('component', ''),
                'created': t.get('created', ''),
                'updated': t.get('updated', ''),
                'resolutiondate': t.get('resolutiondate', ''),
            })
        return normalized

    def _build_bug_summary(
        self,
        release: str,
        tickets: List[Dict[str, Any]],
        previous_tickets: Optional[List[Dict[str, Any]]] = None,
    ) -> BugSummary:
        '''
        Build a BugSummary from a list of tickets for a release.

        Filters to bugs only, counts by status/priority/component.
        If previous_tickets provided, computes new/closed/priority deltas.
        '''
        # Filter to bugs only
        bugs = [
            t for t in tickets
            if str(t.get('issuetype', '')).lower() == 'bug'
        ]

        by_status: Counter[str] = Counter()
        by_priority: Counter[str] = Counter()
        by_component: Counter[str] = Counter()

        for bug in bugs:
            by_status[str(bug.get('status', 'Unknown'))] += 1
            by_priority[str(bug.get('priority', 'Unknown'))] += 1
            # Components may be a list or csv string
            comps = bug.get('components', '')
            if isinstance(comps, list):
                for c in comps:
                    by_component[str(c) or 'Unspecified'] += 1
            elif isinstance(comps, str) and comps.strip():
                by_component[comps.strip()] += 1
            else:
                by_component['Unspecified'] += 1

        summary = BugSummary(
            release=release,
            total_bugs=len(bugs),
            by_status=dict(by_status),
            by_priority=dict(by_priority),
            by_component=dict(by_component),
        )

        # Compute deltas if previous tickets are available
        if previous_tickets is not None:
            prev_bugs = {
                str(t.get('key')): t for t in previous_tickets
                if str(t.get('issuetype', '')).lower() == 'bug'
            }
            curr_bugs = {str(b.get('key')): b for b in bugs}

            summary.new_since_last = [
                k for k in curr_bugs if k not in prev_bugs
            ]
            summary.closed_since_last = [
                k for k in prev_bugs if k not in curr_bugs
            ]
            summary.priority_changes = [
                {
                    'key': k,
                    'from': str(prev_bugs[k].get('priority', '')),
                    'to': str(curr_bugs[k].get('priority', '')),
                }
                for k in curr_bugs
                if k in prev_bugs
                and str(curr_bugs[k].get('priority', ''))
                != str(prev_bugs[k].get('priority', ''))
            ]

        return summary

    def _write_release_monitor_xlsx(
        self,
        report: ReleaseMonitorReport,
        output_file: str,
        all_release_tickets: Optional[Dict[str, List[Dict[str, Any]]]] = None,
    ) -> str:
        '''
        Generate a multi-sheet xlsx workbook for the release monitor report.

        Sheets:
        1. Bug Summary — one row per release with counts
        2. Bug Details — all bugs with hyperlinks and formatting
        3. Changes Since Last — delta rows (if delta data exists)
        4. Readiness — metric/value pairs (if readiness data exists)
        5. Gap Analysis — roadmap gap summary (if roadmap data exists)
        '''
        wb = Workbook()

        # -- Sheet 1: Bug Summary --
        ws_summary = wb.active or wb.create_sheet()
        ws_summary.title = 'Bug Summary'
        self._write_bug_summary_sheet(ws_summary, report)

        # -- Sheet 2: Bug Details --
        ws_details = wb.create_sheet('Bug Details')
        self._write_bug_details_sheet(
            ws_details, report, all_release_tickets or {},
        )

        # -- Sheet 3: Changes Since Last (conditional) --
        if report.delta:
            ws_changes = wb.create_sheet('Changes Since Last')
            self._write_changes_sheet(ws_changes, report)

        # -- Sheet 4: Readiness (conditional) --
        if report.readiness:
            ws_readiness = wb.create_sheet('Readiness')
            self._write_readiness_sheet(ws_readiness, report)

        # -- Sheet 5: Gap Analysis (conditional) --
        if report.roadmap_snapshot:
            ws_gaps = wb.create_sheet('Gap Analysis')
            self._write_gap_analysis_sheet(ws_gaps, report)

        os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
        wb.save(output_file)
        return output_file

    def _write_bug_summary_sheet(
        self,
        ws: Any,
        report: ReleaseMonitorReport,
    ) -> None:
        '''Write the Bug Summary sheet — one row per release.'''
        headers = [
            'Release', 'Total Bugs', 'P0', 'P1', 'P2', 'P3',
            'Open', 'In Progress', 'Verify', 'Closed',
        ]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, bs in enumerate(report.bug_summaries, 2):
            ws.cell(row=row_idx, column=1, value=bs.release)
            ws.cell(row=row_idx, column=2, value=bs.total_bugs)
            # Priority columns — match P0-Stopper, P1-Critical, etc.
            ws.cell(row=row_idx, column=3, value=self._priority_count(bs, 'P0'))
            ws.cell(row=row_idx, column=4, value=self._priority_count(bs, 'P1'))
            ws.cell(row=row_idx, column=5, value=self._priority_count(bs, 'P2'))
            ws.cell(row=row_idx, column=6, value=self._priority_count(bs, 'P3'))
            # Status columns
            ws.cell(row=row_idx, column=7, value=bs.by_status.get('Open', 0))
            ws.cell(row=row_idx, column=8, value=bs.by_status.get('In Progress', 0))
            ws.cell(row=row_idx, column=9, value=bs.by_status.get('Verify', 0))
            ws.cell(row=row_idx, column=10, value=bs.by_status.get('Closed', 0))

        _apply_header_style(ws, len(headers))
        _apply_status_conditional_formatting(ws, headers)
        _apply_priority_conditional_formatting(ws, headers)
        _auto_fit_columns(ws)

    def _write_bug_details_sheet(
        self,
        ws: Any,
        report: ReleaseMonitorReport,
        all_release_tickets: Dict[str, List[Dict[str, Any]]],
    ) -> None:
        '''Write the Bug Details sheet — all bugs sorted by priority then release.'''
        headers = [
            'Release', 'Key', 'Summary', 'Status', 'Priority',
            'Assignee', 'Component', 'Created', 'Updated',
        ]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        row = 2
        for bs in report.bug_summaries:
            tickets = all_release_tickets.get(bs.release, [])
            bugs = [
                t for t in tickets
                if str(t.get('issuetype', '')).lower() == 'bug'
            ]
            # Sort by priority (P0 first)
            bugs.sort(key=lambda t: str(t.get('priority', 'ZZZ')))

            for bug in bugs:
                ws.cell(row=row, column=1, value=bs.release)
                key = str(bug.get('key', ''))
                key_cell = ws.cell(row=row, column=2, value=key)
                if key:
                    key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{key}'
                    key_cell.font = Font(color='0563C1', underline='single')
                ws.cell(row=row, column=3, value=bug.get('summary', ''))
                ws.cell(row=row, column=4, value=bug.get('status', ''))
                ws.cell(row=row, column=5, value=bug.get('priority', ''))
                ws.cell(row=row, column=6, value=bug.get('assignee', ''))
                comps = bug.get('components', '')
                if isinstance(comps, list):
                    comps = ', '.join(str(c) for c in comps)
                ws.cell(row=row, column=7, value=comps)
                ws.cell(row=row, column=8, value=bug.get('created', ''))
                ws.cell(row=row, column=9, value=bug.get('updated', ''))
                row += 1

        _apply_header_style(ws, len(headers))
        _apply_status_conditional_formatting(ws, headers)
        _apply_priority_conditional_formatting(ws, headers)
        ws.freeze_panes = 'A2'
        _auto_fit_columns(ws)

    def _write_changes_sheet(
        self,
        ws: Any,
        report: ReleaseMonitorReport,
    ) -> None:
        '''Write the Changes Since Last sheet — delta rows by change type.'''
        headers = ['Change Type', 'Release', 'Key', 'Summary', 'Detail']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        delta = report.delta or {}
        release = delta.get('release', '')
        row = 2

        # Section: New Bugs
        new_tickets = delta.get('new_tickets', [])
        if new_tickets:
            row = self._write_section_divider(ws, row, 'New Bugs', len(headers))
            for key in new_tickets:
                ws.cell(row=row, column=1, value='New Bug')
                ws.cell(row=row, column=2, value=release)
                key_cell = ws.cell(row=row, column=3, value=key)
                key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{key}'
                key_cell.font = Font(color='0563C1', underline='single')
                ws.cell(row=row, column=4, value='')
                ws.cell(row=row, column=5, value='Added since last snapshot')
                row += 1

        # Section: Closed Bugs
        closed_tickets = delta.get('closed_tickets', [])
        if closed_tickets:
            row = self._write_section_divider(
                ws, row, 'Closed Bugs', len(headers),
            )
            for key in closed_tickets:
                ws.cell(row=row, column=1, value='Closed Bug')
                ws.cell(row=row, column=2, value=release)
                key_cell = ws.cell(row=row, column=3, value=key)
                key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{key}'
                key_cell.font = Font(color='0563C1', underline='single')
                ws.cell(row=row, column=4, value='')
                ws.cell(row=row, column=5, value='Closed since last snapshot')
                row += 1

        # Section: Priority Changes
        priority_changes = delta.get('priority_changes', [])
        if priority_changes:
            row = self._write_section_divider(
                ws, row, 'Priority Changes', len(headers),
            )
            for change in priority_changes:
                ws.cell(row=row, column=1, value='Priority Change')
                ws.cell(row=row, column=2, value=release)
                key = str(change.get('key', ''))
                key_cell = ws.cell(row=row, column=3, value=key)
                if key:
                    key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{key}'
                    key_cell.font = Font(color='0563C1', underline='single')
                ws.cell(row=row, column=4, value='')
                ws.cell(
                    row=row, column=5,
                    value=f'{change.get("from", "")} \u2192 {change.get("to", "")}',
                )
                row += 1

        # Section: Status Changes
        status_changes = delta.get('status_changes', [])
        if status_changes:
            row = self._write_section_divider(
                ws, row, 'Status Changes', len(headers),
            )
            for change in status_changes:
                ws.cell(row=row, column=1, value='Status Change')
                ws.cell(row=row, column=2, value=release)
                key = str(change.get('key', ''))
                key_cell = ws.cell(row=row, column=3, value=key)
                if key:
                    key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{key}'
                    key_cell.font = Font(color='0563C1', underline='single')
                ws.cell(row=row, column=4, value='')
                ws.cell(
                    row=row, column=5,
                    value=f'{change.get("from", "")} \u2192 {change.get("to", "")}',
                )
                row += 1

        _apply_header_style(ws, len(headers))
        _auto_fit_columns(ws)

    def _write_readiness_sheet(
        self,
        ws: Any,
        report: ReleaseMonitorReport,
    ) -> None:
        '''Write the Readiness sheet — metric/value pairs.'''
        headers = ['Metric', 'Value']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        readiness = report.readiness or {}
        metrics = [
            ('Release', readiness.get('release', '')),
            ('Total Open', readiness.get('total_open', 0)),
            ('P0 Open', readiness.get('p0_open', 0)),
            ('P1 Open', readiness.get('p1_open', 0)),
            ('Daily Close Rate', f'{readiness.get("daily_close_rate", 0.0):.2f}'),
            (
                'Estimated Days Remaining',
                readiness.get('estimated_days_remaining', 'N/A'),
            ),
            ('Stale Ticket Count', len(readiness.get('stale_tickets', []))),
        ]

        row = 2
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=str(value))
            row += 1

        # Component risks as sub-rows
        component_risks = readiness.get('component_risks', [])
        if component_risks:
            row = self._write_section_divider(
                ws, row, 'Component Risks', len(headers),
            )
            for risk in component_risks:
                ws.cell(
                    row=row, column=1,
                    value=f'  {risk.get("component", "Unknown")}',
                )
                ws.cell(
                    row=row, column=2,
                    value=str(risk.get('open_tickets', 0)),
                )
                row += 1

        _apply_header_style(ws, len(headers))
        _auto_fit_columns(ws)

    def _write_gap_analysis_sheet(
        self,
        ws: Any,
        report: ReleaseMonitorReport,
    ) -> None:
        '''Write the Gap Analysis sheet — roadmap snapshot summary.'''
        headers = ['Section', 'Jira Items', 'Proposed Gaps', 'Summary']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        roadmap = report.roadmap_snapshot or {}
        sections = roadmap.get('sections', [])

        row = 2
        for section in sections:
            ws.cell(row=row, column=1, value=section.get('title', ''))
            ws.cell(row=row, column=2, value=len(section.get('items', [])))
            ws.cell(row=row, column=3, value=len(section.get('gaps', [])))
            # Build a brief summary of gaps
            gap_summaries = [
                g.get('summary', '') for g in section.get('gaps', [])
            ]
            ws.cell(
                row=row, column=4,
                value='; '.join(gap_summaries[:5]) if gap_summaries else 'No gaps identified',
            )
            row += 1

        _apply_header_style(ws, len(headers))
        _auto_fit_columns(ws)

    @staticmethod
    def _priority_count(bs: BugSummary, prefix: str) -> int:
        '''
        Sum bug counts for all priority values starting with the given prefix.

        Reasoning: Jira priorities may be "P0-Stopper", "P1-Critical", etc.
        We match any key starting with the prefix (e.g. "P0").
        '''
        return sum(
            count for key, count in bs.by_priority.items()
            if key.startswith(prefix)
        )

    @staticmethod
    def _format_release_monitor_summary(
        report: ReleaseMonitorReport,
    ) -> str:
        '''
        Generate a markdown summary of the release monitor report.

        Includes key metrics, bug trends, P0/P1 alerts, and velocity.
        '''
        lines = [
            f'# Release Monitor: {report.project_key}',
            f'',
            f'**Report ID**: {report.report_id}',
            f'**Created**: {report.created_at}',
            f'**Releases Monitored**: {", ".join(report.releases_monitored)}',
            f'',
            '---',
            '',
        ]

        # Bug summary per release
        lines.append('## Bug Summary')
        lines.append('')
        for bs in report.bug_summaries:
            lines.append(f'### {bs.release}')
            lines.append(f'- **Total Bugs**: {bs.total_bugs}')
            for prio, count in sorted(bs.by_priority.items()):
                lines.append(f'  - {prio}: {count}')
            for status, count in sorted(bs.by_status.items()):
                lines.append(f'  - {status}: {count}')
            if bs.new_since_last:
                lines.append(
                    f'- **New since last**: {len(bs.new_since_last)} '
                    f'({", ".join(bs.new_since_last[:5])})'
                )
            if bs.closed_since_last:
                lines.append(
                    f'- **Closed since last**: {len(bs.closed_since_last)} '
                    f'({", ".join(bs.closed_since_last[:5])})'
                )
            lines.append('')

        # P0/P1 alerts
        if report.total_p0 > 0 or report.total_p1 > 0:
            lines.append('## \u26a0 Priority Alerts')
            lines.append('')
            if report.total_p0 > 0:
                lines.append(f'- **P0 (Stopper)**: {report.total_p0} open')
            if report.total_p1 > 0:
                lines.append(f'- **P1 (Critical)**: {report.total_p1} open')
            lines.append('')

        # Velocity
        if report.velocity:
            vel = report.velocity
            lines.append('## Velocity')
            lines.append('')
            lines.append(f'- Opened: {vel.get("opened", 0)}')
            lines.append(f'- Closed: {vel.get("closed", 0)}')
            lines.append(f'- Daily open rate: {vel.get("daily_open_rate", 0.0):.2f}')
            lines.append(f'- Daily close rate: {vel.get("daily_close_rate", 0.0):.2f}')
            lines.append(f'- Daily net rate: {vel.get("daily_net_rate", 0.0):.2f}')
            net = float(vel.get('daily_net_rate', 0.0))
            if net > 0:
                lines.append('- **Trend**: Bug count increasing \u2191')
            elif net < 0:
                lines.append('- **Trend**: Bug count decreasing \u2193')
            else:
                lines.append('- **Trend**: Stable \u2192')
            lines.append('')

        # Readiness
        if report.readiness:
            r = report.readiness
            lines.append('## Readiness')
            lines.append('')
            lines.append(f'- Total open: {r.get("total_open", 0)}')
            lines.append(f'- P0 open: {r.get("p0_open", 0)}')
            lines.append(f'- P1 open: {r.get("p1_open", 0)}')
            est = r.get('estimated_days_remaining')
            if est is not None:
                lines.append(f'- Estimated days remaining: {est:.1f}')
            else:
                lines.append('- Estimated days remaining: N/A (no close rate)')
            stale = r.get('stale_tickets', [])
            if stale:
                lines.append(
                    f'- Stale tickets: {len(stale)} '
                    f'({", ".join(stale[:5])})'
                )
            lines.append('')

        return '\n'.join(lines)
