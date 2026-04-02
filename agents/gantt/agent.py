##########################################################################################
#
# Module: agents/gantt_agent.py
#
# Description: Gantt Project Planner Agent.
#              Produces evidence-backed planning snapshots, milestone proposals,
#              dependency views, and planning-risk summaries from Jira data.
#
# Author: Cornelis Networks
#
##########################################################################################

from __future__ import annotations

import json
import logging
import os
import sys
import time
from datetime import datetime, timezone
from html import escape
from typing import Any, Dict, Iterable, List, Mapping, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from agents.base import BaseAgent, AgentConfig, AgentResponse
from llm.base import Message
from agents.gantt.components import (
    BacklogInterpreter,
    DependencyMapper,
    MilestonePlanner,
    PlanningSummarizer,
    RiskProjector,
)
from agents.gantt.models import (
    BugSummary,
    DependencyEdge,
    DependencyGraph,
    MilestoneProposal,
    PlanningRequest,
    PlanningRiskRecord,
    PlanningSnapshot,
    ReleaseMonitorReport,
    ReleaseMonitorRequest,
    ReleaseSurveyReleaseSummary,
    ReleaseSurveyReport,
    ReleaseSurveyRequest,
    RoadmapGap,
    RoadmapItem,
    RoadmapRequest,
    RoadmapSection,
    RoadmapSnapshot,
)
from agents.pm_runtime import normalize_csv_list, notify_shannon
from collections import Counter
from core.evidence import EvidenceBundle, load_evidence_bundle
from core.release_tracking import (
    ReleaseSnapshot,
    build_snapshot as build_release_snapshot,
    compute_delta,
    compute_velocity,
    compute_cycle_time_stats,
    assess_readiness,
    format_summary as format_release_summary,
    TrackerConfig,
)
from agents.gantt.state.dependency_review_store import GanttDependencyReviewStore
from agents.gantt.state.release_monitor_store import GanttReleaseMonitorStore
from agents.gantt.state.release_survey_store import GanttReleaseSurveyStore
from agents.gantt.state.snapshot_store import GanttSnapshotStore
from excel_utils import (
    STATUS_FILL_COLORS,
    PRIORITY_FILL_COLORS,
    _apply_header_style,
    _apply_status_conditional_formatting,
    _apply_priority_conditional_formatting,
    _auto_fit_columns,
)
from tools.jira_tools import (
    JiraTools, get_jira, get_project_info, get_releases,
    search_tickets, get_children_hierarchy,
)
from tools.knowledge_tools import search_knowledge, list_knowledge_files, read_knowledge_file

# Logging config - follows jira_utils.py pattern
log = logging.getLogger(os.path.basename(sys.argv[0]))

# ---------------------------------------------------------------------------
# Roadmap constants (from roadmap_agent.py)
# ---------------------------------------------------------------------------

# Jira base URL for hyperlinks
JIRA_BASE_URL = 'https://cornelisnetworks.atlassian.net'
CONFLUENCE_JIRA_SERVER = os.getenv('CONFLUENCE_JIRA_SERVER') or 'System Jira'
CONFLUENCE_JIRA_SERVER_ID = (
    os.getenv('CONFLUENCE_JIRA_SERVER_ID')
    or os.getenv('JIRA_SERVER_ID')
    or '332fe428-27be-3c06-ad09-b2cd4d269bee'
)

# Issue types to exclude from the roadmap (Bugs are not roadmap items)
_EXCLUDED_TYPES = {'Bug', 'bug'}

# Depth mapping for issue types
_TYPE_DEPTH = {
    'Initiative': 0,
    'Epic': 1,
    'Story': 2,
    'Task': 2,
    'Sub-task': 2,
}

# Done statuses — used when include_closed is False
_DONE_STATUSES = {'Closed', 'Done', 'Resolved'}


class GanttProjectPlannerAgent(BaseAgent):
    '''
    Agent for producing project-planning snapshots from Jira data.

    The implementation is deterministic-first. Specialized planner components
    handle backlog interpretation, dependency mapping, milestone planning,
    risk projection, and summary formatting.
    '''

    STALE_DAYS = 30
    RELEASE_SURVEY_MODE_FEATURE_DEV = 'feature_dev'
    RELEASE_SURVEY_MODE_BUG = 'bug'
    _RELEASE_SURVEY_MANAGER_ALIASES = {
        'cook sam': 'samuel cook',
        'dennis dalessandro': 'denny dalessandro',
        'nicholas child': 'nick child',
        'sam cook': 'samuel cook',
    }
    _release_survey_manager_lookup_cache: Optional[Dict[str, str]] = None

    def __init__(self, project_key: Optional[str] = None, **kwargs):
        '''
        Initialize the Gantt Project Planner agent.
        '''
        instruction = self._load_prompt_file()
        if not instruction:
            raise FileNotFoundError(
                'agents/gantt/prompts/system.md is required but not found. '
                'The Gantt Project Planner Agent has no hardcoded fallback prompt.'
            )

        config = AgentConfig(
            name='gantt_project_planner',
            description='Builds planning snapshots, milestones, dependencies, and risks',
            instruction=instruction,
            max_iterations=10,
        )

        super().__init__(config=config, tools=[JiraTools()], **kwargs)

        self.register_tool(search_knowledge)
        self.register_tool(list_knowledge_files)
        self.register_tool(read_knowledge_file)

        self.project_key = project_key
        self.backlog_interpreter = BacklogInterpreter(
            jira_provider=get_jira,
            now_provider=self._utc_now,
            stale_days=self.STALE_DAYS,
        )
        self.dependency_mapper = DependencyMapper(
            review_store=GanttDependencyReviewStore()
        )
        self.milestone_planner = MilestonePlanner()
        self.risk_projector = RiskProjector()
        self.planning_summarizer = PlanningSummarizer()

    @staticmethod
    def _load_prompt_file() -> Optional[str]:
        '''Load the Gantt agent prompt from config/prompts/.'''
        prompt_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 'prompts', 'system.md'
        )
        if os.path.exists(prompt_path):
            try:
                with open(prompt_path, 'r', encoding='utf-8') as f:
                    return f.read()
            except Exception as e:
                log.warning(f'Failed to load Gantt agent prompt: {e}')
        return None

    @staticmethod
    def _utc_now() -> datetime:
        '''Return the current UTC time as a timezone-aware datetime.'''
        return datetime.now(timezone.utc)

    def run(self, input_data: Any) -> AgentResponse:
        '''
        Build a planning snapshot from Jira data.
        '''
        log.debug(f'GanttProjectPlannerAgent.run(input_data={input_data})')

        if isinstance(input_data, str):
            request = PlanningRequest(project_key=input_data)
        elif isinstance(input_data, dict):
            request = PlanningRequest(
                project_key=input_data.get('project_key', self.project_key or ''),
                planning_horizon_days=int(input_data.get('planning_horizon_days', 90)),
                limit=int(input_data.get('limit', 200)),
                include_done=bool(input_data.get('include_done', False)),
                backlog_jql=input_data.get('backlog_jql'),
                policy_profile=input_data.get('policy_profile', 'default'),
                evidence_paths=[
                    str(path) for path in (input_data.get('evidence_paths') or [])
                ],
            )
        else:
            return AgentResponse.error_response(
                'Invalid input: expected project key string or request dict'
            )

        if not request.project_key:
            return AgentResponse.error_response('No project_key provided')

        try:
            snapshot = self.create_snapshot(request)
        except Exception as e:
            log.error(f'Gantt planning snapshot failed: {e}')
            return AgentResponse.error_response(str(e))

        return AgentResponse.success_response(
            content=snapshot.summary_markdown,
            metadata={'planning_snapshot': snapshot.to_dict()},
        )

    def run_once(
        self,
        task_type: str,
        request: PlanningRequest | ReleaseMonitorRequest | ReleaseSurveyRequest,
        *,
        persist: bool = True,
    ) -> Dict[str, Any]:
        '''
        Execute one deterministic Gantt task and optionally persist the result.
        '''
        if task_type == 'planning_snapshot':
            if not isinstance(request, PlanningRequest):
                raise TypeError('planning_snapshot requires a PlanningRequest')

            self.project_key = request.project_key or self.project_key
            snapshot = self.create_snapshot(request)
            result: Dict[str, Any] = {
                'ok': True,
                'task_type': task_type,
                'project_key': snapshot.project_key,
                'snapshot': snapshot.to_dict(),
                'summary_markdown': snapshot.summary_markdown,
            }
            if persist:
                result['stored'] = GanttSnapshotStore().save_snapshot(
                    snapshot,
                    summary_markdown=snapshot.summary_markdown,
                )
            return result

        if task_type == 'release_monitor':
            if not isinstance(request, ReleaseMonitorRequest):
                raise TypeError('release_monitor requires a ReleaseMonitorRequest')

            self.project_key = request.project_key or self.project_key
            report = self.create_release_monitor(request)
            result = {
                'ok': True,
                'task_type': task_type,
                'project_key': report.project_key,
                'report': report.to_dict(),
                'summary_markdown': report.summary_markdown,
            }
            if persist:
                result['stored'] = GanttReleaseMonitorStore().save_report(
                    report,
                    summary_markdown=report.summary_markdown,
                )
            return result

        if task_type == 'release_survey':
            if not isinstance(request, ReleaseSurveyRequest):
                raise TypeError('release_survey requires a ReleaseSurveyRequest')

            self.project_key = request.project_key or self.project_key
            report = self.create_release_survey(request)
            result = {
                'ok': True,
                'task_type': task_type,
                'project_key': report.project_key,
                'survey': report.to_dict(),
                'summary_markdown': report.summary_markdown,
            }
            if persist:
                result['stored'] = GanttReleaseSurveyStore().save_survey(
                    report,
                    summary_markdown=report.summary_markdown,
                )
            return result

        raise ValueError(f'Unsupported Gantt task_type: {task_type}')

    @staticmethod
    def _build_snapshot_notification_payload(result: Dict[str, Any]) -> Dict[str, Any]:
        snapshot = result.get('snapshot', {})
        overview = snapshot.get('backlog_overview', {})
        stored = result.get('stored', {})
        text = (
            f'{snapshot.get("project_key", "")}: '
            f'{overview.get("total_issues", 0)} issues, '
            f'{overview.get("blocked_issues", 0)} blocked, '
            f'{len(snapshot.get("milestones", []))} milestones'
        )
        body_lines = [
            f'Snapshot ID: {stored.get("snapshot_id") or snapshot.get("snapshot_id", "")}',
            f'Planning horizon: {snapshot.get("planning_horizon_days", 0)} days',
            f'Risks: {len(snapshot.get("risks", []))}',
            f'Dependencies: {snapshot.get("dependency_graph", {}).get("edge_count", 0)}',
        ]
        return {
            'title': f'Gantt Planning Snapshot — {snapshot.get("project_key", "")}',
            'text': text,
            'body_lines': body_lines,
        }

    @staticmethod
    def _build_release_monitor_notification_payload(result: Dict[str, Any]) -> Dict[str, Any]:
        report = result.get('report', {})
        stored = result.get('stored', {})
        text = (
            f'{report.get("project_key", "")}: '
            f'{report.get("total_bugs", 0)} bugs, '
            f'P0={report.get("total_p0", 0)}, '
            f'P1={report.get("total_p1", 0)}'
        )
        body_lines = [
            f'Report ID: {stored.get("report_id") or report.get("report_id", "")}',
            f'Releases: {", ".join(report.get("releases_monitored", [])) or "none"}',
        ]
        readiness = report.get('readiness') or {}
        if readiness:
            body_lines.append(
                f'Readiness: {readiness.get("total_open", 0)} open, '
                f'{readiness.get("p0_open", 0)} P0, '
                f'{readiness.get("p1_open", 0)} P1'
            )
        return {
            'title': f'Gantt Release Monitor — {report.get("project_key", "")}',
            'text': text,
            'body_lines': body_lines,
        }

    @staticmethod
    def _build_release_survey_notification_payload(result: Dict[str, Any]) -> Dict[str, Any]:
        survey = result.get('survey', {})
        stored = result.get('stored', {})
        text = (
            f'{survey.get("project_key", "")}: '
            f'{survey.get("done_count", 0)} done, '
            f'{survey.get("in_progress_count", 0)} in progress, '
            f'{survey.get("remaining_count", 0)} remaining'
        )
        body_lines = [
            f'Survey ID: {stored.get("survey_id") or survey.get("survey_id", "")}',
            f'Releases: {", ".join(survey.get("releases_surveyed", [])) or "none"}',
            f'Blocked: {survey.get("blocked_count", 0)}',
        ]
        return {
            'title': f'Gantt Release Survey — {survey.get("project_key", "")}',
            'text': text,
            'body_lines': body_lines,
        }

    def tick(self, poller_spec: Dict[str, Any]) -> Dict[str, Any]:
        '''
        Execute one scheduled Gantt cycle.
        '''
        project_key = str(
            poller_spec.get('project_key') or self.project_key or ''
        ).strip()
        if not project_key:
            raise ValueError('Gantt tick requires project_key')

        self.project_key = project_key
        persist = bool(poller_spec.get('persist', True))
        notify_enabled = bool(poller_spec.get('notify_shannon', False))
        shannon_base_url = poller_spec.get('shannon_base_url')
        tasks: List[Dict[str, Any]] = []
        notifications: List[Dict[str, Any]] = []
        errors: List[str] = []

        if bool(poller_spec.get('run_planning', True)):
            try:
                planning_request = PlanningRequest(
                    project_key=project_key,
                    planning_horizon_days=int(
                        poller_spec.get('planning_horizon_days', 90)
                    ),
                    limit=int(poller_spec.get('limit', 200)),
                    include_done=bool(poller_spec.get('include_done', False)),
                    backlog_jql=poller_spec.get('backlog_jql'),
                    policy_profile=str(
                        poller_spec.get('policy_profile', 'default') or 'default'
                    ),
                    evidence_paths=normalize_csv_list(
                        poller_spec.get('evidence_paths')
                    ),
                )
                planning_result = self.run_once(
                    'planning_snapshot',
                    planning_request,
                    persist=persist,
                )
                tasks.append(planning_result)
                if notify_enabled:
                    notifications.append(
                        notify_shannon(
                            agent_id='gantt',
                            shannon_base_url=shannon_base_url,
                            **self._build_snapshot_notification_payload(planning_result),
                        )
                    )
            except Exception as e:
                errors.append(f'planning_snapshot: {e}')

        release_names = normalize_csv_list(poller_spec.get('releases'))
        run_release_monitor = bool(poller_spec.get('run_release_monitor', False))
        if release_names:
            run_release_monitor = True

        if run_release_monitor:
            try:
                release_request = ReleaseMonitorRequest(
                    project_key=project_key,
                    releases=release_names or None,
                    scope_label=poller_spec.get('scope_label'),
                    include_gap_analysis=bool(
                        poller_spec.get('include_gap_analysis', True)
                    ),
                    include_bug_report=bool(
                        poller_spec.get('include_bug_report', True)
                    ),
                    include_velocity=bool(
                        poller_spec.get('include_velocity', True)
                    ),
                    include_readiness=bool(
                        poller_spec.get('include_readiness', True)
                    ),
                    compare_to_previous=bool(
                        poller_spec.get('compare_to_previous', True)
                    ),
                    output_file=poller_spec.get('output_file'),
                )
                release_result = self.run_once(
                    'release_monitor',
                    release_request,
                    persist=persist,
                )
                tasks.append(release_result)
                if notify_enabled:
                    notifications.append(
                        notify_shannon(
                            agent_id='gantt',
                            shannon_base_url=shannon_base_url,
                            **self._build_release_monitor_notification_payload(
                                release_result
                            ),
                        )
                    )
            except Exception as e:
                errors.append(f'release_monitor: {e}')

        run_release_survey = bool(poller_spec.get('run_release_survey', False))
        if run_release_survey:
            try:
                survey_request = ReleaseSurveyRequest(
                    project_key=project_key,
                    releases=release_names or None,
                    scope_label=poller_spec.get('scope_label'),
                    survey_mode=str(
                        poller_spec.get('survey_mode')
                        or poller_spec.get('release_survey_mode')
                        or self.RELEASE_SURVEY_MODE_FEATURE_DEV
                    ),
                    output_file=poller_spec.get('survey_output_file')
                    or poller_spec.get('output_file'),
                )
                survey_result = self.run_once(
                    'release_survey',
                    survey_request,
                    persist=persist,
                )
                tasks.append(survey_result)
                if notify_enabled:
                    notifications.append(
                        notify_shannon(
                            agent_id='gantt',
                            shannon_base_url=shannon_base_url,
                            **self._build_release_survey_notification_payload(
                                survey_result
                            ),
                        )
                    )
            except Exception as e:
                errors.append(f'release_survey: {e}')

        return {
            'ok': not errors,
            'agent_id': 'gantt',
            'project_key': project_key,
            'tasks': tasks,
            'notifications': notifications,
            'errors': errors,
            'completed_at': datetime.now(timezone.utc).isoformat(),
        }

    def run_poller(
        self,
        poller_spec: Dict[str, Any],
        *,
        sleep_fn: Any = time.sleep,
    ) -> Dict[str, Any]:
        '''
        Run the Gantt polling loop for a bounded number of cycles.
        '''
        interval_seconds = max(int(poller_spec.get('interval_seconds', 300) or 0), 0)
        max_cycles = int(poller_spec.get('max_cycles', 1) or 0)
        if max_cycles < 0:
            raise ValueError('max_cycles must be >= 0')
        if max_cycles == 0 and interval_seconds <= 0:
            raise ValueError(
                'Continuous polling requires interval_seconds > 0'
            )

        cycle_summaries: List[Dict[str, Any]] = []
        cycles_run = 0
        last_tick: Optional[Dict[str, Any]] = None

        while True:
            cycles_run += 1
            last_tick = self.tick(poller_spec)
            cycle_summaries.append({
                'cycle_number': cycles_run,
                'ok': last_tick.get('ok', False),
                'task_count': len(last_tick.get('tasks', [])),
                'notification_count': len(last_tick.get('notifications', [])),
                'errors': list(last_tick.get('errors', [])),
                'completed_at': last_tick.get('completed_at', ''),
            })

            if max_cycles > 0 and cycles_run >= max_cycles:
                break
            if interval_seconds > 0:
                sleep_fn(interval_seconds)

        return {
            'ok': all(item.get('ok', False) for item in cycle_summaries),
            'agent_id': 'gantt',
            'project_key': str(
                poller_spec.get('project_key') or self.project_key or ''
            ).strip(),
            'cycles_run': cycles_run,
            'cycle_summaries': cycle_summaries,
            'last_tick': last_tick or {},
        }

    def create_snapshot(self, request: PlanningRequest) -> PlanningSnapshot:
        '''
        Create a deterministic planning snapshot from Jira backlog data.
        '''
        log.info(f'Creating Gantt planning snapshot for {request.project_key}')

        project_info = self._load_project_info(request.project_key)
        releases = self._load_releases(request.project_key)
        issues = self._load_backlog_issues(request)
        evidence_bundle = load_evidence_bundle(request.evidence_paths)
        dependency_graph = self._build_dependency_graph(issues)
        milestones = self._build_milestones(issues, releases, dependency_graph)
        evidence_gaps = self._build_evidence_gaps(issues, evidence_bundle)
        risks = self._build_risks(issues, milestones, dependency_graph)
        backlog_overview = self._build_backlog_overview(
            issues,
            milestones,
            dependency_graph,
            risks,
        )

        snapshot = PlanningSnapshot(
            project_key=request.project_key,
            planning_horizon_days=request.planning_horizon_days,
            project_info=project_info,
            backlog_overview=backlog_overview,
            milestones=milestones,
            dependency_graph=dependency_graph,
            risks=risks,
            issues=issues,
            evidence_summary=evidence_bundle.to_summary(),
            evidence_gaps=evidence_gaps,
        )
        snapshot.summary_markdown = self._format_snapshot(snapshot)
        return snapshot

    def _load_project_info(self, project_key: str) -> Dict[str, Any]:
        result = get_project_info(project_key)
        if result.is_success:
            return result.data
        raise RuntimeError(result.error or f'Failed to load project info for {project_key}')

    def _load_releases(self, project_key: str) -> List[Dict[str, Any]]:
        result = get_releases(
            project_key,
            include_released=True,
            include_unreleased=True,
        )
        if result.is_success:
            return result.data
        raise RuntimeError(result.error or f'Failed to load releases for {project_key}')

    def _load_backlog_issues(self, request: PlanningRequest) -> List[Dict[str, Any]]:
        '''
        Query Jira directly so dependency-related fields remain available.
        '''
        issues = self.backlog_interpreter.load_backlog_issues(request)
        return self.dependency_mapper.attach_dependency_edges(
            issues,
            project_key=request.project_key,
        )

    @staticmethod
    def _build_backlog_jql(request: PlanningRequest) -> str:
        return BacklogInterpreter.build_backlog_jql(request)

    def _normalize_issue(self, issue: Any) -> Dict[str, Any]:
        normalized = self.backlog_interpreter.normalize_issue(issue)
        return self.dependency_mapper.attach_dependency_edges(
            [normalized],
            project_key=self.project_key,
        )[0]

    def _extract_edges(
        self,
        issue_key: str,
        parent_key: str,
        issue_links: Iterable[Any],
    ) -> List[DependencyEdge]:
        return self.dependency_mapper.extract_edges(issue_key, parent_key, issue_links)

    @staticmethod
    def _slugify_relationship(value: str) -> str:
        return DependencyMapper.slugify_relationship(value)

    @staticmethod
    def _parse_jira_datetime(value: str) -> Optional[datetime]:
        return BacklogInterpreter.parse_jira_datetime(value)

    @staticmethod
    def _is_done_status(status: str) -> bool:
        return BacklogInterpreter.is_done_status(status)

    @staticmethod
    def _is_high_priority(priority: str) -> bool:
        return BacklogInterpreter.is_high_priority(priority)

    @staticmethod
    def _is_blocked_status(status: str) -> bool:
        return DependencyMapper.is_blocked_status(status)

    def _build_dependency_graph(self, issues: List[Dict[str, Any]]) -> DependencyGraph:
        return self.dependency_mapper.build_graph(issues)

    def _build_milestones(
        self,
        issues: List[Dict[str, Any]],
        releases: List[Dict[str, Any]],
        dependency_graph: DependencyGraph,
    ) -> List[MilestoneProposal]:
        return self.milestone_planner.build_milestones(
            issues,
            releases,
            dependency_graph,
        )

    def _milestone_risk_level(
        self,
        open_issues: int,
        blocked_issues: int,
        unassigned_issues: int,
        release_date: Optional[str],
    ) -> str:
        return self.milestone_planner.milestone_risk_level(
            open_issues,
            blocked_issues,
            unassigned_issues,
            release_date,
        )

    @staticmethod
    def _milestone_sort_key(milestone: MilestoneProposal) -> tuple[int, str, str]:
        return MilestonePlanner.milestone_sort_key(milestone)

    def _build_evidence_gaps(
        self,
        issues: List[Dict[str, Any]],
        evidence_bundle: Optional[EvidenceBundle] = None,
    ) -> List[str]:
        return self.risk_projector.build_evidence_gaps(issues, evidence_bundle)

    def _build_risks(
        self,
        issues: List[Dict[str, Any]],
        milestones: List[MilestoneProposal],
        dependency_graph: DependencyGraph,
    ) -> List[PlanningRiskRecord]:
        return self.risk_projector.build_risks(issues, milestones, dependency_graph)

    def _build_backlog_overview(
        self,
        issues: List[Dict[str, Any]],
        milestones: List[MilestoneProposal],
        dependency_graph: DependencyGraph,
        risks: List[PlanningRiskRecord],
    ) -> Dict[str, Any]:
        return self.planning_summarizer.build_backlog_overview(
            issues,
            milestones,
            dependency_graph,
            risks,
        )

    def _format_snapshot(self, snapshot: PlanningSnapshot) -> str:
        return self.planning_summarizer.format_snapshot(snapshot)

    # ===========================================================================
    # Roadmap Analysis — Initiative/Epic/Story hierarchy with gap analysis
    # ===========================================================================

    def create_roadmap_snapshot(self, request: RoadmapRequest) -> RoadmapSnapshot:
        '''
        Create a roadmap snapshot from Jira data with optional gap analysis.

        Algorithm:
        1. Discover initiatives (explicit keys or JQL search).
        2. Build hierarchy — children of each initiative (epics, stories).
        3. Organize into RoadmapSection objects.
        4. Optionally run LLM gap analysis.
        5. Generate xlsx output.
        6. Return RoadmapSnapshot.
        '''
        log.info(
            f'Creating roadmap snapshot for {request.project_key} '
            f'(scope={request.scope_label})'
        )

        # Step 1: Discover initiatives
        initiative_keys = self._discover_initiatives(request)
        log.info(f'Discovered {len(initiative_keys)} initiatives')

        # Step 2+3: Build hierarchy and organize into sections
        sections = self._build_sections(request, initiative_keys)
        log.info(
            f'Built {len(sections)} sections with '
            f'{sum(len(s.items) for s in sections)} total items'
        )

        # Step 4: Gap analysis (LLM)
        if request.include_gap_analysis:
            self._run_gap_analysis(request, sections)
            total_gaps = sum(len(s.gaps) for s in sections)
            log.info(f'Gap analysis identified {total_gaps} proposed gaps')

        # Build snapshot
        snapshot = RoadmapSnapshot(
            project_key=request.project_key,
            scope_label=request.scope_label,
            sections=sections,
        )
        snapshot.summary_markdown = self._format_summary(snapshot)

        # Step 5: Generate xlsx
        if request.output_file:
            output_path = self._write_xlsx(snapshot, request.output_file, request)
            log.info(f'Wrote roadmap xlsx to {output_path}')

        return snapshot

    def _discover_initiatives(self, request: RoadmapRequest) -> List[str]:
        '''
        Discover initiative ticket keys — either from explicit list or JQL.

        If request.initiative_keys is provided, use those directly.
        Otherwise, search Jira for initiatives matching the scope_label.
        '''
        if request.initiative_keys:
            return list(request.initiative_keys)

        # Build JQL to find initiatives by scope label
        jql_parts = [
            f'project = {request.project_key}',
            'issuetype = Initiative',
        ]
        if request.scope_label:
            jql_parts.append(f'summary ~ "{request.scope_label}"')
        jql = ' AND '.join(jql_parts) + ' ORDER BY priority ASC'

        log.debug(f'Initiative discovery JQL: {jql}')

        try:
            result = search_tickets(jql=jql, limit=50)
            if not result.is_success:
                log.warning(f'Initiative search failed: {result.error}')
                return []
            return [ticket['key'] for ticket in result.data]
        except Exception as e:
            log.error(f'Initiative discovery failed: {e}')
            return []

    def _build_sections(
        self,
        request: RoadmapRequest,
        initiative_keys: List[str],
    ) -> List[RoadmapSection]:
        '''
        Build RoadmapSection objects from initiative keys.

        For each initiative, fetches the child hierarchy and organizes
        items into a section. Filters out Bugs and optionally closed items.
        '''
        sections: List[RoadmapSection] = []

        for init_key in initiative_keys:
            try:
                section = self._build_section_for_initiative(request, init_key)
                if section:
                    sections.append(section)
            except Exception as e:
                log.error(f'Failed to build section for {init_key}: {e}')

        return sections

    def _build_section_for_initiative(
        self,
        request: RoadmapRequest,
        initiative_key: str,
    ) -> Optional[RoadmapSection]:
        '''
        Build a single RoadmapSection for an initiative and its children.

        Uses get_children_hierarchy to fetch the full tree, then converts
        each Jira issue into a RoadmapItem.
        '''
        # Fetch the full hierarchy rooted at this initiative
        try:
            result = get_children_hierarchy(root_key=initiative_key, limit=500)
            if not result.is_success:
                log.warning(
                    f'get_children_hierarchy failed for {initiative_key}: '
                    f'{result.error}'
                )
                return None
        except Exception as e:
            log.error(f'Hierarchy fetch failed for {initiative_key}: {e}')
            return None

        hierarchy_items = result.data
        if not hierarchy_items:
            return None

        # The first item is the initiative itself — use its summary as title
        root = hierarchy_items[0]
        section_title = root.get('summary', initiative_key)

        items: List[RoadmapItem] = []
        for ticket in hierarchy_items:
            item = self._ticket_to_roadmap_item(ticket, section_title)
            if item is None:
                continue

            # Filter: exclude Bugs
            if item.issue_type in _EXCLUDED_TYPES:
                continue

            # Filter: exclude closed items unless requested
            if not request.include_closed and item.status in _DONE_STATUSES:
                continue

            # Filter: fix_versions if specified
            if request.fix_versions and item.fix_version:
                if not any(
                    fv in item.fix_version for fv in request.fix_versions
                ):
                    continue

            # Filter: respect hierarchy_depth
            if item.depth > request.hierarchy_depth:
                continue

            items.append(item)

        if not items:
            return None

        return RoadmapSection(title=section_title, items=items)

    def _ticket_to_roadmap_item(
        self,
        ticket: Dict[str, Any],
        section_title: str,
    ) -> Optional[RoadmapItem]:
        '''
        Convert a Jira ticket dict (from get_children_hierarchy or
        search_tickets) into a RoadmapItem.

        get_children_hierarchy returns: key, summary, type, status,
        priority, assignee, fix_versions (list), components (list),
        labels (list), depth.

        search_tickets returns: key, summary, issue_type/type, status,
        priority, assignee, fix_version (csv), component (csv),
        labels (list), etc.
        '''
        key = ticket.get('key', '')
        if not key:
            return None

        # Determine issue type — get_children_hierarchy uses 'type',
        # search_tickets uses 'issue_type' or 'type'
        issue_type = (
            ticket.get('issue_type')
            or ticket.get('type')
            or ''
        )

        # Determine depth — get_children_hierarchy provides it directly,
        # otherwise infer from issue type
        depth = ticket.get('depth')
        if depth is None:
            depth = _TYPE_DEPTH.get(issue_type, 2)

        # Handle fix_versions — may be a list or a csv string
        fix_versions_raw = ticket.get('fix_versions') or ticket.get('fix_version', '')
        if isinstance(fix_versions_raw, list):
            fix_version = ', '.join(fix_versions_raw)
        else:
            fix_version = str(fix_versions_raw)

        # Handle components — may be a list or a csv string
        components_raw = ticket.get('components') or ticket.get('component', '')
        if isinstance(components_raw, list):
            component = ', '.join(components_raw)
        else:
            component = str(components_raw)

        # Handle labels — may be a list or a csv string
        labels_raw = ticket.get('labels') or ticket.get('labels_csv', '')
        if isinstance(labels_raw, list):
            labels = ', '.join(labels_raw)
        else:
            labels = str(labels_raw)

        # Parent key — from search_tickets result or hierarchy context
        parent_key = ticket.get('parent_key', '')

        return RoadmapItem(
            key=key,
            summary=ticket.get('summary', ''),
            issue_type=issue_type,
            status=ticket.get('status', ''),
            priority=ticket.get('priority', ''),
            assignee=ticket.get('assignee', ''),
            fix_version=fix_version,
            component=component,
            labels=labels,
            parent_key=parent_key,
            depth=depth,
            source='Jira',
            section=section_title,
        )

    def _run_gap_analysis(
        self,
        request: RoadmapRequest,
        sections: List[RoadmapSection],
    ) -> None:
        '''
        Run LLM gap analysis on the current roadmap sections.

        Uses direct llm.chat() — _run_with_tools() sent tool schemas
        which caused the LLM to make tool calls instead of returning JSON.
        '''
        roadmap_text = self._build_gap_analysis_prompt(request, sections)

        system_prompt = (
            'You are a project planning analyst. '
            'Analyze the Jira roadmap provided by the user and identify gaps — '
            'missing Epics or Stories that should exist but do not yet have '
            'Jira tickets. Return ONLY a JSON block fenced with ```json and ```. '
            'Do not call any tools. Do not include commentary outside the JSON block.'
        )

        messages = [
            Message.system(system_prompt),
            Message.user(roadmap_text),
        ]

        log.info('Running LLM gap analysis (direct chat, no tools)...')
        try:
            llm_response = self.llm.chat(
                messages=messages,
                temperature=0.3,
                max_tokens=self.config.max_tokens,
            )

            gaps_json = self._extract_json_block(llm_response.content or '')
            if not gaps_json:
                log.warning(
                    'No JSON block found in gap analysis response '
                    f'({len(llm_response.content or "")} chars returned)'
                )
                return

            self._assign_gaps_to_sections(gaps_json, sections)
            gap_count = sum(len(s.gaps) for s in sections)
            log.info(f'Gap analysis complete — {gap_count} gaps identified')

        except Exception as e:
            log.error(f'Gap analysis failed: {e}')

    def _build_gap_analysis_prompt(
        self,
        request: RoadmapRequest,
        sections: List[RoadmapSection],
    ) -> str:
        '''Build the user prompt for LLM gap analysis.'''
        lines = [
            f'## Roadmap Gap Analysis Request',
            f'',
            f'**Project**: {request.project_key}',
            f'**Scope**: {request.scope_label or "All"}',
            f'',
            f'Analyze the following Jira roadmap and identify gaps — '
            f'missing Epics or Stories that should exist but do not yet '
            f'have Jira tickets. Return your analysis as a JSON block.',
            f'',
            f'### Current Roadmap',
            f'',
        ]

        for section in sections:
            lines.append(f'#### {section.title}')
            for item in section.items:
                indent = '  ' * item.depth
                status_tag = f' [{item.status}]' if item.status else ''
                lines.append(
                    f'{indent}- {item.key} ({item.issue_type}): '
                    f'{item.summary}{status_tag}'
                )
            lines.append('')

        lines.extend([
            '### Output Format',
            '',
            'Return a JSON block with this structure:',
            '```json',
            '{',
            '  "gaps": [',
            '    {',
            '      "section": "Section Title",',
            '      "summary": "Proposed ticket summary",',
            '      "issue_type": "Epic or Story",',
            '      "priority": "P0-Stopper | P1-Critical | P2-Major | P3-Minor",',
            '      "suggested_component": "component name",',
            '      "acceptance_criteria": "AC text",',
            '      "dependencies": "STL-123; STL-456",',
            '      "suggested_fix_version": "version string",',
            '      "labels": "label1, label2",',
            '      "depth": 1,',
            '      "parent_summary": "Parent epic summary (for stories)"',
            '    }',
            '  ]',
            '}',
            '```',
        ])

        return '\n'.join(lines)

    def _assign_gaps_to_sections(
        self,
        gaps_json: Dict[str, Any],
        sections: List[RoadmapSection],
    ) -> None:
        '''
        Convert parsed gap JSON into RoadmapGap objects and assign
        them to the matching section.
        '''
        # Build a section lookup by title
        section_map: Dict[str, RoadmapSection] = {
            s.title: s for s in sections
        }

        gaps_list = gaps_json.get('gaps', [])
        if isinstance(gaps_json, list):
            # LLM may have returned a bare list instead of {"gaps": [...]}
            gaps_list = gaps_json

        for gap_data in gaps_list:
            if not isinstance(gap_data, dict):
                continue

            gap = RoadmapGap(
                summary=gap_data.get('summary', ''),
                issue_type=gap_data.get('issue_type', 'Epic'),
                priority=gap_data.get('priority', 'P2-Major'),
                suggested_component=gap_data.get('suggested_component', ''),
                acceptance_criteria=gap_data.get('acceptance_criteria', ''),
                dependencies=gap_data.get('dependencies', ''),
                suggested_fix_version=gap_data.get('suggested_fix_version', ''),
                labels=gap_data.get('labels', ''),
                depth=int(gap_data.get('depth', 1)),
                section=gap_data.get('section', ''),
                parent_summary=gap_data.get('parent_summary', ''),
            )

            # Find the matching section — fall back to first section
            target_section = section_map.get(gap.section)
            if not target_section and sections:
                target_section = sections[0]
            if target_section:
                target_section.gaps.append(gap)

    # -- XLSX generation — 3-sheet workbook ------------------------------------

    @staticmethod
    def _col(headers: List[str], name: str) -> int:
        '''Return 1-based column index for *name* in *headers*.'''
        return headers.index(name) + 1

    def _write_xlsx(
        self,
        snapshot: RoadmapSnapshot,
        output_file: str,
        request: Optional[RoadmapRequest] = None,
    ) -> str:
        '''
        Generate a 3-sheet xlsx workbook from the roadmap snapshot.

        Sheet 1: "Current Jira" — all Jira-sourced items
        Sheet 2: "Proposed (Gaps)" — LLM-identified gaps
        Sheet 3: "Merged Roadmap" — both interleaved

        Uses excel_utils formatting helpers for consistent styling.
        '''
        if request is None:
            request = RoadmapRequest()

        wb = Workbook()

        # -- Sheet 1: Current Jira --
        ws_jira = wb.active or wb.create_sheet()
        ws_jira.title = 'Current Jira'
        self._write_jira_sheet(ws_jira, snapshot, request)

        # -- Sheet 2: Proposed (Gaps) --
        ws_gaps = wb.create_sheet('Proposed (Gaps)')
        self._write_gaps_sheet(ws_gaps, snapshot, request)

        # -- Sheet 3: Merged Roadmap --
        ws_merged = wb.create_sheet('Merged Roadmap')
        self._write_merged_sheet(ws_merged, snapshot, request)

        # Save workbook
        os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
        wb.save(output_file)
        return output_file

    def _write_jira_sheet(
        self,
        ws: Any,
        snapshot: RoadmapSnapshot,
        request: RoadmapRequest,
    ) -> None:
        '''Write the "Current Jira" sheet with all Jira-sourced items.'''
        headers = [
            'Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)',
            'project', 'issue_type', 'status', 'priority', 'key',
            'summary', 'assignee', 'fix_version', 'component', 'labels',
        ]
        if not request.show_priority:
            headers.remove('priority')

        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        row = 2
        for section in snapshot.sections:
            row = self._write_section_divider(ws, row, section.title, len(headers))

            for item in section.items:
                if item.source != 'Jira':
                    continue
                self._write_jira_item_row(
                    ws, row, item, snapshot.project_key, headers, request,
                )
                row += 1

        _apply_header_style(ws, len(headers))
        _apply_status_conditional_formatting(ws, headers)
        if request.show_priority:
            _apply_priority_conditional_formatting(ws, headers)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(row - 1, 1)}'
        self._set_column_widths(ws, headers)

    def _write_jira_item_row(
        self,
        ws: Any,
        row: int,
        item: RoadmapItem,
        project_key: str,
        headers: List[str],
        request: RoadmapRequest,
    ) -> None:
        '''Write a single Jira item row to the Current Jira sheet.'''
        c = self._col
        for d in range(3):
            depth_header = ['Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)'][d]
            if item.depth == d:
                ws.cell(row=row, column=c(headers, depth_header), value=item.key)
            else:
                ws.cell(row=row, column=c(headers, depth_header), value='')

        ws.cell(row=row, column=c(headers, 'project'), value=project_key)

        issue_type_cell = ws.cell(
            row=row, column=c(headers, 'issue_type'), value=item.issue_type,
        )
        if request.bold_stories and item.issue_type == 'Story':
            issue_type_cell.font = Font(bold=True)

        ws.cell(row=row, column=c(headers, 'status'), value=item.status)

        if 'priority' in headers:
            ws.cell(row=row, column=c(headers, 'priority'), value=item.priority)

        key_cell = ws.cell(row=row, column=c(headers, 'key'), value=item.key)
        if item.key:
            key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{item.key}'
            key_cell.font = Font(color='0563C1', underline='single')

        ws.cell(row=row, column=c(headers, 'summary'), value=item.summary)

        assignee_val = item.assignee
        if request.blank_unassigned and assignee_val in ('Unassigned', '', None):
            assignee_val = ''
        ws.cell(row=row, column=c(headers, 'assignee'), value=assignee_val)

        ws.cell(row=row, column=c(headers, 'fix_version'), value=item.fix_version)
        ws.cell(row=row, column=c(headers, 'component'), value=item.component)
        ws.cell(row=row, column=c(headers, 'labels'), value=item.labels)

        if item.depth == 0:
            bold_font = Font(bold=True)
            key_col = c(headers, 'key')
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col_idx)
                if col_idx == key_col and item.key:
                    cell.font = Font(
                        bold=True, color='0563C1', underline='single'
                    )
                else:
                    cell.font = bold_font

    def _write_gaps_sheet(
        self,
        ws: Any,
        snapshot: RoadmapSnapshot,
        request: RoadmapRequest,
    ) -> None:
        '''Write the "Proposed (Gaps)" sheet with LLM-identified gaps.'''
        headers = [
            'Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)',
            'project', 'issue_type', 'priority', 'summary',
            'suggested_component', 'acceptance_criteria', 'dependencies',
            'suggested_fix_version', 'labels',
        ]
        if not request.show_priority:
            headers.remove('priority')

        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        row = 2
        for section in snapshot.sections:
            if not section.gaps:
                continue

            row = self._write_section_divider(ws, row, section.title, len(headers))

            for gap in section.gaps:
                self._write_gap_row(
                    ws, row, gap, snapshot.project_key, headers, request,
                )
                row += 1

        _apply_header_style(ws, len(headers))
        if request.show_priority:
            _apply_priority_conditional_formatting(ws, headers)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(row - 1, 1)}'
        self._set_column_widths(ws, headers, is_gaps_sheet=True)

    def _write_gap_row(
        self,
        ws: Any,
        row: int,
        gap: RoadmapGap,
        project_key: str,
        headers: List[str],
        request: RoadmapRequest,
    ) -> None:
        '''Write a single gap row to the Proposed (Gaps) sheet.'''
        c = self._col
        for d in range(3):
            depth_header = ['Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)'][d]
            if gap.depth == d:
                ws.cell(row=row, column=c(headers, depth_header), value=gap.summary[:40])
            else:
                ws.cell(row=row, column=c(headers, depth_header), value='')

        ws.cell(row=row, column=c(headers, 'project'), value=project_key)

        issue_type_cell = ws.cell(
            row=row, column=c(headers, 'issue_type'), value=gap.issue_type,
        )
        if request.bold_stories and gap.issue_type == 'Story':
            issue_type_cell.font = Font(bold=True)

        if 'priority' in headers:
            ws.cell(row=row, column=c(headers, 'priority'), value=gap.priority)

        ws.cell(row=row, column=c(headers, 'summary'), value=gap.summary)
        ws.cell(row=row, column=c(headers, 'suggested_component'), value=gap.suggested_component)
        ws.cell(row=row, column=c(headers, 'acceptance_criteria'), value=gap.acceptance_criteria)
        ws.cell(row=row, column=c(headers, 'dependencies'), value=gap.dependencies)
        ws.cell(row=row, column=c(headers, 'suggested_fix_version'), value=gap.suggested_fix_version)
        ws.cell(row=row, column=c(headers, 'labels'), value=gap.labels)

    def _write_merged_sheet(
        self,
        ws: Any,
        snapshot: RoadmapSnapshot,
        request: RoadmapRequest,
    ) -> None:
        '''Write the "Merged Roadmap" sheet with both Jira and proposed items.'''
        headers = [
            'Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)',
            'project', 'issue_type', 'status', 'priority', 'key',
            'summary', 'assignee', 'fix_version', 'component', 'labels',
            'source', 'acceptance_criteria', 'dependencies',
        ]
        if not request.show_priority:
            headers.remove('priority')

        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        proposed_fill = PatternFill(
            start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'
        )

        row = 2
        for section in snapshot.sections:
            row = self._write_section_divider(ws, row, section.title, len(headers))

            for item in section.items:
                if item.source != 'Jira':
                    continue
                self._write_merged_jira_row(
                    ws, row, item, snapshot.project_key, headers, request,
                )
                row += 1

            for gap in section.gaps:
                self._write_merged_gap_row(
                    ws, row, gap, snapshot.project_key, proposed_fill,
                    headers, request,
                )
                row += 1

        _apply_header_style(ws, len(headers))
        _apply_status_conditional_formatting(ws, headers)
        if request.show_priority:
            _apply_priority_conditional_formatting(ws, headers)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(row - 1, 1)}'
        self._set_column_widths(ws, headers, is_merged_sheet=True)

    def _write_merged_jira_row(
        self,
        ws: Any,
        row: int,
        item: RoadmapItem,
        project_key: str,
        headers: List[str],
        request: RoadmapRequest,
    ) -> None:
        '''Write a Jira item row to the Merged Roadmap sheet.'''
        c = self._col
        for d in range(3):
            depth_header = ['Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)'][d]
            if item.depth == d:
                ws.cell(row=row, column=c(headers, depth_header), value=item.key)
            else:
                ws.cell(row=row, column=c(headers, depth_header), value='')

        ws.cell(row=row, column=c(headers, 'project'), value=project_key)

        issue_type_cell = ws.cell(
            row=row, column=c(headers, 'issue_type'), value=item.issue_type,
        )
        if request.bold_stories and item.issue_type == 'Story':
            issue_type_cell.font = Font(bold=True)

        ws.cell(row=row, column=c(headers, 'status'), value=item.status)

        if 'priority' in headers:
            ws.cell(row=row, column=c(headers, 'priority'), value=item.priority)

        key_cell = ws.cell(row=row, column=c(headers, 'key'), value=item.key)
        if item.key:
            key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{item.key}'
            key_cell.font = Font(color='0563C1', underline='single')

        ws.cell(row=row, column=c(headers, 'summary'), value=item.summary)

        assignee_val = item.assignee
        if request.blank_unassigned and assignee_val in ('Unassigned', '', None):
            assignee_val = ''
        ws.cell(row=row, column=c(headers, 'assignee'), value=assignee_val)

        ws.cell(row=row, column=c(headers, 'fix_version'), value=item.fix_version)
        ws.cell(row=row, column=c(headers, 'component'), value=item.component)
        ws.cell(row=row, column=c(headers, 'labels'), value=item.labels)
        ws.cell(row=row, column=c(headers, 'source'), value='Jira')
        ws.cell(row=row, column=c(headers, 'acceptance_criteria'), value='')
        ws.cell(row=row, column=c(headers, 'dependencies'), value='')

        if item.depth == 0:
            bold_font = Font(bold=True)
            key_col = c(headers, 'key')
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col_idx)
                if col_idx == key_col and item.key:
                    cell.font = Font(
                        bold=True, color='0563C1', underline='single'
                    )
                else:
                    cell.font = bold_font

    def _write_merged_gap_row(
        self,
        ws: Any,
        row: int,
        gap: RoadmapGap,
        project_key: str,
        proposed_fill: PatternFill,
        headers: List[str],
        request: RoadmapRequest,
    ) -> None:
        '''Write a proposed gap row to the Merged Roadmap sheet (highlighted).'''
        c = self._col
        for d in range(3):
            depth_header = ['Depth 0 (Initiative)', 'Depth 1 (Epic)', 'Depth 2 (Story)'][d]
            if gap.depth == d:
                ws.cell(row=row, column=c(headers, depth_header), value=gap.summary[:40])
            else:
                ws.cell(row=row, column=c(headers, depth_header), value='')

        ws.cell(row=row, column=c(headers, 'project'), value=project_key)

        issue_type_cell = ws.cell(
            row=row, column=c(headers, 'issue_type'), value=gap.issue_type,
        )
        if request.bold_stories and gap.issue_type == 'Story':
            issue_type_cell.font = Font(bold=True)

        ws.cell(row=row, column=c(headers, 'status'), value='')

        if 'priority' in headers:
            ws.cell(row=row, column=c(headers, 'priority'), value=gap.priority)

        ws.cell(row=row, column=c(headers, 'key'), value='')
        ws.cell(row=row, column=c(headers, 'summary'), value=gap.summary)
        ws.cell(row=row, column=c(headers, 'assignee'), value='')
        ws.cell(row=row, column=c(headers, 'fix_version'), value=gap.suggested_fix_version)
        ws.cell(row=row, column=c(headers, 'component'), value=gap.suggested_component)
        ws.cell(row=row, column=c(headers, 'labels'), value=gap.labels)
        ws.cell(row=row, column=c(headers, 'source'), value='Proposed')
        ws.cell(row=row, column=c(headers, 'acceptance_criteria'), value=gap.acceptance_criteria)
        ws.cell(row=row, column=c(headers, 'dependencies'), value=gap.dependencies)

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row, column=col_idx).fill = proposed_fill

    @staticmethod
    def _write_section_divider(
        ws: Any,
        row: int,
        title: str,
        num_cols: int,
    ) -> int:
        '''
        Write a section divider row — bold white text on dark blue fill.

        Returns the next row number (row + 1).
        '''
        divider_font = Font(bold=True, color='FFFFFF')
        divider_fill = PatternFill(
            start_color='2F5496', end_color='2F5496', fill_type='solid'
        )

        ws.cell(row=row, column=1, value=title)
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = divider_font
            cell.fill = divider_fill

        return row + 1

    @staticmethod
    def _set_column_widths(
        ws: Any,
        headers: List[str],
        is_gaps_sheet: bool = False,
        is_merged_sheet: bool = False,
    ) -> None:
        '''
        Set explicit column widths for the worksheet.

        Column widths are tuned for readability in each sheet type.
        Uses header-name lookup so widths stay correct when columns are
        added or removed (e.g. priority omitted).
        '''
        _BASE: Dict[str, int] = {
            'Depth 0 (Initiative)': 42,
            'Depth 1 (Epic)': 42,
            'Depth 2 (Story)': 48,
            'project': 7,
            'issue_type': 12,
            'status': 12,
            'priority': 6,
            'key': 12,
            'summary': 65,
            'assignee': 20,
            'fix_version': 18,
            'component': 30,
            'labels': 14,
            'source': 50,
            'acceptance_criteria': 25,
            'dependencies': 25,
        }

        _GAPS: Dict[str, int] = {
            'summary': 65,
            'suggested_component': 25,
            'acceptance_criteria': 50,
            'dependencies': 25,
            'suggested_fix_version': 18,
            'labels': 25,
        }

        width_map = {**_BASE, **(_GAPS if is_gaps_sheet else {})}

        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width_map.get(header, 14)

    @staticmethod
    def _format_summary(snapshot: RoadmapSnapshot) -> str:
        '''Build a human-readable Markdown summary of the roadmap snapshot.'''
        lines = [
            f'# Roadmap Analysis: {snapshot.project_key}',
            f'',
            f'**Scope**: {snapshot.scope_label or "All"}',
            f'**Snapshot ID**: {snapshot.snapshot_id}',
            f'**Created**: {snapshot.created_at}',
            f'**Jira Items**: {snapshot.total_jira_items}',
            f'**Proposed Gaps**: {snapshot.total_proposed_gaps}',
            f'**Total Items**: {snapshot.total_items}',
            f'',
            '---',
            '',
        ]

        for section in snapshot.sections:
            lines.append(f'## {section.title}')
            lines.append(f'  Items: {len(section.items)}, Gaps: {len(section.gaps)}')
            lines.append('')

            for item in section.items:
                indent = '  ' * item.depth
                lines.append(
                    f'{indent}- [{item.key}] {item.summary} '
                    f'({item.status}, {item.priority})'
                )

            if section.gaps:
                lines.append('')
                lines.append('  **Proposed Gaps:**')
                for gap in section.gaps:
                    lines.append(
                        f'  - [{gap.issue_type}] {gap.summary} '
                        f'({gap.priority})'
                    )

            lines.append('')

        return '\n'.join(lines)

    # ===========================================================================
    # Release Monitor — Bug tracking, velocity, readiness, and gap analysis
    # ===========================================================================

    @staticmethod
    def _serialize_release_snapshot(snapshot: ReleaseSnapshot) -> Dict[str, Any]:
        return {
            'release': snapshot.release,
            'timestamp': snapshot.timestamp,
            'total_tickets': snapshot.total_tickets,
            'tickets': list(snapshot.tickets),
        }

    @staticmethod
    def _deserialize_release_snapshot(data: Optional[Dict[str, Any]]) -> Optional[ReleaseSnapshot]:
        if not data:
            return None

        release = str(data.get('release') or '').strip()
        timestamp = str(data.get('timestamp') or '').strip()
        tickets = list(data.get('tickets') or [])
        total_tickets = int(data.get('total_tickets') or len(tickets))

        if not release or not timestamp:
            return None

        return ReleaseSnapshot(
            release=release,
            timestamp=timestamp,
            total_tickets=total_tickets,
            tickets=tickets,
        )

    @staticmethod
    def _build_bug_release_snapshot(snapshot: ReleaseSnapshot) -> ReleaseSnapshot:
        bug_tickets = [
            dict(ticket)
            for ticket in snapshot.tickets
            if str(ticket.get('issuetype', '')).lower() == 'bug'
        ]
        return ReleaseSnapshot(
            release=snapshot.release,
            timestamp=snapshot.timestamp,
            total_tickets=len(bug_tickets),
            tickets=bug_tickets,
        )

    @staticmethod
    def _cycle_time_sample_key(sample: Dict[str, Any]) -> tuple[str, str, str, str]:
        return (
            str(sample.get('release') or ''),
            str(sample.get('ticket_key') or ''),
            str(sample.get('component') or ''),
            str(sample.get('priority') or ''),
        )

    def _merge_cycle_time_samples(
        self,
        historical_samples: List[Dict[str, Any]],
        current_samples: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        '''
        Merge historical/current samples without re-counting previously closed tickets.
        '''
        merged: Dict[tuple[str, str, str, str], Dict[str, Any]] = {}

        for sample in historical_samples:
            if not isinstance(sample, dict):
                continue
            merged[self._cycle_time_sample_key(sample)] = dict(sample)

        for sample in current_samples:
            if not isinstance(sample, dict):
                continue
            merged[self._cycle_time_sample_key(sample)] = dict(sample)

        return list(merged.values())

    def _build_cycle_time_samples(
        self,
        release: str,
        tickets: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        '''
        Estimate cycle-time samples from closed tickets using created/resolved dates.
        '''
        samples: List[Dict[str, Any]] = []

        for ticket in tickets:
            created_dt = self._parse_jira_datetime(str(ticket.get('created', '') or ''))
            resolved_dt = self._parse_jira_datetime(
                str(ticket.get('resolutiondate', '') or '')
            )
            if not created_dt or not resolved_dt or resolved_dt < created_dt:
                continue

            duration_hours = round(
                (resolved_dt - created_dt).total_seconds() / 3600.0,
                2,
            )
            if duration_hours < 0:
                continue

            components_raw = ticket.get('components') or []
            if isinstance(components_raw, str):
                components = [
                    item.strip() for item in components_raw.split(',')
                    if item.strip()
                ]
            elif isinstance(components_raw, list):
                components = [
                    str(item).strip() for item in components_raw
                    if str(item).strip()
                ]
            else:
                components = []

            if not components:
                components = ['Unspecified']

            for component in components:
                samples.append({
                    'release': release,
                    'ticket_key': str(ticket.get('key') or ''),
                    'component': component,
                    'priority': str(ticket.get('priority') or 'Unknown'),
                    'duration_hours': duration_hours,
                })

        return samples

    def _build_cycle_time_stats(
        self,
        snapshot: ReleaseSnapshot,
        cycle_time_samples: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        '''
        Build component/priority cycle-time statistics for readiness assessment.
        '''
        relevant_samples = [
            dict(item)
            for item in cycle_time_samples
            if str(item.get('release') or snapshot.release) == snapshot.release
        ]
        components = list(snapshot.by_component.keys()) or ['Unspecified']
        tracked_priorities = ['P0-Stopper', 'P1-Critical']
        stats: List[Dict[str, Any]] = []

        for component in components:
            for priority in tracked_priorities:
                stat = compute_cycle_time_stats(
                    relevant_samples,
                    component=component,
                    priority=priority,
                )
                if stat.sample_size <= 0:
                    continue
                stats.append({
                    'release': snapshot.release,
                    'component': stat.component,
                    'priority': stat.priority,
                    'avg_hours': stat.avg_hours,
                    'median_hours': stat.median_hours,
                    'sample_size': stat.sample_size,
                })

        return stats

    def _build_velocity_summary(
        self,
        release_names: List[str],
        current_snapshots: Dict[str, ReleaseSnapshot],
        previous_snapshots: Dict[str, ReleaseSnapshot],
    ) -> Optional[Dict[str, Any]]:
        if not current_snapshots:
            return None

        by_release: Dict[str, Dict[str, Any]] = {}
        total_opened = 0
        total_closed = 0
        total_daily_open = 0.0
        total_daily_close = 0.0
        total_daily_net = 0.0
        total_snapshots_used = 0

        for release in release_names:
            current_snapshot = current_snapshots.get(release)
            if current_snapshot is None:
                continue

            history: List[ReleaseSnapshot] = []
            previous_snapshot = previous_snapshots.get(release)
            if previous_snapshot is not None:
                history.append(previous_snapshot)
            history.append(current_snapshot)

            velocity = compute_velocity(history, window_days=14)
            by_release[release] = velocity
            total_opened += int(velocity.get('opened', 0) or 0)
            total_closed += int(velocity.get('closed', 0) or 0)
            total_daily_open += float(velocity.get('daily_open_rate', 0.0) or 0.0)
            total_daily_close += float(velocity.get('daily_close_rate', 0.0) or 0.0)
            total_daily_net += float(velocity.get('daily_net_rate', 0.0) or 0.0)
            total_snapshots_used += int(velocity.get('snapshots_used', 0) or 0)

        if not by_release:
            return None

        if len(by_release) == 1:
            return next(iter(by_release.values()))

        return {
            'by_release': by_release,
            'snapshots_used': total_snapshots_used,
            'opened': total_opened,
            'closed': total_closed,
            'daily_open_rate': total_daily_open,
            'daily_close_rate': total_daily_close,
            'daily_net_rate': total_daily_net,
        }

    def create_release_monitor(
        self,
        request: ReleaseMonitorRequest,
    ) -> ReleaseMonitorReport:
        '''
        Create a release health monitoring report.

        Algorithm:
        1. Discover releases (explicit or unreleased versions from Jira).
        2. Query all tickets per release; build bug summaries.
        3. Build ReleaseSnapshots via core.release_tracking.
        4. Optionally compute deltas, velocity, readiness, and gap analysis.
        5. Generate xlsx output.
        6. Return ReleaseMonitorReport.
        '''
        project_key = request.project_key or self.project_key or 'STL'
        self.project_key = project_key
        log.info(
            f'Creating release monitor for {project_key} '
            f'(releases={request.releases}, scope={request.scope_label})'
        )

        # Step 1: Discover releases to monitor
        if request.releases:
            release_names = list(request.releases)
        else:
            # Query Jira for unreleased versions
            result = get_releases(
                project_key,
                include_released=False,
                include_unreleased=True,
            )
            if result.is_success:
                release_names = [
                    v.get('name', '') for v in result.data
                    if not v.get('released', False) and v.get('name')
                ]
            else:
                log.warning(f'Failed to fetch releases: {result.error}')
                release_names = []

        if not release_names:
            log.warning('No releases found to monitor')

        # Step 2: Query tickets and build snapshots/bug summaries per release
        all_tickets_by_release: Dict[str, List[Dict[str, Any]]] = {}
        bug_summaries: List[BugSummary] = []
        current_snapshots: Dict[str, ReleaseSnapshot] = {}
        current_cycle_time_samples: List[Dict[str, Any]] = []

        for release in release_names:
            tickets = self._query_release_tickets(release, request.scope_label)
            all_tickets_by_release[release] = tickets

            mapping_tickets: list[Mapping[str, Any]] = list(tickets)
            snapshot = build_release_snapshot(mapping_tickets, release)
            current_snapshots[release] = snapshot
            current_cycle_time_samples.extend(
                self._build_cycle_time_samples(release, tickets)
            )

            bug_summary = self._build_bug_summary(release, tickets)
            bug_summaries.append(bug_summary)

        # Step 3: Load the most recent compatible stored report
        previous_report_data: Optional[Dict[str, Any]] = None
        previous_snapshots: Dict[str, ReleaseSnapshot] = {}
        historical_cycle_time_samples: List[Dict[str, Any]] = []
        if request.compare_to_previous and release_names:
            try:
                previous_record = GanttReleaseMonitorStore().get_latest_compatible_report(
                    project_key=project_key,
                    releases=release_names,
                    scope_label=request.scope_label,
                )
                if previous_record:
                    previous_report_data = previous_record.get('report') or {}
                    historical_cycle_time_samples = [
                        dict(item)
                        for item in (previous_report_data.get('cycle_time_samples') or [])
                        if isinstance(item, dict)
                    ]
                    for release in release_names:
                        snapshot = self._deserialize_release_snapshot(
                            (previous_report_data.get('release_snapshots') or {}).get(release)
                        )
                        if snapshot is not None:
                            previous_snapshots[release] = snapshot
            except Exception as e:
                log.warning(f'Previous report lookup failed: {e}')

        # Step 4: Delta comparison with the previous compatible report
        delta_data: Optional[Dict[str, Any]] = None
        if previous_snapshots:
            try:
                delta_by_release: Dict[str, Dict[str, Any]] = {}
                flat_new: List[Dict[str, str]] = []
                flat_closed: List[Dict[str, str]] = []
                flat_status_changes: List[Dict[str, str]] = []
                flat_priority_changes: List[Dict[str, str]] = []
                flat_new_p0_p1: List[Dict[str, str]] = []
                total_velocity = 0.0

                for bug_summary in bug_summaries:
                    current_snapshot = current_snapshots.get(bug_summary.release)
                    previous_snapshot = previous_snapshots.get(bug_summary.release)
                    if current_snapshot is None or previous_snapshot is None:
                        continue

                    delta = compute_delta(
                        self._build_bug_release_snapshot(current_snapshot),
                        self._build_bug_release_snapshot(previous_snapshot),
                    )
                    delta_entry = {
                        'release': delta.release,
                        'new_tickets': list(delta.new_tickets),
                        'closed_tickets': list(delta.closed_tickets),
                        'status_changes': list(delta.status_changes),
                        'priority_changes': list(delta.priority_changes),
                        'new_p0_p1': list(delta.new_p0_p1),
                        'velocity': delta.velocity,
                    }
                    delta_by_release[bug_summary.release] = delta_entry
                    bug_summary.new_since_last = list(delta.new_tickets)
                    bug_summary.closed_since_last = list(delta.closed_tickets)
                    bug_summary.priority_changes = list(delta.priority_changes)

                    flat_new.extend(
                        {'release': delta.release, 'key': key}
                        for key in delta.new_tickets
                    )
                    flat_closed.extend(
                        {'release': delta.release, 'key': key}
                        for key in delta.closed_tickets
                    )
                    flat_status_changes.extend(
                        {
                            'release': delta.release,
                            'key': str(change.get('key', '')),
                            'from': str(change.get('from', '')),
                            'to': str(change.get('to', '')),
                        }
                        for change in delta.status_changes
                    )
                    flat_priority_changes.extend(
                        {
                            'release': delta.release,
                            'key': str(change.get('key', '')),
                            'from': str(change.get('from', '')),
                            'to': str(change.get('to', '')),
                        }
                        for change in delta.priority_changes
                    )
                    flat_new_p0_p1.extend(
                        {'release': delta.release, 'key': key}
                        for key in delta.new_p0_p1
                    )
                    total_velocity += float(delta.velocity or 0.0)

                if delta_by_release:
                    primary_release = release_names[0] if len(release_names) == 1 else 'multiple'
                    delta_data = {
                        'comparison_basis': 'previous_report',
                        'previous_report_id': str(
                            previous_report_data.get('report_id') or ''
                        ) if previous_report_data else '',
                        'release': primary_release,
                        'new_tickets': flat_new,
                        'closed_tickets': flat_closed,
                        'status_changes': flat_status_changes,
                        'priority_changes': flat_priority_changes,
                        'new_p0_p1': flat_new_p0_p1,
                        'velocity': total_velocity,
                        'by_release': delta_by_release,
                    }
            except Exception as e:
                log.warning(f'Delta computation failed: {e}')

        # Step 5: Velocity
        velocity_data: Optional[Dict[str, Any]] = None
        if request.include_velocity and current_snapshots:
            try:
                velocity_data = self._build_velocity_summary(
                    release_names,
                    current_snapshots,
                    previous_snapshots,
                )
            except Exception as e:
                log.warning(f'Velocity computation failed: {e}')

        # Step 6: Readiness assessment
        readiness_data: Optional[Dict[str, Any]] = None
        cycle_time_samples = self._merge_cycle_time_samples(
            historical_cycle_time_samples,
            current_cycle_time_samples,
        )
        cycle_time_stats: List[Dict[str, Any]] = []
        if request.include_readiness and current_snapshots:
            try:
                primary_release = release_names[0] if release_names else ''
                primary_snapshot = current_snapshots.get(primary_release)
                if primary_snapshot is not None:
                    config = TrackerConfig(
                        project=project_key,
                        releases=release_names,
                    )
                    cycle_time_stats = self._build_cycle_time_stats(
                        primary_snapshot,
                        cycle_time_samples,
                    )
                    velocity_for_readiness = velocity_data or {}
                    if 'by_release' in velocity_for_readiness:
                        velocity_for_readiness = (
                            velocity_for_readiness.get('by_release') or {}
                        ).get(primary_release, {})
                    readiness_report = assess_readiness(
                        primary_snapshot,
                        velocity_for_readiness,
                        cycle_time_stats,
                        config,
                    )
                    readiness_data = {
                        'release': readiness_report.release,
                        'total_open': readiness_report.total_open,
                        'p0_open': readiness_report.p0_open,
                        'p1_open': readiness_report.p1_open,
                        'daily_close_rate': readiness_report.daily_close_rate,
                        'estimated_days_remaining': readiness_report.estimated_days_remaining,
                        'stale_tickets': readiness_report.stale_tickets,
                        'component_risks': readiness_report.component_risks,
                    }
            except Exception as e:
                log.warning(f'Readiness assessment failed: {e}')

        # Step 7: Gap analysis (reuse roadmap snapshot machinery)
        roadmap_data: Optional[Dict[str, Any]] = None
        if request.include_gap_analysis:
            try:
                roadmap_request = RoadmapRequest(
                    project_key=project_key,
                    scope_label=request.scope_label or '',
                    fix_versions=release_names,
                    include_gap_analysis=True,
                )
                roadmap_snap = self.create_roadmap_snapshot(roadmap_request)
                roadmap_data = roadmap_snap.to_dict()
            except Exception as e:
                log.warning(f'Gap analysis failed: {e}')

        # Build the report
        report = ReleaseMonitorReport(
            project_key=project_key,
            created_at=datetime.now(timezone.utc).isoformat(),
            scope_label=request.scope_label or '',
            releases_monitored=release_names,
            bug_summaries=bug_summaries,
            velocity=velocity_data,
            readiness=readiness_data,
            roadmap_snapshot=roadmap_data,
            delta=delta_data,
            release_snapshots={
                release: self._serialize_release_snapshot(snapshot)
                for release, snapshot in current_snapshots.items()
            },
            cycle_time_samples=cycle_time_samples,
            cycle_time_stats=cycle_time_stats,
        )

        # Generate markdown summary
        report.summary_markdown = self._format_release_monitor_summary(report)

        # Step 8: Generate xlsx
        if request.output_file:
            output_path = self._write_release_monitor_xlsx(
                report, request.output_file, all_tickets_by_release,
            )
            report.output_file = output_path
            log.info(f'Wrote release monitor xlsx to {output_path}')

        return report

    def create_release_survey(
        self,
        request: ReleaseSurveyRequest,
    ) -> ReleaseSurveyReport:
        '''
        Create a release execution survey focused on done, active, and remaining work.
        '''
        project_key = request.project_key or self.project_key or 'STL'
        self.project_key = project_key
        survey_mode = self._normalize_release_survey_mode(request.survey_mode)
        log.info(
            f'Creating release survey for {project_key} '
            f'(releases={request.releases}, scope={request.scope_label}, '
            f'survey_mode={survey_mode})'
        )

        if request.releases:
            release_names = list(request.releases)
        else:
            result = get_releases(
                project_key,
                include_released=False,
                include_unreleased=True,
            )
            if result.is_success:
                release_names = [
                    item.get('name', '')
                    for item in result.data
                    if not item.get('released', False) and item.get('name')
                ]
            else:
                log.warning(f'Failed to fetch releases for survey: {result.error}')
                release_names = []

        release_summaries: List[ReleaseSurveyReleaseSummary] = []
        for release in release_names:
            tickets = self._query_release_tickets(release, request.scope_label)
            tickets = [
                ticket
                for ticket in tickets
                if self._ticket_matches_release_survey_mode(ticket, survey_mode)
            ]
            classified = [
                self._classify_release_survey_ticket(release, ticket)
                for ticket in tickets
            ]
            family_names = self._release_survey_family_names(
                request.scope_label,
                classified,
            )
            family_breakdowns = self._build_release_survey_family_breakdowns(
                family_names,
                classified,
            )
            hierarchy_cache: Dict[str, Dict[str, Any]] = {}
            family_epic_analysis = self._build_release_survey_family_epic_analysis(
                family_names,
                family_breakdowns,
                survey_mode,
                hierarchy_cache=hierarchy_cache,
            )

            by_status: Counter[str] = Counter()
            by_priority: Counter[str] = Counter()
            by_issue_type: Counter[str] = Counter()
            by_component: Counter[str] = Counter()
            done_tickets: List[Dict[str, Any]] = []
            in_progress_tickets: List[Dict[str, Any]] = []
            remaining_tickets: List[Dict[str, Any]] = []
            blocked_tickets: List[Dict[str, Any]] = []
            stale_tickets: List[str] = []
            unassigned_tickets: List[str] = []

            for ticket in classified:
                by_status[str(ticket.get('status') or 'Unknown')] += 1
                by_priority[str(ticket.get('priority') or 'Unspecified')] += 1
                by_issue_type[str(ticket.get('issue_type') or 'Unknown')] += 1

                components = ticket.get('components') or ['Unspecified']
                for component in components:
                    by_component[str(component or 'Unspecified')] += 1

                bucket = str(ticket.get('bucket') or 'remaining')
                if bucket == 'done':
                    done_tickets.append(ticket)
                elif bucket == 'in_progress':
                    in_progress_tickets.append(ticket)
                elif bucket == 'blocked':
                    blocked_tickets.append(ticket)
                else:
                    remaining_tickets.append(ticket)

                if ticket.get('is_stale'):
                    stale_tickets.append(str(ticket.get('key') or ''))
                if ticket.get('is_unassigned'):
                    unassigned_tickets.append(str(ticket.get('key') or ''))

            release_summaries.append(
                ReleaseSurveyReleaseSummary(
                    release=release,
                    total_tickets=len(classified),
                    status_breakdown=dict(by_status),
                    priority_breakdown=dict(by_priority),
                    issue_type_breakdown=dict(by_issue_type),
                    component_breakdown=dict(by_component),
                    family_breakdowns=family_breakdowns,
                    family_epic_analysis=family_epic_analysis,
                    done_tickets=self._sort_release_survey_bucket(
                        done_tickets,
                        bucket='done',
                    ),
                    in_progress_tickets=self._sort_release_survey_bucket(
                        in_progress_tickets,
                        bucket='in_progress',
                    ),
                    remaining_tickets=self._sort_release_survey_bucket(
                        remaining_tickets,
                        bucket='remaining',
                    ),
                    blocked_tickets=self._sort_release_survey_bucket(
                        blocked_tickets,
                        bucket='blocked',
                    ),
                    stale_tickets=sorted(key for key in stale_tickets if key),
                    unassigned_tickets=sorted(
                        key for key in unassigned_tickets if key
                    ),
                )
            )

        report = ReleaseSurveyReport(
            project_key=project_key,
            created_at=datetime.now(timezone.utc).isoformat(),
            scope_label=request.scope_label or '',
            survey_mode=survey_mode,
            releases_surveyed=release_names,
            release_summaries=release_summaries,
        )
        report.summary_markdown = self._format_release_survey_summary(report)

        if request.output_file:
            output_path = self._write_release_survey_xlsx(report, request.output_file)
            report.output_file = output_path
            log.info(f'Wrote release survey xlsx to {output_path}')

        return report

    @classmethod
    def _normalize_release_survey_mode(cls, survey_mode: Optional[str]) -> str:
        '''
        Normalize survey-mode aliases to the supported release-survey modes.
        '''
        normalized = str(survey_mode or '').strip().lower().replace('-', '_')
        if not normalized:
            return cls.RELEASE_SURVEY_MODE_FEATURE_DEV
        if normalized in (
            cls.RELEASE_SURVEY_MODE_FEATURE_DEV,
            'feature',
            'featuredev',
        ):
            return cls.RELEASE_SURVEY_MODE_FEATURE_DEV
        if normalized in (
            cls.RELEASE_SURVEY_MODE_BUG,
            'bugs',
            'bug_only',
            'bugs_only',
            'bugonly',
        ):
            return cls.RELEASE_SURVEY_MODE_BUG
        raise ValueError(
            f'Unsupported release survey mode: {survey_mode}. '
            'Use "feature_dev" or "bug".'
        )

    @staticmethod
    def _ticket_matches_release_survey_mode(
        ticket: Dict[str, Any],
        survey_mode: str,
    ) -> bool:
        '''
        Apply the issue-type boundary for the requested release-survey mode.
        '''
        issue_type = str(ticket.get('issuetype') or '').strip().casefold()
        is_bug = issue_type == 'bug'
        if survey_mode == GanttProjectPlannerAgent.RELEASE_SURVEY_MODE_BUG:
            return is_bug
        return not is_bug

    def _classify_release_survey_ticket(
        self,
        release: str,
        ticket: Dict[str, Any],
    ) -> Dict[str, Any]:
        '''
        Normalize and bucket a release ticket for survey reporting.
        '''
        status = str(ticket.get('status') or '').strip()
        priority = str(ticket.get('priority') or 'Unspecified').strip() or 'Unspecified'
        issue_type = str(
            ticket.get('issuetype') or ticket.get('issue_type') or 'Unknown'
        ).strip() or 'Unknown'
        assignee = str(ticket.get('assignee') or 'Unassigned').strip() or 'Unassigned'
        updated_raw = str(ticket.get('updated') or '').strip()
        resolution_raw = str(ticket.get('resolutiondate') or '').strip()
        updated_dt = self.backlog_interpreter.parse_jira_datetime(updated_raw)
        resolved_dt = self.backlog_interpreter.parse_jira_datetime(resolution_raw)
        age_days = 0
        if updated_dt is not None:
            age_days = max((self._utc_now() - updated_dt).days, 0)

        raw_components = ticket.get('components') or []
        if isinstance(raw_components, list):
            components = [
                str(component).strip()
                for component in raw_components
                if str(component).strip()
            ]
        elif str(raw_components).strip():
            components = [
                part.strip() for part in str(raw_components).split(',')
                if part.strip()
            ]
        else:
            components = []

        product_family = self._normalize_release_survey_field_values(
            ticket.get('product_family')
        )

        status_lower = status.casefold()
        is_done = bool(resolved_dt) or self.backlog_interpreter.is_done_status(status)
        is_blocked = (
            not is_done
            and any(
                token in status_lower
                for token in ('block', 'hold', 'wait', 'stalled', 'depend')
            )
        )
        is_in_progress = (
            not is_done
            and not is_blocked
            and any(
                token in status_lower
                for token in (
                    'progress',
                    'review',
                    'verify',
                    'test',
                    'qa',
                    'develop',
                    'implement',
                    'validation',
                    'integration',
                )
            )
        )

        if is_done:
            bucket = 'done'
        elif is_blocked:
            bucket = 'blocked'
        elif is_in_progress:
            bucket = 'in_progress'
        else:
            bucket = 'remaining'

        return {
            'release': release,
            'key': str(ticket.get('key') or ''),
            'summary': str(ticket.get('summary') or ''),
            'status': status,
            'priority': priority,
            'issue_type': issue_type,
            'assignee': assignee,
            'components': components or ['Unspecified'],
            'component_csv': ', '.join(components) if components else 'Unspecified',
            'product_family': product_family or ['Unspecified'],
            'product_family_csv': (
                ', '.join(product_family) if product_family else 'Unspecified'
            ),
            'created': str(ticket.get('created') or ''),
            'updated': updated_raw,
            'resolutiondate': resolution_raw,
            'reporter': str(ticket.get('reporter') or ''),
            'labels': list(ticket.get('labels') or []),
            'url': str(ticket.get('url') or ''),
            'bucket': bucket,
            'age_days': age_days,
            'is_stale': age_days >= self.STALE_DAYS and not is_done,
            'is_unassigned': assignee.casefold() == 'unassigned',
        }

    def _release_survey_family_names(
        self,
        scope_label: Optional[str],
        tickets: List[Dict[str, Any]],
    ) -> List[str]:
        '''
        Determine which product-family buckets should be rendered for the survey.
        '''
        requested_families = normalize_csv_list(scope_label)
        if requested_families:
            return requested_families

        discovered: Dict[str, str] = {}
        for ticket in tickets:
            for family in list(ticket.get('product_family') or []):
                family_name = str(family or '').strip()
                if not family_name or family_name.casefold() == 'unspecified':
                    continue
                discovered.setdefault(family_name.casefold(), family_name)

        return list(discovered.values())

    @staticmethod
    def _ticket_matches_release_survey_family(
        ticket: Dict[str, Any],
        family_name: str,
    ) -> bool:
        '''
        Return whether the ticket belongs to the named product-family bucket.
        '''
        target = str(family_name or '').strip().casefold()
        if not target:
            return False

        ticket_families = [
            str(value or '').strip().casefold()
            for value in list(ticket.get('product_family') or [])
            if str(value or '').strip()
        ]
        return target in ticket_families

    def _sort_release_survey_epics(
        self,
        tickets: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        '''
        Sort epics with unfinished items first, then by priority.
        '''
        return sorted(
            tickets,
            key=lambda item: (
                1 if str(item.get('bucket') or '') == 'done' else 0,
                self._release_survey_priority_rank(str(item.get('priority') or '')),
                -int(item.get('age_days') or 0),
                str(item.get('key') or ''),
            ),
        )

    def _build_release_survey_family_breakdowns(
        self,
        family_names: List[str],
        tickets: List[Dict[str, Any]],
    ) -> Dict[str, Dict[str, Any]]:
        '''
        Build per-family ticket buckets for the survey markdown.
        '''
        breakdowns: Dict[str, Dict[str, Any]] = {}
        for family_name in family_names:
            family_tickets = [
                ticket
                for ticket in tickets
                if self._ticket_matches_release_survey_family(ticket, family_name)
            ]
            epics = [
                ticket
                for ticket in family_tickets
                if str(ticket.get('issue_type') or '').casefold() == 'epic'
            ]
            breakdowns[family_name] = {
                'total_tickets': len(family_tickets),
                'in_progress_tickets': self._sort_release_survey_bucket(
                    [
                        ticket for ticket in family_tickets
                        if str(ticket.get('bucket') or '') == 'in_progress'
                    ],
                    bucket='in_progress',
                ),
                'remaining_tickets': self._sort_release_survey_bucket(
                    [
                        ticket for ticket in family_tickets
                        if str(ticket.get('bucket') or '') == 'remaining'
                    ],
                    bucket='remaining',
                ),
                'epics': self._sort_release_survey_epics(epics),
            }

        return breakdowns

    def _build_release_survey_family_epic_analysis(
        self,
        family_names: List[str],
        family_breakdowns: Dict[str, Dict[str, Any]],
        survey_mode: str,
        *,
        hierarchy_cache: Optional[Dict[str, Dict[str, Any]]] = None,
    ) -> Dict[str, List[Dict[str, Any]]]:
        '''
        Build open-child analysis for unfinished epics in each product family.
        '''
        analysis_by_family: Dict[str, List[Dict[str, Any]]] = {}
        resolved_cache = hierarchy_cache if hierarchy_cache is not None else {}

        for family_name in family_names:
            family_epics = list(
                (family_breakdowns.get(family_name) or {}).get('epics') or []
            )
            analysis_entries: List[Dict[str, Any]] = []
            for epic in family_epics:
                if str(epic.get('bucket') or '') == 'done':
                    continue
                analysis_entries.append(
                    self._build_release_survey_epic_analysis_entry(
                        family_name,
                        epic,
                        survey_mode,
                        hierarchy_cache=resolved_cache,
                    )
                )
            analysis_by_family[family_name] = analysis_entries

        return analysis_by_family

    def _build_release_survey_epic_analysis_entry(
        self,
        family_name: str,
        epic: Dict[str, Any],
        survey_mode: str,
        *,
        hierarchy_cache: Dict[str, Dict[str, Any]],
    ) -> Dict[str, Any]:
        '''
        Analyze the unfinished descendants of one epic.
        '''
        epic_key = str(epic.get('key') or '').strip()
        cached_entry = hierarchy_cache.get(epic_key)
        if cached_entry is None:
            try:
                result = get_children_hierarchy(root_key=epic_key, limit=500)
                if result.is_success:
                    cached_entry = {
                        'items': list(result.data or []),
                        'error': '',
                    }
                else:
                    cached_entry = {
                        'items': [],
                        'error': str(result.error or 'Hierarchy lookup failed'),
                    }
            except Exception as e:
                cached_entry = {
                    'items': [],
                    'error': str(e),
                }
            hierarchy_cache[epic_key] = cached_entry

        hierarchy_items = list(cached_entry.get('items') or [])
        error_text = str(cached_entry.get('error') or '')
        open_children: List[Dict[str, Any]] = []
        open_by_status: Counter[str] = Counter()
        open_by_type: Counter[str] = Counter()
        total_descendants = 0

        for item in hierarchy_items[1:]:
            issue_type = str(item.get('type') or item.get('issue_type') or 'Unknown')
            if not self._ticket_matches_release_survey_mode(
                {'issuetype': issue_type},
                survey_mode,
            ):
                continue

            total_descendants += 1
            status = str(item.get('status') or '')
            if self.backlog_interpreter.is_done_status(status):
                continue

            fix_versions = list(item.get('fix_versions') or [])
            fix_version_csv = ', '.join(
                str(value).strip()
                for value in fix_versions
                if str(value).strip()
            )

            open_child = {
                'key': str(item.get('key') or ''),
                'summary': str(item.get('summary') or ''),
                'issue_type': issue_type or 'Unknown',
                'status': status,
                'priority': str(item.get('priority') or 'Unspecified'),
                'assignee': str(item.get('assignee') or 'Unassigned'),
                'depth': int(item.get('depth') or 0),
                'fix_version_csv': fix_version_csv or 'Unspecified',
                'url': str(item.get('url') or ''),
            }
            open_children.append(open_child)
            open_by_status[open_child['status'] or 'Unknown'] += 1
            open_by_type[open_child['issue_type'] or 'Unknown'] += 1

        return {
            'family': family_name,
            'epic': dict(epic),
            'error': error_text,
            'total_descendant_count': total_descendants,
            'open_child_count': len(open_children),
            'open_by_status': dict(open_by_status),
            'open_by_type': dict(open_by_type),
            'ready_to_close': not open_children and not error_text,
            'open_children': open_children,
        }

    @classmethod
    def _release_survey_priority_rank(cls, priority: str) -> int:
        '''
        Return a sortable priority rank for survey ticket ordering.
        '''
        normalized = str(priority or '').casefold()
        if 'p0' in normalized or 'stopper' in normalized or 'blocker' in normalized:
            return 0
        if 'p1' in normalized or 'critical' in normalized:
            return 1
        if 'p2' in normalized or 'high' in normalized:
            return 2
        if 'p3' in normalized or 'medium' in normalized:
            return 3
        if 'p4' in normalized or 'low' in normalized:
            return 4
        return 5

    @classmethod
    def _normalize_release_survey_person_name(cls, raw_name: Any) -> str:
        '''
        Normalize a Jira or org-chart display name for manager lookups.
        '''
        name = str(raw_name or '').strip()
        if not name:
            return ''

        if ',' in name:
            parts = [part.strip() for part in name.split(',') if part.strip()]
            if len(parts) >= 2:
                name = ' '.join(parts[1:] + [parts[0]])

        normalized = ' '.join(name.split()).casefold()
        return cls._RELEASE_SURVEY_MANAGER_ALIASES.get(normalized, normalized)

    @classmethod
    def _release_survey_person_lookup_keys(cls, raw_name: Any) -> List[str]:
        '''
        Build candidate lookup keys for a Jira assignee display name.
        '''
        normalized = cls._normalize_release_survey_person_name(raw_name)
        if not normalized:
            return []

        lookup_keys = [normalized]
        name_parts = normalized.split()
        if len(name_parts) > 2:
            lookup_keys.append(f'{name_parts[0]} {name_parts[-1]}')

        return lookup_keys

    @classmethod
    def _get_release_survey_manager_lookup(cls) -> Dict[str, str]:
        '''
        Load direct-manager mappings from the local org chart.
        '''
        if cls._release_survey_manager_lookup_cache is not None:
            return cls._release_survey_manager_lookup_cache

        org_chart_path = os.path.join('data', 'knowledge', 'heqing_org.json')
        if not os.path.exists(org_chart_path):
            cls._release_survey_manager_lookup_cache = {}
            return cls._release_survey_manager_lookup_cache

        try:
            with open(org_chart_path, 'r', encoding='utf-8') as f:
                org_chart_data = json.load(f)
        except Exception as e:
            log.warning(f'Failed to load release survey org chart {org_chart_path}: {e}')
            cls._release_survey_manager_lookup_cache = {}
            return cls._release_survey_manager_lookup_cache

        root = (
            org_chart_data.get('org_chart', {}).get('root')
            if isinstance(org_chart_data, dict)
            else None
        )
        if not isinstance(root, dict):
            cls._release_survey_manager_lookup_cache = {}
            return cls._release_survey_manager_lookup_cache

        manager_lookup: Dict[str, str] = {}

        def _walk(node: Mapping[str, Any], manager_name: str = '') -> None:
            current_name = str(node.get('name') or '').strip()
            normalized_name = cls._normalize_release_survey_person_name(current_name)
            if normalized_name and manager_name:
                manager_lookup[normalized_name] = manager_name

            for child in list(node.get('children') or []):
                if isinstance(child, dict):
                    _walk(child, current_name)

        _walk(root)
        cls._release_survey_manager_lookup_cache = manager_lookup
        return cls._release_survey_manager_lookup_cache

    @classmethod
    def _lookup_release_survey_manager_name(
        cls,
        assignee: Any,
        manager_lookup: Optional[Mapping[str, str]] = None,
    ) -> str:
        '''
        Resolve the assignee's direct manager from the local org chart.
        '''
        assignee_name = str(assignee or '').strip()
        if not assignee_name:
            return 'Unassigned'

        if assignee_name.casefold() == 'unassigned':
            return 'Unassigned'

        resolved_lookup = (
            manager_lookup
            if manager_lookup is not None
            else cls._get_release_survey_manager_lookup()
        )
        for lookup_key in cls._release_survey_person_lookup_keys(assignee_name):
            manager_name = str(resolved_lookup.get(lookup_key) or '').strip()
            if manager_name:
                return manager_name

        return 'Manager Not Found'

    @staticmethod
    def _escape_release_survey_table_cell(value: Any) -> str:
        '''
        Escape a Markdown table cell.
        '''
        text = ' '.join(str(value or '').splitlines()).strip()
        if not text:
            return ''
        return text.replace('|', '\\|')

    @staticmethod
    def _normalize_release_survey_field_values(raw_value: Any) -> List[str]:
        '''
        Normalize Jira multi-value fields into a list of display strings.
        '''
        if raw_value is None:
            return []

        values = raw_value if isinstance(raw_value, list) else [raw_value]
        normalized: List[str] = []
        for item in values:
            if item is None:
                continue

            if isinstance(item, str):
                parts = [part.strip() for part in item.split(',') if part.strip()]
                normalized.extend(parts or [item.strip()])
                continue

            if isinstance(item, Mapping):
                candidate = item.get('name') or item.get('value')
                if candidate:
                    normalized.append(str(candidate).strip())
                continue

            candidate = getattr(item, 'name', None) or getattr(item, 'value', None)
            if candidate:
                normalized.append(str(candidate).strip())
                continue

            text = str(item).strip()
            if text:
                normalized.append(text)

        return [value for value in normalized if value]

    def _release_survey_datetime_sort_value(self, raw_value: Any) -> datetime:
        '''
        Parse a Jira timestamp for survey sorting.
        '''
        parsed = self.backlog_interpreter.parse_jira_datetime(str(raw_value or ''))
        if parsed is not None:
            return parsed
        return datetime.min.replace(tzinfo=timezone.utc)

    def _sort_release_survey_bucket(
        self,
        tickets: List[Dict[str, Any]],
        *,
        bucket: str,
    ) -> List[Dict[str, Any]]:
        '''
        Sort survey ticket buckets for readable summaries and exports.
        '''
        if bucket == 'done':
            return sorted(
                tickets,
                key=lambda item: (
                    self._release_survey_datetime_sort_value(
                        item.get('resolutiondate') or item.get('updated')
                    ),
                    str(item.get('key') or ''),
                ),
                reverse=True,
            )

        return sorted(
            tickets,
            key=lambda item: (
                self._release_survey_priority_rank(str(item.get('priority') or '')),
                -int(item.get('age_days') or 0),
                str(item.get('key') or ''),
            ),
        )

    def _query_release_tickets(
        self,
        release: str,
        scope_label: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        '''
        Query Jira for all tickets in a given release.

        Builds JQL with fixVersion filter and optional scope_label keyword.
        Returns a list of normalized ticket dicts.
        '''
        jql_parts = [
            f'project = {self.project_key or "STL"}',
            f'fixVersion = "{release}"',
        ]
        if scope_label:
            scope_values = normalize_csv_list(scope_label)
            if scope_values:
                scope_clauses: List[str] = []
                for scope_value in scope_values:
                    scope_clauses.append(f'labels = "{scope_value}"')
                    scope_clauses.append(f'"product family" = "{scope_value}"')
                jql_parts.append(f'({" OR ".join(scope_clauses)})')
        jql = ' AND '.join(jql_parts) + ' ORDER BY priority ASC, updated DESC'

        log.debug(f'Release ticket query JQL: {jql}')

        try:
            result = search_tickets(jql=jql, limit=500)
            if not result.is_success:
                log.warning(f'Release ticket search failed: {result.error}')
                return []
            tickets = result.data or []
        except Exception as e:
            log.error(f'Release ticket query failed for {release}: {e}')
            return []

        # Normalize ticket dicts to ensure consistent field names
        normalized: List[Dict[str, Any]] = []
        for t in tickets:
            normalized.append({
                'key': t.get('key', ''),
                'summary': t.get('summary', ''),
                'status': t.get('status', ''),
                'priority': t.get('priority', ''),
                'issuetype': t.get('issue_type') or t.get('type') or '',
                'assignee': t.get('assignee', ''),
                'reporter': t.get('reporter', ''),
                'fix_versions': t.get('fix_versions') or t.get('fix_version', ''),
                'components': t.get('components') or t.get('component', ''),
                'product_family': (
                    t.get('product_family')
                    or t.get('customfield_28434')
                    or t.get('customfield_28382')
                    or []
                ),
                'labels': t.get('labels') or [],
                'created': t.get('created', ''),
                'updated': t.get('updated', ''),
                'resolutiondate': t.get('resolutiondate', ''),
                'url': t.get('url', ''),
            })
        return normalized

    def _build_bug_summary(
        self,
        release: str,
        tickets: List[Dict[str, Any]],
        previous_tickets: Optional[List[Dict[str, Any]]] = None,
    ) -> BugSummary:
        '''
        Build a BugSummary from a list of tickets for a release.

        Filters to bugs only, counts by status/priority/component.
        If previous_tickets provided, computes new/closed/priority deltas.
        '''
        # Filter to bugs only
        bugs = [
            t for t in tickets
            if str(t.get('issuetype', '')).lower() == 'bug'
        ]

        by_status: Counter[str] = Counter()
        by_priority: Counter[str] = Counter()
        by_component: Counter[str] = Counter()

        for bug in bugs:
            by_status[str(bug.get('status', 'Unknown'))] += 1
            by_priority[str(bug.get('priority', 'Unknown'))] += 1
            # Components may be a list or csv string
            comps = bug.get('components', '')
            if isinstance(comps, list):
                for c in comps:
                    by_component[str(c) or 'Unspecified'] += 1
            elif isinstance(comps, str) and comps.strip():
                by_component[comps.strip()] += 1
            else:
                by_component['Unspecified'] += 1

        summary = BugSummary(
            release=release,
            total_bugs=len(bugs),
            by_status=dict(by_status),
            by_priority=dict(by_priority),
            by_component=dict(by_component),
        )

        # Compute deltas if previous tickets are available
        if previous_tickets is not None:
            prev_bugs = {
                str(t.get('key')): t for t in previous_tickets
                if str(t.get('issuetype', '')).lower() == 'bug'
            }
            curr_bugs = {str(b.get('key')): b for b in bugs}

            summary.new_since_last = [
                k for k in curr_bugs if k not in prev_bugs
            ]
            summary.closed_since_last = [
                k for k in prev_bugs if k not in curr_bugs
            ]
            summary.priority_changes = [
                {
                    'key': k,
                    'from': str(prev_bugs[k].get('priority', '')),
                    'to': str(curr_bugs[k].get('priority', '')),
                }
                for k in curr_bugs
                if k in prev_bugs
                and str(curr_bugs[k].get('priority', ''))
                != str(prev_bugs[k].get('priority', ''))
            ]

        return summary

    def _write_release_monitor_xlsx(
        self,
        report: ReleaseMonitorReport,
        output_file: str,
        all_release_tickets: Optional[Dict[str, List[Dict[str, Any]]]] = None,
    ) -> str:
        '''
        Generate a multi-sheet xlsx workbook for the release monitor report.

        Sheets:
        1. Bug Summary — one row per release with counts
        2. Bug Details — all bugs with hyperlinks and formatting
        3. Changes Since Last — delta rows (if delta data exists)
        4. Readiness — metric/value pairs (if readiness data exists)
        5. Gap Analysis — roadmap gap summary (if roadmap data exists)
        '''
        wb = Workbook()

        # -- Sheet 1: Bug Summary --
        ws_summary = wb.active or wb.create_sheet()
        ws_summary.title = 'Bug Summary'
        self._write_bug_summary_sheet(ws_summary, report)

        # -- Sheet 2: Bug Details --
        ws_details = wb.create_sheet('Bug Details')
        self._write_bug_details_sheet(
            ws_details, report, all_release_tickets or {},
        )

        # -- Sheet 3: Changes Since Last (conditional) --
        if report.delta:
            ws_changes = wb.create_sheet('Changes Since Last')
            self._write_changes_sheet(ws_changes, report)

        # -- Sheet 4: Readiness (conditional) --
        if report.readiness:
            ws_readiness = wb.create_sheet('Readiness')
            self._write_readiness_sheet(ws_readiness, report)

        # -- Sheet 5: Gap Analysis (conditional) --
        if report.roadmap_snapshot:
            ws_gaps = wb.create_sheet('Gap Analysis')
            self._write_gap_analysis_sheet(ws_gaps, report)

        os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
        wb.save(output_file)
        return output_file

    def _write_bug_summary_sheet(
        self,
        ws: Any,
        report: ReleaseMonitorReport,
    ) -> None:
        '''Write the Bug Summary sheet — one row per release.'''
        headers = [
            'Release', 'Total Bugs', 'P0', 'P1', 'P2', 'P3',
            'Open', 'In Progress', 'Verify', 'Closed',
        ]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, bs in enumerate(report.bug_summaries, 2):
            ws.cell(row=row_idx, column=1, value=bs.release)
            ws.cell(row=row_idx, column=2, value=bs.total_bugs)
            # Priority columns — match P0-Stopper, P1-Critical, etc.
            ws.cell(row=row_idx, column=3, value=self._priority_count(bs, 'P0'))
            ws.cell(row=row_idx, column=4, value=self._priority_count(bs, 'P1'))
            ws.cell(row=row_idx, column=5, value=self._priority_count(bs, 'P2'))
            ws.cell(row=row_idx, column=6, value=self._priority_count(bs, 'P3'))
            # Status columns
            ws.cell(row=row_idx, column=7, value=bs.by_status.get('Open', 0))
            ws.cell(row=row_idx, column=8, value=bs.by_status.get('In Progress', 0))
            ws.cell(row=row_idx, column=9, value=bs.by_status.get('Verify', 0))
            ws.cell(row=row_idx, column=10, value=bs.by_status.get('Closed', 0))

        _apply_header_style(ws, len(headers))
        _apply_status_conditional_formatting(ws, headers)
        _apply_priority_conditional_formatting(ws, headers)
        _auto_fit_columns(ws)

    def _write_bug_details_sheet(
        self,
        ws: Any,
        report: ReleaseMonitorReport,
        all_release_tickets: Dict[str, List[Dict[str, Any]]],
    ) -> None:
        '''Write the Bug Details sheet — all bugs sorted by priority then release.'''
        headers = [
            'Release', 'Key', 'Summary', 'Status', 'Priority',
            'Assignee', 'Component', 'Created', 'Updated',
        ]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        row = 2
        for bs in report.bug_summaries:
            tickets = all_release_tickets.get(bs.release, [])
            bugs = [
                t for t in tickets
                if str(t.get('issuetype', '')).lower() == 'bug'
            ]
            # Sort by priority (P0 first)
            bugs.sort(key=lambda t: str(t.get('priority', 'ZZZ')))

            for bug in bugs:
                ws.cell(row=row, column=1, value=bs.release)
                key = str(bug.get('key', ''))
                key_cell = ws.cell(row=row, column=2, value=key)
                if key:
                    key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{key}'
                    key_cell.font = Font(color='0563C1', underline='single')
                ws.cell(row=row, column=3, value=bug.get('summary', ''))
                ws.cell(row=row, column=4, value=bug.get('status', ''))
                ws.cell(row=row, column=5, value=bug.get('priority', ''))
                ws.cell(row=row, column=6, value=bug.get('assignee', ''))
                comps = bug.get('components', '')
                if isinstance(comps, list):
                    comps = ', '.join(str(c) for c in comps)
                ws.cell(row=row, column=7, value=comps)
                ws.cell(row=row, column=8, value=bug.get('created', ''))
                ws.cell(row=row, column=9, value=bug.get('updated', ''))
                row += 1

        _apply_header_style(ws, len(headers))
        _apply_status_conditional_formatting(ws, headers)
        _apply_priority_conditional_formatting(ws, headers)
        ws.freeze_panes = 'A2'
        _auto_fit_columns(ws)

    def _write_changes_sheet(
        self,
        ws: Any,
        report: ReleaseMonitorReport,
    ) -> None:
        '''Write the Changes Since Last sheet — delta rows by change type.'''
        headers = ['Change Type', 'Release', 'Key', 'Summary', 'Detail']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        delta = report.delta or {}
        release = delta.get('release', '')
        row = 2

        # Section: New Bugs
        new_tickets = delta.get('new_tickets', [])
        if new_tickets:
            row = self._write_section_divider(ws, row, 'New Bugs', len(headers))
            for entry in new_tickets:
                if isinstance(entry, dict):
                    key = str(entry.get('key', ''))
                    entry_release = str(entry.get('release', release))
                else:
                    key = str(entry)
                    entry_release = release
                ws.cell(row=row, column=1, value='New Bug')
                ws.cell(row=row, column=2, value=entry_release)
                key_cell = ws.cell(row=row, column=3, value=key)
                key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{key}'
                key_cell.font = Font(color='0563C1', underline='single')
                ws.cell(row=row, column=4, value='')
                ws.cell(row=row, column=5, value='Added since last snapshot')
                row += 1

        # Section: Closed Bugs
        closed_tickets = delta.get('closed_tickets', [])
        if closed_tickets:
            row = self._write_section_divider(
                ws, row, 'Closed Bugs', len(headers),
            )
            for entry in closed_tickets:
                if isinstance(entry, dict):
                    key = str(entry.get('key', ''))
                    entry_release = str(entry.get('release', release))
                else:
                    key = str(entry)
                    entry_release = release
                ws.cell(row=row, column=1, value='Closed Bug')
                ws.cell(row=row, column=2, value=entry_release)
                key_cell = ws.cell(row=row, column=3, value=key)
                key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{key}'
                key_cell.font = Font(color='0563C1', underline='single')
                ws.cell(row=row, column=4, value='')
                ws.cell(row=row, column=5, value='Closed since last snapshot')
                row += 1

        # Section: Priority Changes
        priority_changes = delta.get('priority_changes', [])
        if priority_changes:
            row = self._write_section_divider(
                ws, row, 'Priority Changes', len(headers),
            )
            for change in priority_changes:
                ws.cell(row=row, column=1, value='Priority Change')
                ws.cell(
                    row=row,
                    column=2,
                    value=str(change.get('release', release)),
                )
                key = str(change.get('key', ''))
                key_cell = ws.cell(row=row, column=3, value=key)
                if key:
                    key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{key}'
                    key_cell.font = Font(color='0563C1', underline='single')
                ws.cell(row=row, column=4, value='')
                ws.cell(
                    row=row, column=5,
                    value=f'{change.get("from", "")} \u2192 {change.get("to", "")}',
                )
                row += 1

        # Section: Status Changes
        status_changes = delta.get('status_changes', [])
        if status_changes:
            row = self._write_section_divider(
                ws, row, 'Status Changes', len(headers),
            )
            for change in status_changes:
                ws.cell(row=row, column=1, value='Status Change')
                ws.cell(
                    row=row,
                    column=2,
                    value=str(change.get('release', release)),
                )
                key = str(change.get('key', ''))
                key_cell = ws.cell(row=row, column=3, value=key)
                if key:
                    key_cell.hyperlink = f'{JIRA_BASE_URL}/browse/{key}'
                    key_cell.font = Font(color='0563C1', underline='single')
                ws.cell(row=row, column=4, value='')
                ws.cell(
                    row=row, column=5,
                    value=f'{change.get("from", "")} \u2192 {change.get("to", "")}',
                )
                row += 1

        _apply_header_style(ws, len(headers))
        _auto_fit_columns(ws)

    def _write_readiness_sheet(
        self,
        ws: Any,
        report: ReleaseMonitorReport,
    ) -> None:
        '''Write the Readiness sheet — metric/value pairs.'''
        headers = ['Metric', 'Value']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        readiness = report.readiness or {}
        metrics = [
            ('Release', readiness.get('release', '')),
            ('Total Open', readiness.get('total_open', 0)),
            ('P0 Open', readiness.get('p0_open', 0)),
            ('P1 Open', readiness.get('p1_open', 0)),
            ('Daily Close Rate', f'{readiness.get("daily_close_rate", 0.0):.2f}'),
            (
                'Estimated Days Remaining',
                readiness.get('estimated_days_remaining', 'N/A'),
            ),
            ('Stale Ticket Count', len(readiness.get('stale_tickets', []))),
        ]

        row = 2
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=str(value))
            row += 1

        # Component risks as sub-rows
        component_risks = readiness.get('component_risks', [])
        if component_risks:
            row = self._write_section_divider(
                ws, row, 'Component Risks', len(headers),
            )
            for risk in component_risks:
                ws.cell(
                    row=row, column=1,
                    value=f'  {risk.get("component", "Unknown")}',
                )
                ws.cell(
                    row=row, column=2,
                    value=str(risk.get('open_tickets', 0)),
                )
                row += 1

        _apply_header_style(ws, len(headers))
        _auto_fit_columns(ws)

    def _write_gap_analysis_sheet(
        self,
        ws: Any,
        report: ReleaseMonitorReport,
    ) -> None:
        '''Write the Gap Analysis sheet — roadmap snapshot summary.'''
        headers = ['Section', 'Jira Items', 'Proposed Gaps', 'Summary']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        roadmap = report.roadmap_snapshot or {}
        sections = roadmap.get('sections', [])

        row = 2
        for section in sections:
            ws.cell(row=row, column=1, value=section.get('title', ''))
            ws.cell(row=row, column=2, value=len(section.get('items', [])))
            ws.cell(row=row, column=3, value=len(section.get('gaps', [])))
            # Build a brief summary of gaps
            gap_summaries = [
                g.get('summary', '') for g in section.get('gaps', [])
            ]
            ws.cell(
                row=row, column=4,
                value='; '.join(gap_summaries[:5]) if gap_summaries else 'No gaps identified',
            )
            row += 1

        _apply_header_style(ws, len(headers))
        _auto_fit_columns(ws)

    @staticmethod
    def _priority_count(bs: BugSummary, prefix: str) -> int:
        '''
        Sum bug counts for all priority values starting with the given prefix.

        Reasoning: Jira priorities may be "P0-Stopper", "P1-Critical", etc.
        We match any key starting with the prefix (e.g. "P0").
        '''
        return sum(
            count for key, count in bs.by_priority.items()
            if key.startswith(prefix)
        )

    @staticmethod
    def _format_release_monitor_summary(
        report: ReleaseMonitorReport,
    ) -> str:
        '''
        Generate a markdown summary of the release monitor report.

        Includes key metrics, bug trends, P0/P1 alerts, and velocity.
        '''
        lines = [
            f'# Release Monitor: {report.project_key}',
            f'',
            f'**Report ID**: {report.report_id}',
            f'**Created**: {report.created_at}',
            f'**Releases Monitored**: {", ".join(report.releases_monitored)}',
            f'',
            '---',
            '',
        ]

        # Bug summary per release
        lines.append('## Bug Summary')
        lines.append('')
        for bs in report.bug_summaries:
            lines.append(f'### {bs.release}')
            lines.append(f'- **Total Bugs**: {bs.total_bugs}')
            for prio, count in sorted(bs.by_priority.items()):
                lines.append(f'  - {prio}: {count}')
            for status, count in sorted(bs.by_status.items()):
                lines.append(f'  - {status}: {count}')
            if bs.new_since_last:
                lines.append(
                    f'- **New since last**: {len(bs.new_since_last)} '
                    f'({", ".join(bs.new_since_last[:5])})'
                )
            if bs.closed_since_last:
                lines.append(
                    f'- **Closed since last**: {len(bs.closed_since_last)} '
                    f'({", ".join(bs.closed_since_last[:5])})'
                )
            lines.append('')

        # P0/P1 alerts
        if report.total_p0 > 0 or report.total_p1 > 0:
            lines.append('## \u26a0 Priority Alerts')
            lines.append('')
            if report.total_p0 > 0:
                lines.append(f'- **P0 (Stopper)**: {report.total_p0} open')
            if report.total_p1 > 0:
                lines.append(f'- **P1 (Critical)**: {report.total_p1} open')
            lines.append('')

        # Velocity
        if report.velocity:
            vel = report.velocity
            lines.append('## Velocity')
            lines.append('')
            lines.append(f'- Opened: {vel.get("opened", 0)}')
            lines.append(f'- Closed: {vel.get("closed", 0)}')
            lines.append(f'- Daily open rate: {vel.get("daily_open_rate", 0.0):.2f}')
            lines.append(f'- Daily close rate: {vel.get("daily_close_rate", 0.0):.2f}')
            lines.append(f'- Daily net rate: {vel.get("daily_net_rate", 0.0):.2f}')
            net = float(vel.get('daily_net_rate', 0.0))
            if net > 0:
                lines.append('- **Trend**: Bug count increasing \u2191')
            elif net < 0:
                lines.append('- **Trend**: Bug count decreasing \u2193')
            else:
                lines.append('- **Trend**: Stable \u2192')
            lines.append('')

        # Readiness
        if report.readiness:
            r = report.readiness
            lines.append('## Readiness')
            lines.append('')
            lines.append(f'- Total open: {r.get("total_open", 0)}')
            lines.append(f'- P0 open: {r.get("p0_open", 0)}')
            lines.append(f'- P1 open: {r.get("p1_open", 0)}')
            est = r.get('estimated_days_remaining')
            if est is not None:
                lines.append(f'- Estimated days remaining: {est:.1f}')
            else:
                lines.append('- Estimated days remaining: N/A (no close rate)')
            stale = r.get('stale_tickets', [])
            if stale:
                lines.append(
                    f'- Stale tickets: {len(stale)} '
                    f'({", ".join(stale[:5])})'
                )
            lines.append('')

        return '\n'.join(lines)

    def _write_release_survey_xlsx(
        self,
        report: ReleaseSurveyReport,
        output_file: str,
    ) -> str:
        '''
        Generate a multi-sheet xlsx workbook for the release survey report.
        '''
        wb = Workbook()

        ws_overview = wb.active or wb.create_sheet()
        ws_overview.title = 'Overview'
        self._write_release_survey_overview_sheet(ws_overview, report)

        ws_in_progress = wb.create_sheet('In Progress')
        self._write_release_survey_ticket_sheet(
            ws_in_progress,
            self._aggregate_release_survey_tickets(report, 'in_progress_tickets'),
        )

        ws_blocked = wb.create_sheet('Blocked')
        self._write_release_survey_ticket_sheet(
            ws_blocked,
            self._aggregate_release_survey_tickets(report, 'blocked_tickets'),
        )

        ws_remaining = wb.create_sheet('Remaining')
        self._write_release_survey_ticket_sheet(
            ws_remaining,
            self._aggregate_release_survey_tickets(report, 'remaining_tickets'),
        )

        ws_done = wb.create_sheet('Done')
        self._write_release_survey_ticket_sheet(
            ws_done,
            self._aggregate_release_survey_tickets(report, 'done_tickets'),
        )

        os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
        wb.save(output_file)
        return output_file

    @staticmethod
    def _aggregate_release_survey_tickets(
        report: ReleaseSurveyReport,
        attribute: str,
    ) -> List[Dict[str, Any]]:
        '''
        Flatten per-release survey ticket buckets into one export list.
        '''
        rows: List[Dict[str, Any]] = []
        for summary in report.release_summaries:
            rows.extend(list(getattr(summary, attribute, [])))
        return rows

    def _write_release_survey_overview_sheet(
        self,
        ws: Any,
        report: ReleaseSurveyReport,
    ) -> None:
        '''
        Write the Overview sheet for the release survey workbook.
        '''
        headers = [
            'Release',
            'Total',
            'Done',
            'In Progress',
            'Remaining',
            'Blocked',
            'Stale',
            'Unassigned',
            'Completion %',
        ]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, summary in enumerate(report.release_summaries, 2):
            ws.cell(row=row_idx, column=1, value=summary.release)
            ws.cell(row=row_idx, column=2, value=summary.total_tickets)
            ws.cell(row=row_idx, column=3, value=summary.done_count)
            ws.cell(row=row_idx, column=4, value=summary.in_progress_count)
            ws.cell(row=row_idx, column=5, value=summary.remaining_count)
            ws.cell(row=row_idx, column=6, value=summary.blocked_count)
            ws.cell(row=row_idx, column=7, value=summary.stale_count)
            ws.cell(row=row_idx, column=8, value=summary.unassigned_count)
            ws.cell(row=row_idx, column=9, value=round(summary.completion_pct, 1))

        total_row = len(report.release_summaries) + 2
        ws.cell(row=total_row, column=1, value='TOTAL')
        ws.cell(row=total_row, column=2, value=report.total_tickets)
        ws.cell(row=total_row, column=3, value=report.done_count)
        ws.cell(row=total_row, column=4, value=report.in_progress_count)
        ws.cell(row=total_row, column=5, value=report.remaining_count)
        ws.cell(row=total_row, column=6, value=report.blocked_count)
        ws.cell(row=total_row, column=7, value=report.stale_count)
        ws.cell(row=total_row, column=8, value=report.unassigned_count)
        ws.cell(row=total_row, column=9, value=round(report.completion_pct, 1))

        _apply_header_style(ws, len(headers))
        _auto_fit_columns(ws)

    def _write_release_survey_ticket_sheet(
        self,
        ws: Any,
        tickets: List[Dict[str, Any]],
    ) -> None:
        '''
        Write a release survey ticket bucket sheet.
        '''
        headers = [
            'Release',
            'Key',
            'Summary',
            'Status',
            'Priority',
            'Issue Type',
            'Assignee',
            'Components',
            'Updated',
            'Resolved',
            'Age Days',
        ]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        row = 2
        for ticket in tickets:
            ws.cell(row=row, column=1, value=ticket.get('release', ''))
            key = str(ticket.get('key', ''))
            key_cell = ws.cell(row=row, column=2, value=key)
            if key:
                key_cell.hyperlink = str(
                    ticket.get('url') or f'{JIRA_BASE_URL}/browse/{key}'
                )
                key_cell.font = Font(color='0563C1', underline='single')
            ws.cell(row=row, column=3, value=ticket.get('summary', ''))
            ws.cell(row=row, column=4, value=ticket.get('status', ''))
            ws.cell(row=row, column=5, value=ticket.get('priority', ''))
            ws.cell(row=row, column=6, value=ticket.get('issue_type', ''))
            ws.cell(row=row, column=7, value=ticket.get('assignee', ''))
            ws.cell(row=row, column=8, value=ticket.get('component_csv', ''))
            ws.cell(row=row, column=9, value=ticket.get('updated', ''))
            ws.cell(row=row, column=10, value=ticket.get('resolutiondate', ''))
            ws.cell(row=row, column=11, value=ticket.get('age_days', 0))
            row += 1

        _apply_header_style(ws, len(headers))
        _apply_status_conditional_formatting(ws, headers)
        _apply_priority_conditional_formatting(ws, headers)
        ws.freeze_panes = 'A2'
        _auto_fit_columns(ws)

    @classmethod
    def _format_release_survey_confluence_markdown(
        cls,
        report: ReleaseSurveyReport,
    ) -> str:
        '''
        Generate a Confluence-oriented survey view with live Jira macros.
        '''
        lines = [
            f'# Release Survey: {report.project_key}',
            '',
            f'**Survey ID**: {report.survey_id}',
            f'**Created**: {report.created_at}',
            f'**Survey Mode**: {report.survey_mode.replace("_", " ").title()}',
            f'**Releases Surveyed**: {", ".join(report.releases_surveyed)}',
        ]
        if report.scope_label:
            lines.append(f'**Scope Label**: {report.scope_label}')
        lines.extend([
            '',
            '## Snapshot Totals',
            '',
            '- Counts in this section are a point-in-time snapshot.',
            '- Jira macro sections below are live Jira views and may drift after publication.',
            f'- Total tickets: {report.total_tickets}',
            f'- Done: {report.done_count} ({report.completion_pct:.1f}%)',
            f'- In progress: {report.in_progress_count}',
            f'- Remaining: {report.remaining_count}',
            f'- Blocked: {report.blocked_count}',
            f'- Stale: {report.stale_count}',
            f'- Unassigned: {report.unassigned_count}',
            '',
        ])

        if not report.release_summaries:
            lines.append('No releases matched this survey.')
            return '\n'.join(lines)

        for summary in report.release_summaries:
            lines.extend([
                f'## {summary.release}',
                '',
                f'- Snapshot total tickets: {summary.total_tickets}',
                f'- Snapshot done: {summary.done_count} ({summary.completion_pct:.1f}%)',
                f'- Snapshot in progress: {summary.in_progress_count}',
                f'- Snapshot remaining: {summary.remaining_count}',
                f'- Snapshot blocked: {summary.blocked_count}',
                f'- Snapshot stale: {summary.stale_count}',
                f'- Snapshot unassigned: {summary.unassigned_count}',
            ])

            if summary.status_breakdown:
                status_line = ', '.join(
                    f'{status}={count}'
                    for status, count in sorted(summary.status_breakdown.items())
                )
                lines.append(f'- Snapshot status mix: {status_line}')

            lines.append('')

            cls._append_release_survey_confluence_jira_macro(
                lines,
                'In Progress',
                project_key=report.project_key,
                release=summary.release,
                scope_label=report.scope_label,
                survey_mode=report.survey_mode,
                statuses=cls._release_survey_ticket_statuses(
                    summary.in_progress_tickets
                ),
                snapshot_count=len(summary.in_progress_tickets),
            )
            cls._append_release_survey_ticket_lines(
                lines,
                'Blocked',
                summary.blocked_tickets,
                limit=25,
            )
            cls._append_release_survey_confluence_jira_macro(
                lines,
                'Remaining',
                project_key=report.project_key,
                release=summary.release,
                scope_label=report.scope_label,
                survey_mode=report.survey_mode,
                statuses=cls._release_survey_ticket_statuses(
                    summary.remaining_tickets
                ),
                snapshot_count=len(summary.remaining_tickets),
            )
            cls._append_release_survey_ticket_lines(
                lines,
                'Done',
                summary.done_tickets,
                limit=20,
            )

            family_breakdowns = dict(summary.family_breakdowns or {})
            family_epic_analysis = dict(summary.family_epic_analysis or {})
            if family_breakdowns:
                lines.extend([
                    '### Product Family Breakouts',
                    '',
                    '- The Jira tables in this section are live views filtered by product family.',
                    '',
                ])
                for family_name, family_view in family_breakdowns.items():
                    cls._append_release_survey_confluence_jira_macro(
                        lines,
                        f'{family_name} In Progress',
                        project_key=report.project_key,
                        release=summary.release,
                        scope_label=report.scope_label,
                        survey_mode=report.survey_mode,
                        family_name=family_name,
                        statuses=cls._release_survey_ticket_statuses(
                            list(family_view.get('in_progress_tickets') or [])
                        ),
                        snapshot_count=len(
                            list(family_view.get('in_progress_tickets') or [])
                        ),
                        heading_level='####',
                    )
                    cls._append_release_survey_confluence_jira_macro(
                        lines,
                        f'{family_name} Remaining',
                        project_key=report.project_key,
                        release=summary.release,
                        scope_label=report.scope_label,
                        survey_mode=report.survey_mode,
                        family_name=family_name,
                        statuses=cls._release_survey_ticket_statuses(
                            list(family_view.get('remaining_tickets') or [])
                        ),
                        snapshot_count=len(
                            list(family_view.get('remaining_tickets') or [])
                        ),
                        heading_level='####',
                    )
                    cls._append_release_survey_confluence_jira_macro(
                        lines,
                        f'{family_name} Epics',
                        project_key=report.project_key,
                        release=summary.release,
                        scope_label=report.scope_label,
                        survey_mode=report.survey_mode,
                        family_name=family_name,
                        issue_type='Epic',
                        snapshot_count=len(list(family_view.get('epics') or [])),
                        heading_level='####',
                    )
                    cls._append_release_survey_epic_analysis(
                        lines,
                        family_name,
                        list(family_epic_analysis.get(family_name) or []),
                        limit=40,
                        heading_level='####',
                    )

        return '\n'.join(lines)

    @staticmethod
    def _release_survey_jql_literal(value: Any) -> str:
        '''
        Escape a string for a quoted JQL literal.
        '''
        return str(value or '').replace('\\', '\\\\').replace('"', '\\"')

    @classmethod
    def _release_survey_scope_clause(cls, scope_label: Optional[str]) -> str:
        '''
        Build the release-survey scope clause used for Jira macro queries.
        '''
        scope_values = normalize_csv_list(scope_label)
        if not scope_values:
            return ''

        scope_clauses: List[str] = []
        for scope_value in scope_values:
            escaped_scope = cls._release_survey_jql_literal(scope_value)
            scope_clauses.append(f'labels = "{escaped_scope}"')
            scope_clauses.append(f'"product family" = "{escaped_scope}"')
        return f'({" OR ".join(scope_clauses)})'

    @staticmethod
    def _release_survey_ticket_statuses(
        tickets: List[Dict[str, Any]],
    ) -> List[str]:
        '''
        Collect the distinct Jira statuses present in one ticket bucket.
        '''
        seen: Dict[str, str] = {}
        for ticket in tickets:
            status = str(ticket.get('status') or '').strip()
            if status:
                seen.setdefault(status.casefold(), status)
        return list(seen.values())

    @classmethod
    def _build_release_survey_jql(
        cls,
        *,
        project_key: str,
        release: str,
        scope_label: Optional[str],
        survey_mode: str,
        family_name: Optional[str] = None,
        statuses: Optional[List[str]] = None,
        issue_type: Optional[str] = None,
    ) -> str:
        '''
        Build a Jira query that approximates one survey section dynamically.
        '''
        clauses = [
            f'project = {project_key}',
            f'fixVersion = "{cls._release_survey_jql_literal(release)}"',
        ]

        scope_clause = cls._release_survey_scope_clause(scope_label)
        if scope_clause:
            clauses.append(scope_clause)

        if survey_mode == cls.RELEASE_SURVEY_MODE_BUG:
            clauses.append('issuetype = Bug')
        else:
            clauses.append('issuetype != Bug')

        if family_name:
            clauses.append(
                f'"product family" = "{cls._release_survey_jql_literal(family_name)}"'
            )

        if issue_type:
            clauses.append(
                f'issuetype = "{cls._release_survey_jql_literal(issue_type)}"'
            )

        if statuses:
            status_literals = ', '.join(
                f'"{cls._release_survey_jql_literal(status)}"'
                for status in statuses
                if str(status or '').strip()
            )
            if status_literals:
                clauses.append(f'status in ({status_literals})')

        return ' AND '.join(clauses) + ' ORDER BY priority ASC, updated DESC'

    @classmethod
    def _build_release_survey_jira_macro_block(
        cls,
        *,
        jql: str,
        columns: Optional[List[str]] = None,
        column_ids: Optional[List[str]] = None,
        maximum_issues: int = 200,
    ) -> str:
        '''
        Build a Confluence Jira Issues macro block for one live section.
        '''
        macro_columns = columns or [
            'key',
            'summary',
            'type',
            'status',
            'assignee',
            'priority',
            'updated',
        ]
        macro_column_ids = column_ids or [
            'issuekey',
            'summary',
            'issuetype',
            'status',
            'assignee',
            'priority',
            'updated',
        ]
        column_csv = ','.join(macro_columns)
        column_id_csv = ','.join(macro_column_ids)
        escaped_jql = escape(jql)
        escaped_server = escape(CONFLUENCE_JIRA_SERVER)
        escaped_server_id = escape(CONFLUENCE_JIRA_SERVER_ID)

        return '\n'.join([
            '<div>',
            '<ac:structured-macro ac:name="jira" ac:schema-version="1" data-layout="full-width">',
            f'<ac:parameter ac:name="server">{escaped_server}</ac:parameter>',
            f'<ac:parameter ac:name="columns">{column_csv}</ac:parameter>',
            f'<ac:parameter ac:name="columnIds">{column_id_csv}</ac:parameter>',
            f'<ac:parameter ac:name="maximumIssues">{maximum_issues}</ac:parameter>',
            f'<ac:parameter ac:name="jqlQuery">{escaped_jql}</ac:parameter>',
            f'<ac:parameter ac:name="serverId">{escaped_server_id}</ac:parameter>',
            '</ac:structured-macro>',
            '</div>',
        ])

    @classmethod
    def _append_release_survey_confluence_jira_macro(
        cls,
        lines: List[str],
        heading: str,
        *,
        project_key: str,
        release: str,
        scope_label: Optional[str],
        survey_mode: str,
        snapshot_count: int,
        family_name: Optional[str] = None,
        statuses: Optional[List[str]] = None,
        issue_type: Optional[str] = None,
        heading_level: str = '###',
    ) -> None:
        '''
        Append a live Jira macro section for the Confluence survey page.
        '''
        lines.append(f'{heading_level} {heading}')
        if snapshot_count <= 0:
            lines.extend([
                'Snapshot count at publish time: 0.',
                '',
                '- None',
                '',
            ])
            return

        lines.extend([
            f'Snapshot count at publish time: {snapshot_count}. '
            'This Jira table stays live after publication.',
            '',
            cls._build_release_survey_jira_macro_block(
                jql=cls._build_release_survey_jql(
                    project_key=project_key,
                    release=release,
                    scope_label=scope_label,
                    survey_mode=survey_mode,
                    family_name=family_name,
                    statuses=statuses,
                    issue_type=issue_type,
                ),
            ),
            '',
        ])

    @staticmethod
    def _format_release_survey_summary(
        report: ReleaseSurveyReport,
    ) -> str:
        '''
        Generate a markdown summary for a release execution survey.
        '''
        lines = [
            f'# Release Survey: {report.project_key}',
            '',
            f'**Survey ID**: {report.survey_id}',
            f'**Created**: {report.created_at}',
            f'**Survey Mode**: {report.survey_mode.replace("_", " ").title()}',
            f'**Releases Surveyed**: {", ".join(report.releases_surveyed)}',
        ]
        if report.scope_label:
            lines.append(f'**Scope Label**: {report.scope_label}')
        lines.extend([
            '',
            '## Overall',
            '',
            f'- Total tickets: {report.total_tickets}',
            f'- Done: {report.done_count} ({report.completion_pct:.1f}%)',
            f'- In progress: {report.in_progress_count}',
            f'- Remaining: {report.remaining_count}',
            f'- Blocked: {report.blocked_count}',
            f'- Stale: {report.stale_count}',
            f'- Unassigned: {report.unassigned_count}',
            '',
        ])

        if not report.release_summaries:
            lines.append('No releases matched this survey.')
            return '\n'.join(lines)

        for summary in report.release_summaries:
            lines.extend([
                f'## {summary.release}',
                '',
                f'- Total tickets: {summary.total_tickets}',
                f'- Done: {summary.done_count} ({summary.completion_pct:.1f}%)',
                f'- In progress: {summary.in_progress_count}',
                f'- Remaining: {summary.remaining_count}',
                f'- Blocked: {summary.blocked_count}',
                f'- Stale: {summary.stale_count}',
                f'- Unassigned: {summary.unassigned_count}',
            ])

            if summary.status_breakdown:
                status_line = ', '.join(
                    f'{status}={count}'
                    for status, count in sorted(summary.status_breakdown.items())
                )
                lines.append(f'- Status mix: {status_line}')

            if summary.priority_breakdown:
                priority_line = ', '.join(
                    f'{priority}={count}'
                    for priority, count in sorted(summary.priority_breakdown.items())
                )
                lines.append(f'- Priority mix: {priority_line}')

            lines.append('')
            GanttProjectPlannerAgent._append_release_survey_ticket_table(
                lines,
                'In Progress',
                summary.in_progress_tickets,
                limit=25,
                sort_by_priority=True,
            )
            GanttProjectPlannerAgent._append_release_survey_ticket_lines(
                lines,
                'Blocked',
                summary.blocked_tickets,
                limit=25,
            )
            GanttProjectPlannerAgent._append_release_survey_ticket_table(
                lines,
                'Remaining',
                summary.remaining_tickets,
                limit=20,
                sort_by_priority=True,
            )
            GanttProjectPlannerAgent._append_release_survey_ticket_lines(
                lines,
                'Done',
                summary.done_tickets,
                limit=20,
            )

            family_breakdowns = dict(summary.family_breakdowns or {})
            family_epic_analysis = dict(summary.family_epic_analysis or {})
            if family_breakdowns:
                lines.extend([
                    '### Product Family Breakouts',
                    '',
                    '- Tickets can appear in more than one family table when Jira lists multiple product families.',
                    '',
                ])
                for family_name, family_view in family_breakdowns.items():
                    GanttProjectPlannerAgent._append_release_survey_ticket_table(
                        lines,
                        f'{family_name} In Progress',
                        list(family_view.get('in_progress_tickets') or []),
                        limit=25,
                        sort_by_priority=True,
                        heading_level='####',
                    )
                    GanttProjectPlannerAgent._append_release_survey_ticket_table(
                        lines,
                        f'{family_name} Remaining',
                        list(family_view.get('remaining_tickets') or []),
                        limit=25,
                        sort_by_priority=True,
                        heading_level='####',
                    )
                    GanttProjectPlannerAgent._append_release_survey_ticket_table(
                        lines,
                        f'{family_name} Epics',
                        list(family_view.get('epics') or []),
                        limit=25,
                        sort_by_priority=True,
                        heading_level='####',
                    )
                    GanttProjectPlannerAgent._append_release_survey_epic_analysis(
                        lines,
                        family_name,
                        list(family_epic_analysis.get(family_name) or []),
                        limit=40,
                        heading_level='####',
                    )

        return '\n'.join(lines)

    @staticmethod
    def _append_release_survey_ticket_lines(
        lines: List[str],
        heading: str,
        tickets: List[Dict[str, Any]],
        *,
        limit: int,
        sort_by_priority: bool = False,
        heading_level: str = '###',
    ) -> None:
        '''
        Append one survey ticket bucket to the markdown summary.
        '''
        lines.append(f'{heading_level} {heading} ({len(tickets)})')
        if not tickets:
            lines.extend(['- None', ''])
            return

        ordered_tickets = list(tickets)
        if sort_by_priority:
            ordered_tickets = sorted(
                ordered_tickets,
                key=lambda item: (
                    GanttProjectPlannerAgent._release_survey_priority_rank(
                        str(item.get('priority') or '')
                    ),
                    -int(item.get('age_days') or 0),
                    str(item.get('key') or ''),
                ),
            )

        for ticket in ordered_tickets[:limit]:
            key = str(ticket.get('key') or '')
            ticket_url = str(ticket.get('url') or f'{JIRA_BASE_URL}/browse/{key}')
            key_text = f'[{key}]({ticket_url})' if key else ticket_url
            summary = str(ticket.get('summary') or '')
            status = str(ticket.get('status') or '')
            assignee = str(ticket.get('assignee') or 'Unassigned')
            priority = str(ticket.get('priority') or 'Unspecified')
            updated = str(ticket.get('updated') or '')[:10]
            lines.append(
                f'- {key_text}: {summary} '
                f'({status}; {assignee}; {priority}; updated {updated})'
            )

        if len(ordered_tickets) > limit:
            lines.append(f'- ...and {len(ordered_tickets) - limit} more')

        lines.append('')

    @classmethod
    def _append_release_survey_ticket_table(
        cls,
        lines: List[str],
        heading: str,
        tickets: List[Dict[str, Any]],
        *,
        limit: int,
        sort_by_priority: bool = False,
        heading_level: str = '###',
    ) -> None:
        '''
        Append one survey ticket bucket to the markdown summary as a table.
        '''
        lines.append(f'{heading_level} {heading} ({len(tickets)})')
        if not tickets:
            lines.extend(['- None', ''])
            return

        ordered_tickets = list(tickets)
        if sort_by_priority:
            ordered_tickets = sorted(
                ordered_tickets,
                key=lambda item: (
                    cls._release_survey_priority_rank(
                        str(item.get('priority') or '')
                    ),
                    -int(item.get('age_days') or 0),
                    str(item.get('key') or ''),
                ),
            )

        lines.extend([
            '| Ticket | Summary | Product Family | Status | Assignee | Manager | Priority | Updated |',
            '| --- | --- | --- | --- | --- | --- | --- | --- |',
        ])

        for ticket in ordered_tickets[:limit]:
            key = str(ticket.get('key') or '')
            ticket_url = str(ticket.get('url') or f'{JIRA_BASE_URL}/browse/{key}')
            key_text = f'[{key}]({ticket_url})' if key else ticket_url
            summary = cls._escape_release_survey_table_cell(ticket.get('summary'))
            product_family = cls._escape_release_survey_table_cell(
                ticket.get('product_family_csv') or 'Unspecified'
            )
            status = cls._escape_release_survey_table_cell(ticket.get('status'))
            assignee = str(ticket.get('assignee') or 'Unassigned')
            assignee_text = cls._escape_release_survey_table_cell(assignee)
            manager_name = cls._lookup_release_survey_manager_name(assignee)
            manager_text = cls._escape_release_survey_table_cell(manager_name)
            priority = cls._escape_release_survey_table_cell(
                ticket.get('priority') or 'Unspecified'
            )
            updated = cls._escape_release_survey_table_cell(
                str(ticket.get('updated') or '')[:10]
            )
            lines.append(
                f'| {key_text} | {summary} | {product_family} | {status} | '
                f'{assignee_text} | {manager_text} | {priority} | {updated} |'
            )

        if len(ordered_tickets) > limit:
            lines.append(f'- ...and {len(ordered_tickets) - limit} more')

        lines.append('')

    @classmethod
    def _append_release_survey_epic_analysis(
        cls,
        lines: List[str],
        family_name: str,
        analysis_entries: List[Dict[str, Any]],
        *,
        limit: int,
        heading_level: str = '###',
    ) -> None:
        '''
        Append open-child analysis for unfinished epics in one product family.
        '''
        lines.append(
            f'{heading_level} {family_name} Open Epic Child Analysis '
            f'({len(analysis_entries)})'
        )
        if not analysis_entries:
            lines.extend(['- None', ''])
            return

        for entry in analysis_entries:
            epic = dict(entry.get('epic') or {})
            epic_key = str(epic.get('key') or '')
            epic_url = str(epic.get('url') or f'{JIRA_BASE_URL}/browse/{epic_key}')
            epic_key_text = f'[{epic_key}]({epic_url})' if epic_key else epic_url
            epic_summary = cls._escape_release_survey_table_cell(epic.get('summary'))
            epic_status = cls._escape_release_survey_table_cell(epic.get('status'))
            lines.append(f'##### {epic_key_text}: {epic_summary}')
            lines.append(
                f'- Epic status: {epic_status or "Unknown"}; '
                f'open descendants: {int(entry.get("open_child_count") or 0)} of '
                f'{int(entry.get("total_descendant_count") or 0)} '
                f'non-bug descendants'
            )

            open_by_type = dict(entry.get('open_by_type') or {})
            if open_by_type:
                type_line = ', '.join(
                    f'{issue_type}={count}'
                    for issue_type, count in sorted(open_by_type.items())
                )
                lines.append(f'- Open type mix: {type_line}')

            error_text = str(entry.get('error') or '').strip()
            if error_text:
                lines.append(f'- Hierarchy lookup failed: {error_text}')
                lines.append('')
                continue

            open_children = list(entry.get('open_children') or [])
            if not open_children:
                lines.extend([
                    '- No open non-bug descendants found. This epic may be ready to close.',
                    '',
                ])
                continue

            lines.append('')
            lines.extend([
                '| Depth | Ticket | Type | Summary | Status | Assignee | Manager | Priority | Fix Versions |',
                '| --- | --- | --- | --- | --- | --- | --- | --- | --- |',
            ])
            for child in open_children[:limit]:
                child_key = str(child.get('key') or '')
                child_url = str(child.get('url') or f'{JIRA_BASE_URL}/browse/{child_key}')
                child_key_text = (
                    f'[{child_key}]({child_url})' if child_key else child_url
                )
                child_type = cls._escape_release_survey_table_cell(
                    child.get('issue_type') or 'Unknown'
                )
                child_summary = cls._escape_release_survey_table_cell(
                    child.get('summary') or ''
                )
                child_status = cls._escape_release_survey_table_cell(
                    child.get('status') or ''
                )
                child_assignee = str(child.get('assignee') or 'Unassigned')
                child_assignee_text = cls._escape_release_survey_table_cell(
                    child_assignee
                )
                child_manager = cls._escape_release_survey_table_cell(
                    cls._lookup_release_survey_manager_name(child_assignee)
                )
                child_priority = cls._escape_release_survey_table_cell(
                    child.get('priority') or 'Unspecified'
                )
                child_fix_versions = cls._escape_release_survey_table_cell(
                    child.get('fix_version_csv') or 'Unspecified'
                )
                lines.append(
                    f'| {int(child.get("depth") or 0)} | {child_key_text} | '
                    f'{child_type} | {child_summary} | {child_status} | '
                    f'{child_assignee_text} | {child_manager} | {child_priority} | '
                    f'{child_fix_versions} |'
                )

            if len(open_children) > limit:
                lines.append(
                    f'- ...and {len(open_children) - limit} more open descendants'
                )

            lines.append('')
